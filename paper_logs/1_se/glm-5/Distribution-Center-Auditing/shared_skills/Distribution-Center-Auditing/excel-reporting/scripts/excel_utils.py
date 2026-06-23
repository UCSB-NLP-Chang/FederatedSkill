#!/usr/bin/env python3
"""Reusable utilities for Excel reporting tasks."""

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


def read_excel_data(filepath, sheet_name=None):
    """Read all data from an Excel file, returning headers and rows.
    
    Args:
        filepath: Path to Excel file
        sheet_name: Specific sheet name (uses active sheet if None)
    
    Returns:
        Tuple of (headers, rows) where headers is a tuple of column names
        and rows is a list of tuples containing row data
    """
    wb = load_workbook(filepath, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active
    
    all_rows = list(ws.iter_rows(values_only=True))
    if not all_rows:
        return (), []
    
    headers = all_rows[0]
    rows = all_rows[1:]
    return headers, rows


def create_formatted_workbook(sheet_configs):
    """Create a workbook with multiple formatted sheets.
    
    Args:
        sheet_configs: List of dicts with keys:
            - name: Sheet name
            - headers: Tuple of column headers
            - rows: List of row tuples
            - header_style: Optional dict with font, fill, alignment options
    
    Returns:
        Workbook object ready to save
    """
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    default_header_style = {
        'font': Font(bold=True),
        'alignment': Alignment(horizontal='center'),
        'fill': PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    }
    
    for config in sheet_configs:
        ws = wb.create_sheet(config['name'])
        headers = config.get('headers', ())
        rows = config.get('rows', [])
        style = config.get('header_style', default_header_style)
        
        # Write headers
        if headers:
            ws.append(headers)
            for cell in ws[1]:
                if style.get('font'):
                    cell.font = style['font']
                if style.get('alignment'):
                    cell.alignment = style['alignment']
                if style.get('fill'):
                    cell.fill = style['fill']
        
        # Write data rows
        for row in rows:
            ws.append(row)
    
    return wb


def add_calculated_column(rows, headers, column_name, calc_func):
    """Add a calculated column to data rows.
    
    Args:
        rows: List of row tuples
        headers: Tuple of column headers
        column_name: Name for the new column
        calc_func: Function that takes a row dict and returns calculated value
    
    Returns:
        Tuple of (new_headers, new_rows)
    """
    new_headers = headers + (column_name,)
    header_dict = {h: i for i, h in enumerate(headers)}
    
    new_rows = []
    for row in rows:
        row_dict = {h: row[i] for i, h in enumerate(headers)}
        calc_value = calc_func(row_dict)
        new_rows.append(row + (calc_value,))
    
    return new_headers, new_rows


def verify_workbook(filepath):
    """Verify workbook contents by reading it back.
    
    Args:
        filepath: Path to Excel file to verify
    
    Returns:
        Dict with sheet names and row counts
    """
    wb = load_workbook(filepath)
    result = {'sheets': {}}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        row_count = ws.max_row
        col_count = ws.max_column
        result['sheets'][sheet_name] = {
            'rows': row_count,
            'columns': col_count
        }
    return result