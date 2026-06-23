#!/usr/bin/env python3
"""
Audit Template: Validates PDF claims against Excel/CSV/JSON reference data.
Adapt field names, regex extraction patterns, and validation rules to the specific task.
Run via: python3 audit_template.py
"""
import json
import csv
import re
import subprocess
from difflib import SequenceMatcher

# --- Configuration (Update paths per task) ---
PDF_PATH = "claims.pdf"
PROVIDER_XLSX = "directory.xlsx"
ORDERS_JSON = "orders.json"
ADJUSTMENTS_CSV = "adjustments.csv"
OUTPUT_PATH = "flags.json"

def load_providers(xlsx_path):
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    providers = {}
    aliases = {}
    # Load main sheet
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        pid, name, region, acct = row[0], row[1], row[2], row[3]
        providers[pid] = {"name": name, "region": region, "account": acct}
        aliases[name.lower().strip()] = pid
    # Load alias sheet if present
    for sheet_name in wb.sheetnames:
        if "alias" in sheet_name.lower():
            ws_alias = wb[sheet_name]
            for row in ws_alias.iter_rows(min_row=2, values_only=True):
                pid, alias_name = row[0], row[1]
                aliases[alias_name.lower().strip()] = pid
    return providers, aliases

def flatten_json(data, keys_to_extract):
    """Recursively flatten nested JSON to find target keys."""
    results = []
    if isinstance(data, dict):
        if all(k in data for k in keys_to_extract):
            results.append(data)
        for v in data.values():
            results.extend(flatten_json(v, keys_to_extract))
    elif isinstance(data, list):
        for item in data:
            results.extend(flatten_json(item, keys_to_extract))
    return results

def load_orders(json_path):
    with open(json_path, "r") as f:
        data = json.load(f)
    # Flatten nested structure (adapt keys_to_extract to actual schema)
    flat_records = flatten_json(data, ["order_id", "provider_id", "approved_charge", "lifecycle"])
    orders = {}
    for rec in flat_records:
        orders[rec["order_id"]] = {
            "provider_id": rec["provider_id"],
            "approved_charge": float(rec["approved_charge"]),
            "lifecycle": rec.get("lifecycle", "approved")
        }
    return orders

def load_adjustments(csv_path):
    adjustments = {}
    with open(csv_path, "r") as f:
        reader = csv.DictReader(f)
        for row in reader:
            if row.get("decision", "").lower() == "approved" or row.get("approval_state", "").lower() == "approved":
                oid = row.get("order_id") or row.get("award_ref")
                if not oid: continue
                # Keep highest amendment/version number
                ver_key = "amendment_no" if "amendment_no" in row else "version_no"
                amt_key = "amended_charge" if "amended_charge" in row else "version_amount"
                current_ver = int(row.get(ver_key, 0))
                new_amt = row.get(amt_key, "").strip()
                
                if oid not in adjustments or current_ver > int(adjustments[oid]["version"]):
                    # Handle empty/null override amounts by falling back to base or storing None
                    adj_amt = float(new_amt) if new_amt else None
                    adjustments[oid] = {
                        "version": current_ver,
                        "amended_charge": adj_amt
                    }
    return adjustments

def extract_pdf_text(pdf_path):
    # Try pdfplumber first, fallback to pdftotext
    try:
        import pdfplumber
        with pdfplumber.open(pdf_path) as pdf:
            return [page.extract_text() or "" for page in pdf.pages]
    except ImportError:
        result = subprocess.run(["pdftotext", "-layout", pdf_path, "-"], capture_output=True, text=True)
        return result.stdout.split("\f")  # Split by form feed

def match_name(claimed_name, aliases):
    claimed_lower = claimed_name.lower().strip()
    if claimed_lower in aliases:
        return aliases[claimed_lower]
    # Fuzzy fallback
    best_match = None
    best_score = 0.0
    for alias_name, pid in aliases.items():
        score = SequenceMatcher(None, claimed_lower, alias_name).ratio()
        if score > best_score:
            best_score = score
            best_match = pid
    return best_match if best_score > 0.80 else None

def parse_page(text):
    """Extract fields from a single page. Adapt regex to match PDF layout."""
    provider = re.search(r"Provider:\s*(.+)", text)
    order_id = re.search(r"Order ID:\s*(.+)", text)
    amount = re.search(r"Total:\s*\$?([\d,]+\.?\d*)", text)
    account = re.search(r"Account:\s*(.+)", text)
    revision = re.search(r"Revision:\s*(\d+)", text)
    
    return {
        "provider": provider.group(1).strip() if provider else None,
        "order_id": order_id.group(1).strip() if order_id else None,
        "amount": float(amount.group(1).replace(",", "")) if amount else 0.0,
        "account": account.group(1).strip() if account else None,
        "revision": int(revision.group(1)) if revision else 1
    }

def deduplicate_claims(parsed_claims):
    """Keep only the highest revision per claim ID."""
    latest = {}
    for claim in parsed_claims:
        oid = claim.get("order_id")
        if oid:
            if oid not in latest or claim["revision"] > latest[oid]["revision"]:
                latest[oid] = claim
    return list(latest.values())

def validate_claims(pages_text, providers, aliases, orders, adjustments):
    # Parse all pages first
    all_claims = []
    for i, text in enumerate(pages_text, start=1):
        fields = parse_page(text)
        if not all(v for k, v in fields.items() if k != "revision"):
            continue  # Skip malformed pages
        fields["page_number"] = i
        all_claims.append(fields)
        
    # Deduplicate by claim ID, keeping highest revision
    claims = deduplicate_claims(all_claims)
    
    flags = []
    for claim in claims:
        provider, order_id, amount, account, page = claim["provider"], claim["order_id"], claim["amount"], claim["account"], claim["page_number"]
        
        # Validation Chain
        pid = match_name(provider, aliases)
        if not pid:
            flags.append({"request_page_number": page, "participant_name": provider, "reason": "Unknown Participant"})
            continue
            
        if order_id not in orders:
            flags.append({"request_page_number": page, "participant_name": provider, "award_ref": order_id, "reason": "Invalid Award Ref"})
            continue
            
        order = orders[order_id]
        if order["provider_id"] != pid:
            flags.append({"request_page_number": page, "participant_name": provider, "award_ref": order_id, "reason": "Participant Mismatch"})
            continue
            
        expected_amount = order["approved_charge"]
        if order_id in adjustments and adjustments[order_id]["amended_charge"] is not None:
            expected_amount = adjustments[order_id]["amended_charge"]
            
        if abs(amount - expected_amount) > 0.001:
            flags.append({"request_page_number": page, "participant_name": provider, "requested_amount": amount, "award_ref": order_id, "reason": "Amount Mismatch"})
            continue
            
        if account != providers[pid]["account"]:
            flags.append({"request_page_number": page, "participant_name": provider, "payment_account": account, "award_ref": order_id, "reason": "Account Mismatch"})
            continue
            
    return flags

if __name__ == "__main__":
    providers, aliases = load_providers(PROVIDER_XLSX)
    orders = load_orders(ORDERS_JSON)
    adjustments = load_adjustments(ADJUSTMENTS_CSV)
    pages = extract_pdf_text(PDF_PATH)
    flags = validate_claims(pages, providers, aliases, orders, adjustments)
    with open(OUTPUT_PATH, "w") as f:
        json.dump(flags, f, indent=2)
    print(f"Generated {len(flags)} flags -> {OUTPUT_PATH}")
