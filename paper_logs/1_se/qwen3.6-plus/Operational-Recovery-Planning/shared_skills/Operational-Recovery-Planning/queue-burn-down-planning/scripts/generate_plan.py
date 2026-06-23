#!/usr/bin/env python3
"""Generates a deterministic queue burn-down plan and summary from reference Excel data."""
import sys
import argparse
import openpyxl
from openpyxl import Workbook

def main():
    parser = argparse.ArgumentParser(description="Generate queue burn-down plan")
    parser.add_argument("source", help="Path to source Excel file")
    parser.add_argument("plan", help="Path to output plan Excel file")
    parser.add_argument("summary", help="Path to output summary text file")
    parser.add_argument("--backlog", type=float, required=True, help="Initial backlog value")
    parser.add_argument("--cap6", type=float, default=168, help="Capacity for 6-day phase")
    parser.add_argument("--cap5", type=float, default=140, help="Capacity for 5-day phase")
    parser.add_argument("--cap4", type=float, default=112, help="Capacity for 4-day phase")
    parser.add_argument("--ot6", type=float, default=16, help="OT for 6-day phase")
    parser.add_argument("--ot5", type=float, default=8, help="OT for 5-day phase")
    parser.add_argument("--ot4", type=float, default=0, help="OT for 4-day phase")
    parser.add_argument("--threshold", type=float, default=112, help="Demand threshold for phase transition")
    parser.add_argument("--demand_labels", type=str, default=None, help="Comma-separated row labels for demand. Values will be summed.")
    parser.add_argument("--demand_label", type=str, default="Security Alert Load Total", help="Single row label for demand (legacy).")
    parser.add_argument("--headers", type=str, default=None, help="Comma-separated exact column headers for output Excel")
    parser.add_argument("--summary_line3", type=str, default=None, help="Exact text for the 3rd summary line")
    args = parser.parse_args()

    wb_src = openpyxl.load_workbook(args.source, data_only=True)
    ws_src = wb_src.active
    
    # Determine demand labels
    if args.demand_labels:
        demand_labels = [l.strip() for l in args.demand_labels.split(",")]
    else:
        demand_labels = [args.demand_label]
        
    weeks = []
    demands = []
    
    for row in ws_src.iter_rows(min_row=1, max_row=ws_src.max_row, values_only=True):
        if row[0] == "Week":
            weeks = [v for v in row[1:] if v is not None and isinstance(v, (int, float))]
        elif row[0] in demand_labels:
            vals = [v for v in row[1:] if v is not None and isinstance(v, (int, float))]
            if not demands:
                demands = [0.0] * len(vals)
            for i, v in enumerate(vals):
                demands[i] += v
            
    if not weeks or not demands:
        print("Error: Could not find 'Week' or demand data.")
        sys.exit(1)
        
    n_weeks = min(len(weeks), len(demands))
    weeks = weeks[:n_weeks]
    demands = demands[:n_weeks]
    
    sow = args.backlog
    first_5 = None
    first_4 = None
    plan_rows = []
    
    for i in range(n_weeks):
        week = weeks[i]
        demand = demands[i]
        
        if sow > 0:
            days, cap, ot = 6, args.cap6, args.ot6
        elif demand > args.threshold:
            days, cap, ot = 5, args.cap5, args.ot5
        else:
            days, cap, ot = 4, args.cap4, args.ot4
            
        eow = sow + demand - cap
        
        if first_5 is None and days == 5:
            first_5 = week
        if first_4 is None and days == 4:
            first_4 = week
            
        plan_rows.append([week, days, demand, cap, sow, eow, ot])
        sow = max(0, eow)
        
    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = "Plan"
    
    if args.headers:
        headers = [h.strip() for h in args.headers.split(",")]
    else:
        headers = ["Week", "On-Call Days", "Forecast Demand", "Weekly Capacity", "Start-of-Week Queue", "End-of-Week Queue/Buffer", "Overtime Hours"]
        
    ws_out.append(headers)
    for row in plan_rows:
        ws_out.append(row)
    wb_out.save(args.plan)
    
    summary_line3 = args.summary_line3 or f"The queue cleared to normal operations by Week {first_5} (first 5-day week). Step-down to minimum staffing occurred at Week {first_4} (first 4-day week). Backlog was fully resolved and demand stabilized within standard capacity thresholds."
    
    with open(args.summary, "w") as f:
        f.write(f"First_Week_5_Days: {first_5}\n")
        f.write(f"First_Week_4_Days: {first_4}\n")
        f.write(f"Summary: {summary_line3}\n")
        
    print(f"Generated {args.plan} and {args.summary}")

if __name__ == "__main__":
    main()