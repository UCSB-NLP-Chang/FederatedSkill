#!/usr/bin/env python3
"""Calculate route dispatch tracker from template, stock, and queue workbooks.
Preserves template sheets, parses grouped stock, resolves aliases,
deduplicates dispatches by revision, filters by state, and computes
per-route/SKU load requirements with route/SKU-specific configs.
"""
import sys
import math
import re
from datetime import datetime, timedelta
from openpyxl import Workbook, load_workbook

VALID_STATES = {"released", "approved", "committed", "confirmed", "booked", "loaded", "firm", "arranged"}
IGNORE_STATES = {"pending", "tentative", "cancelled", "rejected", "draft", "hold"}

def parse_date(val):
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, str):
        try:
            return datetime.strptime(val.strip(), "%Y-%m-%d").date()
        except ValueError:
            return None
    return None

def find_sheet(wb, keywords):
    for name in wb.sheetnames:
        if any(k.lower() in name.lower() for k in keywords):
            return wb[name]
    return None

def find_label_value(ws, label, max_scan=15):
    for r in range(1, max_scan + 1):
        for c in range(1, 6):
            cell = ws.cell(row=r, column=c)
            if cell.value and str(cell.value).strip().lower() == label.lower():
                if c < 5 and ws.cell(row=r, column=c+1).value:
                    return ws.cell(row=r, column=c+1).value
                if r < max_scan and ws.cell(row=r+1, column=c).value:
                    return ws.cell(row=r+1, column=c).value
    return None

def is_section_header(val):
    if not val or not isinstance(val, str):
        return False
    val = val.strip()
    return bool(re.match(r'^(Route|Lane|Zone|Aisle|Section|Category|Group)[:\s]', val, re.IGNORECASE))

def extract_route_name(header_val):
    if not header_val:
        return None
    val = str(header_val).strip()
    match = re.match(r'^(?:Route|Lane|Zone|Aisle|Section|Category|Group)[:\s]+(.+)$', val, re.IGNORECASE)
    return match.group(1).strip() if match else val

def map_headers(ws, targets, max_row=5):
    for row in ws.iter_rows(min_row=1, max_row=max_row, values_only=False):
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                hdr = cell.value.strip().lower()
                for t in targets:
                    if t in hdr:
                        targets[t] = cell.column
        if all(v is not None for v in targets.values()):
            break
    return targets

def main():
    if len(sys.argv) != 5:
        print("Usage: python3 calculate_route_tracker.py <template.xlsx> <stock.xlsx> <queue.xlsx> <output.xlsx>")
        sys.exit(1)
    
    tmpl_path, stock_path, queue_path, out_path = sys.argv[1], sys.argv[2], sys.argv[3], sys.argv[4]
    
    tmpl_wb = load_workbook(tmpl_path)
    sheet_order = tmpl_wb.sheetnames
    
    stock_wb = load_workbook(stock_path, data_only=True)
    queue_wb = load_workbook(queue_path, data_only=True)
    
    stock_ws = find_sheet(stock_wb, ["stock", "snapshot", "current", "inventory"])
    queue_ws = find_sheet(queue_wb, ["queue", "dispatch", "log", "shipment", "booking"])
    alias_ws = find_sheet(tmpl_wb, ["alias", "route map", "location map"])
    pack_ws = find_sheet(tmpl_wb, ["pack matrix", "pallet guide", "config", "setup"])
    
    if not stock_ws or not queue_ws:
        print("Error: Could not locate Stock or Dispatch Queue sheets.")
        sys.exit(1)
    
    asof = parse_date(find_label_value(stock_ws, "asofdate"))
    horizon = parse_date(find_label_value(stock_ws, "horizonend"))
    planning_days = (horizon - asof).days if horizon and asof else 30
    
    # Parse Route Alias Map
    alias_map = {}
    if alias_ws:
        for row in alias_ws.iter_rows(min_row=2, values_only=True):
            if row and row[0] and row[1]:
                alias_map[str(row[0]).strip()] = str(row[1]).strip()
    
    # Parse Pack Matrix (Route/SKU -> Cases Per Load)
    pack_config = {}
    if pack_ws:
        for row in pack_ws.iter_rows(min_row=2, values_only=True):
            if row and row[0] and row[1] and row[2]:
                try:
                    route = str(row[0]).strip()
                    sku = str(row[1]).strip()
                    cases = int(row[2])
                    pack_config[(route, sku)] = cases
                except (ValueError, TypeError):
                    pass
    
    # Parse Grouped Stock
    stock_hdrs = map_headers(stock_ws, {"sku": None, "on_hand": None, "daily": None})
    stock = {}
    current_route = None
    for row in stock_ws.iter_rows(min_row=1, values_only=True):
        if not row or not row[0]:
            continue
        if is_section_header(row[0]):
            current_route = extract_route_name(row[0])
            continue
        sku = row[stock_hdrs["sku"]-1] if stock_hdrs["sku"] else None
        if not sku or not isinstance(sku, str):
            continue
        try:
            on_hand = float(row[stock_hdrs["on_hand"]-1] if stock_hdrs["on_hand"] else 0)
            daily = float(row[stock_hdrs["daily"]-1] if stock_hdrs["daily"] else 0)
            key = (current_route, sku) if current_route else sku
            stock[key] = {"route": current_route, "sku": sku, "on_hand": on_hand, "daily": daily}
        except (ValueError, TypeError):
            continue
    
    # Parse Dispatch Queue
    queue_hdrs = map_headers(queue_ws, {"row_type": None, "queue_id": None, "revision": None, "route_alias": None, "sku": None, "date": None, "cases": None, "state": None})
    loads_by_id = {}
    for row in queue_ws.iter_rows(min_row=1, values_only=True):
        if not row or not row[0]:
            continue
        row_type = str(row[queue_hdrs["row_type"]-1] if queue_hdrs["row_type"] else "").strip().upper()
        if row_type not in ("DISPATCH", "SHIPMENT", "TRANSFER"):
            continue
            
        queue_id = row[queue_hdrs["queue_id"]-1] if queue_hdrs["queue_id"] else None
        if not queue_id:
            continue
        try:
            revision = int(row[queue_hdrs["revision"]-1] if queue_hdrs["revision"] else 0)
        except (ValueError, TypeError):
            revision = 0
            
        if queue_id not in loads_by_id or revision > loads_by_id[queue_id]["revision"]:
            alias = row[queue_hdrs["route_alias"]-1] if queue_hdrs["route_alias"] else None
            route = alias_map.get(str(alias).strip()) if alias else None
            sku = row[queue_hdrs["sku"]-1] if queue_hdrs["sku"] else None
            if not sku or not isinstance(sku, str):
                continue
            if not route:
                continue
                
            loads_by_id[queue_id] = {
                "queue_id": queue_id,
                "revision": revision,
                "route": route,
                "sku": str(sku).strip(),
                "date": parse_date(row[queue_hdrs["date"]-1] if queue_hdrs["date"] else None),
                "cases": float(row[queue_hdrs["cases"]-1] if queue_hdrs["cases"] else 0),
                "state": str(row[queue_hdrs["state"]-1] if queue_hdrs["state"] else "").strip().lower()
            }
    
    # Filter by state and horizon
    qualifying_dispatches = []
    for load in loads_by_id.values():
        if load["state"] in IGNORE_STATES or load["state"] not in VALID_STATES:
            continue
        if load["date"] and load["date"] > horizon:
            continue
        qualifying_dispatches.append(load)
    
    # Compute per-route/SKU metrics
    results = []
    for key, data in stock.items():
        route, sku = data["route"], data["sku"]
        on_hand, daily = data["on_hand"], data["daily"]
        days_oh = float('inf') if daily == 0 else on_hand / daily
        oos_date = horizon + timedelta(days=1) if daily == 0 else asof + timedelta(days=days_oh)
        
        sku_dispatches = [d for d in qualifying_dispatches if d["route"] == route and d["sku"] == sku]
        inbound_cases = sum(d["cases"] for d in sku_dispatches)
        earliest_dispatch = min((d["date"] for d in sku_dispatches if d["date"]), default=None)
        
        delivered_doh = (on_hand + inbound_cases) / daily if daily > 0 else 0
        remaining_demand = daily * planning_days
        additional = max(0, remaining_demand - (on_hand + inbound_cases))
        cases_per_load = pack_config.get((route, sku), 40)
        loads_req = math.ceil(additional / cases_per_load) if cases_per_load > 0 else 0
        
        earlier = True
        if inbound_cases > 0 and earliest_dispatch and earliest_dispatch <= oos_date:
            earlier = False
        
        results.append({
            "route": route, "sku": sku, "on_hand": on_hand, "daily": daily,
            "days_oh": round(days_oh, 2) if days_oh != float('inf') else "N/A",
            "oos_date": oos_date, "inbound": inbound_cases,
            "delivered_doh": round(delivered_doh, 2), "demand": remaining_demand,
            "additional": additional, "loads_req": loads_req,
            "rdd": oos_date, "earlier": earlier
        })
    
    # Build output workbook preserving template structure
    out_wb = Workbook()
    for sname in sheet_order:
        src_ws = tmpl_wb[sname]
        if sname.lower() in ["overview", "pack matrix", "route alias map", "alias map", "instructions"]:
            dst_ws = out_wb.create_sheet(sname)
            for row in src_ws.iter_rows(values_only=True):
                dst_ws.append(list(row))
    
    # Write Coverage_Detail
    cd_ws = out_wb.create_sheet("Coverage_Detail")
    cd_ws.append(["Field", "Value"])
    cd_ws.append(["AsOfDate", asof.isoformat() if asof else ""])
    cd_ws.append(["HorizonEnd", horizon.isoformat() if horizon else ""])
    cd_ws.append(["PlanningDays", planning_days])
    cd_ws.append([])
    
    headers = ["Route", "SKU", "On_Hand_Cases", "Daily_Demand_Cases_Per_Day", "Current_Days_On_Hand",
               "Projected_OOS_Date", "Inbound_Cases_By_Horizon", "Delivered_Days_On_Hand",
               "Remaining_Demand_Cases", "Additional_Cases_Needed", "Loads_Required",
               "Required_Delivery_Date", "Earlier_Delivery_Required"]
    cd_ws.append(headers)
    
    for r in results:
        cd_ws.append([
            r["route"], r["sku"], r["on_hand"], r["daily"], r["days_oh"],
            r["oos_date"].isoformat() if hasattr(r["oos_date"], 'isoformat') else r["oos_date"],
            r["inbound"], r["delivered_doh"], r["demand"],
            r["additional"], r["loads_req"],
            r["rdd"].isoformat() if hasattr(r["rdd"], 'isoformat') else r["rdd"],
            r["earlier"]
        ])
    
    # Write Dispatch_Plan
    dp_ws = out_wb.create_sheet("Dispatch_Plan")
    dp_headers = ["Route", "SKU", "Required_Delivery_Date", "Loads_Required", "Additional_Cases_Needed", "Earlier_Delivery_Required"]
    dp_ws.append(dp_headers)
    for r in results:
        if r["loads_req"] > 0:
            dp_ws.append([
                r["route"], r["sku"],
                r["rdd"].isoformat() if hasattr(r["rdd"], 'isoformat') else r["rdd"],
                r["loads_req"], r["additional"], r["earlier"]
            ])
    
    out_wb.save(out_path)
    print(f"Saved route tracker to {out_path}")

if __name__ == "__main__":
    main()
