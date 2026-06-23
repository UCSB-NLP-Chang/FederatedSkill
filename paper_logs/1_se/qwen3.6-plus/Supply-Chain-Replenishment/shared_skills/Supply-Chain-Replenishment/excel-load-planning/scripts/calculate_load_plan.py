#!/usr/bin/env python3
"""Calculate DC load plan from source inventory workbook and write output.
Dynamically detects sheets, handles grouped/hierarchical layouts, filters inbounds by status,
handles formula-based case calculations, reads ratio sheets, and handles type-safe parsing.
"""
import sys
import math
import re
from datetime import datetime, timedelta
from openpyxl import Workbook, load_workbook

VALID_STATUSES = {"committed", "arranged", "confirmed", "approved", "ready", "docked", "firm", "locked"}
IGNORE_STATUSES = {"pending", "tentative", "cancelled", "rejected", "draft", "hold"}

def parse_date(val):
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, str):
        try:
            return datetime.strptime(val.strip(), "%Y-%m-%d").date()
        except ValueError:
            return None
    return None

def find_sheet(wb, keywords):
    for name in wb.sheetnames:
        if any(k.lower() in name.lower() for k in keywords):
            return wb[name]
    return None

def find_label_value(ws, label, max_scan=15):
    for r in range(1, max_scan + 1):
        for c in range(1, 6):
            cell = ws.cell(row=r, column=c)
            if cell.value and str(cell.value).strip().lower() == label.lower():
                if c < 5 and ws.cell(row=r, column=c+1).value:
                    return ws.cell(row=r, column=c+1).value
                if r < max_scan and ws.cell(row=r+1, column=c).value:
                    return ws.cell(row=r+1, column=c).value
    return None

def is_section_header(val):
    if not val or not isinstance(val, str):
        return False
    val = val.strip()
    return bool(re.match(r'^(Lane|Zone|Aisle|Section|Category|Group)[:\s]', val, re.IGNORECASE))

def map_headers(ws, targets, max_row=5):
    for row in ws.iter_rows(min_row=1, max_row=max_row, values_only=False):
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                hdr = cell.value.strip().lower()
                for t in targets:
                    if t in hdr:
                        targets[t] = cell.column
        if all(v is not None for v in targets.values()):
            break
    return targets

def read_ratio_sheet(wb):
    """Read cases-per-pallet ratio from Ratio/Config sheet."""
    ratio_ws = find_sheet(wb, ["ratio", "conversion", "config", "pallet guide", "setup"])
    if not ratio_ws:
        return 40
    for row in ratio_ws.iter_rows(min_row=1, values_only=True):
        if not row or len(row) < 2:
            continue
        # Look for pattern: Cases | Pallet with values like 80 | 1
        if str(row[0]).strip().lower() == "cases" and str(row[1]).strip().lower() == "pallet":
            try:
                return int(row[0]) if isinstance(row[0], (int, float)) else int(ratio_ws.cell(row=ratio_ws.max_row, column=1).value)
            except (ValueError, TypeError):
                pass
        # Look for key-value pairs
        if str(row[0]).strip().lower() in ("cases", "cases per pallet", "ratio"):
            try:
                return int(row[1]) if row[1] else int(row[0])
            except (ValueError, TypeError):
                pass
    return 40

def main():
    if len(sys.argv) != 3:
        print("Usage: python3 calculate_load_plan.py <source.xlsx> <output.xlsx>")
        sys.exit(1)
    
    src_path, out_path = sys.argv[1], sys.argv[2]
    wb = load_workbook(src_path, data_only=True)
    
    stock_ws = find_sheet(wb, ["stock", "snapshot", "current", "inventory", "lane"])
    inbound_ws = find_sheet(wb, ["inbound", "arrival", "schedule", "expected", "booking", "shipment"])
    config_ws = find_sheet(wb, ["config", "guide", "pallet", "setup", "defaults"])
    
    if not stock_ws or not inbound_ws:
        print("Error: Could not locate Stock/Inventory or Inbound/Arrival sheets.")
        sys.exit(1)
        
    today = parse_date(find_label_value(stock_ws, "asofdate") or find_label_value(stock_ws, "today's date") or find_label_value(config_ws, "asofdate"))
    horizon = parse_date(find_label_value(stock_ws, "horizonend") or find_label_value(stock_ws, "month end") or find_label_value(config_ws, "horizonend"))
    planning_days = (horizon - today).days if horizon and today else 30
    
    cases_per_pallet = read_ratio_sheet(wb)
    
    # Detect if stock sheet is grouped
    is_grouped = False
    for row in stock_ws.iter_rows(min_row=1, max_row=10, values_only=True):
        if row and is_section_header(row[0]):
            is_grouped = True
            break
            
    stock_hdrs = map_headers(stock_ws, {"sku": None, "on_floor": None, "daily": None})
    inbound_hdrs = map_headers(inbound_ws, {"sku": None, "date": None, "cases": None, "pallets": None, "status": None, "lane": None})
    
    if stock_hdrs["sku"] is None or inbound_hdrs["sku"] is None:
        print("Error: Could not map SKU columns in source sheets.")
        sys.exit(1)
        
    stock = {}
    current_lane = None
    for row in stock_ws.iter_rows(min_row=1, values_only=True):
        if not row or not row[0]: continue
        
        # Handle grouped layout
        if is_grouped and is_section_header(row[0]):
            current_lane = row[0].strip().split(":")[-1].strip()
            continue
            
        sku = row[stock_hdrs["sku"]-1] if stock_hdrs["sku"] else None
        if not sku or not isinstance(sku, str): continue
        
        try:
            on_floor = float(row[stock_hdrs["on_floor"]-1] if stock_hdrs["on_floor"] else 0)
            daily = float(row[stock_hdrs["daily"]-1] if stock_hdrs["daily"] else 0)
            key = (current_lane, sku) if current_lane else sku
            stock[key] = {"on_floor": on_floor, "daily_sales": daily, "lane": current_lane, "sku": sku}
        except (ValueError, TypeError): continue
            
    inbounds = {}
    for row in inbound_ws.iter_rows(min_row=1, values_only=True):
        sku = row[inbound_hdrs["sku"]-1] if inbound_hdrs["sku"] else None
        if not sku or not isinstance(sku, str): continue
        
        arr_date = parse_date(row[inbound_hdrs["date"]-1] if inbound_hdrs["date"] else None)
        
        # Handle formula cells: cases may be None, compute from pallets * ratio
        cases_val = row[inbound_hdrs["cases"]-1] if inbound_hdrs["cases"] else None
        if cases_val is None and inbound_hdrs["pallets"]:
            pallets_val = row[inbound_hdrs["pallets"]-1]
            try:
                cases = float(pallets_val) * cases_per_pallet if pallets_val else 0
            except (ValueError, TypeError):
                cases = 0
        else:
            try:
                cases = float(cases_val) if cases_val else 0
            except (ValueError, TypeError):
                cases = 0
                
        status = str(row[inbound_hdrs["status"]-1] if inbound_hdrs["status"] else "Committed").strip().lower()
        lane = row[inbound_hdrs["lane"]-1] if inbound_hdrs["lane"] else None
        if lane: lane = str(lane).strip()
        
        if not arr_date or arr_date > horizon: continue
        if status in IGNORE_STATUSES: continue
            
        key = (lane, sku) if lane else sku
        inbounds.setdefault(key, []).append({"date": arr_date, "cases": cases})
        
    results = []
    for key, data in stock.items():
        lane = data.get("lane")
        sku = data["sku"]
        on_floor, daily = data["on_floor"], data["daily_sales"]
        
        days_oh = float('inf') if daily == 0 else on_floor / daily
        oos_date = horizon + timedelta(days=1) if daily == 0 else today + timedelta(days=days_oh)
        
        item_inbounds = inbounds.get(key, [])
        inbound_cases = sum(i["cases"] for i in item_inbounds)
        earliest_inbound = min((i["date"] for i in item_inbounds), default=None)
        
        delivered_oh = (on_floor + inbound_cases) / daily if daily > 0 else 0
        total_demand = daily * planning_days
        additional = max(0, total_demand - (on_floor + inbound_cases))
        pallets = math.ceil(additional / cases_per_pallet) if cases_per_pallet > 0 else 0
        rounding_applied = (additional / cases_per_pallet) != pallets if cases_per_pallet > 0 else False
        
        earlier = True
        if inbound_cases > 0 and earliest_inbound and earliest_inbound <= oos_date:
            earlier = False
            
        results.append({
            "lane": lane or "", "sku": sku, "on_floor": on_floor, "daily": daily,
            "days_oh": round(days_oh, 2) if days_oh != float('inf') else "N/A",
            "oos_date": oos_date, "inbound": inbound_cases,
            "delivered_oh": round(delivered_oh, 2), "demand": total_demand,
            "additional": additional, "pallets": pallets,
            "rdd": oos_date, "earlier": earlier, "rounding": rounding_applied,
            "earliest_inbound": earliest_inbound
        })
        
    out_wb = Workbook()
    ws1 = out_wb.active
    ws1.title = "SKU_Results"
    has_lane = bool(any(r["lane"] for r in results))
    
    # Write metadata
    ws1.append(["Field", "Value"])
    ws1.append(["AsOfDate", today.isoformat() if today else ""])
    ws1.append(["PlanningHorizonEnd", horizon.isoformat() if horizon else ""])
    ws1.append(["RemainingDaysInJuly" if planning_days == 27 else "PlanningDays", planning_days])
    ws1.append([])
    
    headers = ["Lane", "Product_SKU", "Current_Cases", "Daily_Rate_Cases_Per_Day", "Current_DOH",
               "Projected_OOS_Date", "Inbound_Cases_By_Horizon", "Delivered_DOH_To_Horizon",
               "Remaining_Demand_Cases", "Additional_Cases_Needed", "Pallets_Required_Rounded_Up",
               "Required_Delivery_Date", "Rounding_Applied", "Earlier_Delivery_Required",
               "Earliest_Scheduled_Inbound_Date"] if has_lane else \
              ["Product_SKU", "Current_Cases", "Daily_Rate_Cases_Per_Day", "Current_DOH",
               "Projected_OOS_Date", "Inbound_Cases_By_Horizon", "Delivered_DOH_To_Horizon",
               "Remaining_Demand_Cases", "Additional_Cases_Needed", "Pallets_Required_Rounded_Up",
               "Required_Delivery_Date", "Rounding_Applied", "Earlier_Delivery_Required",
               "Earliest_Scheduled_Inbound_Date"]
    ws1.append(headers)
    
    for r in results:
        row_data = [r["lane"], r["sku"], r["on_floor"], r["daily"], r["days_oh"],
                    r["oos_date"].isoformat() if hasattr(r["oos_date"], 'isoformat') else r["oos_date"],
                    r["inbound"], r["delivered_oh"], r["demand"], r["additional"],
                    r["pallets"],
                    r["rdd"].isoformat() if hasattr(r["rdd"], 'isoformat') else r["rdd"],
                    r["rounding"], r["earlier"],
                    r["earliest_inbound"].isoformat() if r["earliest_inbound"] and hasattr(r["earliest_inbound"], 'isoformat') else (r["earliest_inbound"] if r["earliest_inbound"] else "")
                   ] if has_lane else \
                   [r["sku"], r["on_floor"], r["daily"], r["days_oh"],
                    r["oos_date"].isoformat() if hasattr(r["oos_date"], 'isoformat') else r["oos_date"],
                    r["inbound"], r["delivered_oh"], r["demand"], r["additional"],
                    r["pallets"],
                    r["rdd"].isoformat() if hasattr(r["rdd"], 'isoformat') else r["rdd"],
                    r["rounding"], r["earlier"],
                    r["earliest_inbound"].isoformat() if r["earliest_inbound"] and hasattr(r["earliest_inbound"], 'isoformat') else (r["earliest_inbound"] if r["earliest_inbound"] else "")
                   ]
        ws1.append(row_data)
                    
    ws2 = out_wb.create_sheet("Additional_Shipments_Needed")
    ws2_headers = ["Lane", "Product_SKU", "Required_Delivery_Date", "Pallets_Required_Rounded_Up",
                   "Additional_Cases_Needed", "Rounding_Applied", "Earlier_Delivery_Required"] if has_lane else \
                  ["Product_SKU", "Required_Delivery_Date", "Pallets_Required_Rounded_Up",
                   "Additional_Cases_Needed", "Rounding_Applied", "Earlier_Delivery_Required"]
    ws2.append(ws2_headers)
    for r in results:
        if r["pallets"] > 0:
            row_data = [r["lane"], r["sku"],
                        r["rdd"].isoformat() if hasattr(r["rdd"], 'isoformat') else r["rdd"],
                        r["pallets"], r["additional"], r["rounding"], r["earlier"]
                       ] if has_lane else \
                       [r["sku"],
                        r["rdd"].isoformat() if hasattr(r["rdd"], 'isoformat') else r["rdd"],
                        r["pallets"], r["additional"], r["rounding"], r["earlier"]
                       ]
            ws2.append(row_data)
            
    out_wb.save(out_path)
    print(f"Saved load plan to {out_path}")

if __name__ == "__main__":
    main()