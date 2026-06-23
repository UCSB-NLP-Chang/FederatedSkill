#!/usr/bin/env python3
"""Calculate recovery tracker from template, stock, and recovery log workbooks.
Preserves template sheets, deduplicates loads by revision, filters by stage,
and computes per-SKU load requirements with SKU-specific pallet configs.
"""
import sys
import math
from datetime import datetime, timedelta
from openpyxl import Workbook, load_workbook

VALID_STAGES = {"booked", "loaded", "confirmed", "committed", "firm"}
IGNORE_STAGES = {"tentative", "pending", "cancelled", "rejected", "draft", "hold"}

def parse_date(val):
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, str):
        try:
            return datetime.strptime(val.strip(), "%Y-%m-%d").date()
        except ValueError:
            return None
    return None

def find_sheet(wb, keywords):
    for name in wb.sheetnames:
        if any(k.lower() in name.lower() for k in keywords):
            return wb[name]
    return None

def find_label_value(ws, label, max_scan=15):
    for r in range(1, max_scan + 1):
        for c in range(1, 6):
            cell = ws.cell(row=r, column=c)
            if cell.value and str(cell.value).strip().lower() == label.lower():
                if c < 5 and ws.cell(row=r, column=c+1).value:
                    return ws.cell(row=r, column=c+1).value
                if r < max_scan and ws.cell(row=r+1, column=c).value:
                    return ws.cell(row=r+1, column=c).value
    return None

def map_headers(ws, targets, max_row=5):
    for row in ws.iter_rows(min_row=1, max_row=max_row, values_only=False):
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                hdr = cell.value.strip().lower()
                for t in targets:
                    if t in hdr:
                        targets[t] = cell.column
        if all(v is not None for v in targets.values()):
            break
    return targets

def main():
    if len(sys.argv) != 5:
        print("Usage: python3 calculate_recovery_tracker.py <template.xlsx> <stock.xlsx> <recovery_log.xlsx> <output.xlsx>")
        sys.exit(1)
    
    tmpl_path, stock_path, log_path, out_path = sys.argv[1], sys.argv[2], sys.argv[3], sys.argv[4]
    
    # Load template to preserve structure
    tmpl_wb = load_workbook(tmpl_path)
    sheet_order = tmpl_wb.sheetnames
    
    # Load source data
    stock_wb = load_workbook(stock_path, data_only=True)
    log_wb = load_workbook(log_path, data_only=True)
    
    stock_ws = find_sheet(stock_wb, ["stock", "snapshot", "current", "inventory"])
    log_ws = find_sheet(log_wb, ["recovery", "log", "load", "shipment", "booking"])
    pallet_ws = find_sheet(tmpl_wb, ["pallet", "guide", "config", "setup"])
    
    if not stock_ws or not log_ws:
        print("Error: Could not locate Stock or Recovery Log sheets.")
        sys.exit(1)
    
    # Extract parameters
    asof = parse_date(find_label_value(stock_ws, "asofdate"))
    horizon = parse_date(find_label_value(stock_ws, "horizonend"))
    planning_days = (horizon - asof).days if horizon and asof else 30
    
    # Read pallet guide (SKU-specific)
    pallet_config = {}
    if pallet_ws:
        for row in pallet_ws.iter_rows(min_row=2, values_only=True):
            if row and row[0] and row[1]:
                try:
                    pallet_config[str(row[0]).strip()] = int(row[1])
                except (ValueError, TypeError):
                    pass
    
    # Read stock
    stock_hdrs = map_headers(stock_ws, {"sku": None, "units": None, "daily": None})
    stock = {}
    for row in stock_ws.iter_rows(min_row=1, values_only=True):
        if not row or not row[0]:
            continue
        sku = row[stock_hdrs["sku"]-1] if stock_hdrs["sku"] else None
        if not sku or not isinstance(sku, str):
            continue
        try:
            units = float(row[stock_hdrs["units"]-1] if stock_hdrs["units"] else 0)
            daily = float(row[stock_hdrs["daily"]-1] if stock_hdrs["daily"] else 0)
            stock[sku] = {"units": units, "daily": daily}
        except (ValueError, TypeError):
            continue
    
    # Read recovery log and deduplicate by revision
    log_hdrs = map_headers(log_ws, {"load_id": None, "revision": None, "sku": None, "date": None, "units": None, "stage": None})
    loads_by_id = {}
    for row in log_ws.iter_rows(min_row=1, values_only=True):
        if not row or not row[0]:
            continue
        load_id = row[log_hdrs["load_id"]-1] if log_hdrs["load_id"] else None
        if not load_id:
            continue
        try:
            revision = int(row[log_hdrs["revision"]-1] if log_hdrs["revision"] else 0)
        except (ValueError, TypeError):
            revision = 0
        
        # Keep highest revision per Load ID
        if load_id not in loads_by_id or revision > loads_by_id[load_id]["revision"]:
            loads_by_id[load_id] = {
                "load_id": load_id,
                "revision": revision,
                "sku": row[log_hdrs["sku"]-1] if log_hdrs["sku"] else None,
                "date": parse_date(row[log_hdrs["date"]-1] if log_hdrs["date"] else None),
                "units": float(row[log_hdrs["units"]-1] if log_hdrs["units"] else 0),
                "stage": str(row[log_hdrs["stage"]-1] if log_hdrs["stage"] else "").strip().lower()
            }
    
    # Filter by stage and horizon
    qualifying_loads = []
    for load in loads_by_id.values():
        if load["stage"] in IGNORE_STAGES:
            continue
        if load["stage"] not in VALID_STAGES:
            continue
        if load["date"] and load["date"] > horizon:
            continue
        qualifying_loads.append(load)
    
    # Compute per-SKU metrics
    results = []
    for sku, data in stock.items():
        units, daily = data["units"], data["daily"]
        days_oh = float('inf') if daily == 0 else units / daily
        oos_date = horizon + timedelta(days=1) if daily == 0 else asof + timedelta(days=days_oh)
        
        sku_loads = [l for l in qualifying_loads if l["sku"] == sku]
        inbound_units = sum(l["units"] for l in sku_loads)
        earliest_load = min((l["date"] for l in sku_loads if l["date"]), default=None)
        
        delivered_doh = (units + inbound_units) / daily if daily > 0 else 0
        remaining_demand = daily * planning_days
        additional = max(0, remaining_demand - (units + inbound_units))
        cases_per_pallet = pallet_config.get(sku, 40)
        loads_req = math.ceil(additional / cases_per_pallet) if cases_per_pallet > 0 else 0
        
        earlier = True
        if inbound_units > 0 and earliest_load and earliest_load <= oos_date:
            earlier = False
        
        results.append({
            "sku": sku, "units": units, "daily": daily,
            "days_oh": round(days_oh, 2) if days_oh != float('inf') else "N/A",
            "oos_date": oos_date, "inbound": inbound_units,
            "delivered_doh": round(delivered_doh, 2), "demand": remaining_demand,
            "additional": additional, "loads_req": loads_req,
            "rdd": oos_date, "earlier": earlier
        })
    
    # Build output workbook preserving template structure
    out_wb = Workbook()
    
    # Copy unmodified sheets first
    for sname in sheet_order:
        src_ws = tmpl_wb[sname]
        if sname.lower() in ["instructions", "pallet guide", "guide", "config"]:
            dst_ws = out_wb.create_sheet(sname)
            for row in src_ws.iter_rows(values_only=True):
                dst_ws.append(list(row))
    
    # Write Coverage_Detail
    cd_ws = out_wb.create_sheet("Coverage_Detail")
    cd_ws.append(["Field", "Value"])
    cd_ws.append(["AsOfDate", asof.isoformat() if asof else ""])
    cd_ws.append(["HorizonEnd", horizon.isoformat() if horizon else ""])
    cd_ws.append(["PlanningDays", planning_days])
    cd_ws.append([])
    
    headers = ["SKU", "Units_On_Hand", "Daily_Rate_Units_Per_Day", "Current_Days_On_Hand",
               "Projected_OOS_Date", "Inbound_Units_By_Horizon", "Delivered_Days_On_Hand",
               "Remaining_Demand_Units", "Additional_Units_Needed", "Loads_Required",
               "Required_Delivery_Date", "Earlier_Delivery_Required"]
    cd_ws.append(headers)
    
    for r in results:
        cd_ws.append([
            r["sku"], r["units"], r["daily"], r["days_oh"],
            r["oos_date"].isoformat() if isinstance(r["oos_date"], (datetime,)) else r["oos_date"],
            r["inbound"], r["delivered_doh"], r["demand"],
            r["additional"], r["loads_req"],
            r["rdd"].isoformat() if isinstance(r["rdd"], (datetime,)) else r["rdd"],
            r["earlier"]
        ])
    
    # Write Recovery_Loads
    rl_ws = out_wb.create_sheet("Recovery_Loads")
    rl_headers = ["SKU", "Required_Delivery_Date", "Loads_Required", "Additional_Units_Needed", "Earlier_Delivery_Required"]
    rl_ws.append(rl_headers)
    for r in results:
        if r["loads_req"] > 0:
            rl_ws.append([
                r["sku"],
                r["rdd"].isoformat() if isinstance(r["rdd"], (datetime,)) else r["rdd"],
                r["loads_req"], r["additional"], r["earlier"]
            ])
    
    out_wb.save(out_path)
    print(f"Saved recovery tracker to {out_path}")

if __name__ == "__main__":
    main()