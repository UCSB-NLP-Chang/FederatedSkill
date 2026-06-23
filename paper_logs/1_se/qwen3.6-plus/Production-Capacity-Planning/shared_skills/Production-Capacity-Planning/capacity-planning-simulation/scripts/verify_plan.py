#!/usr/bin/env python3
"""Verifies capacity planning outputs against strict constraints."""
import sys
import openpyxl
import re

def verify_excel(path):
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    if ws.title != "Plan":
        return False, f"Sheet name is '{ws.title}', expected 'Plan'"
    
    headers = [cell.value for cell in ws[1]]
    expected = ["Week", "Days Worked", "Scheduled Demand (Std Hrs)", "Weekly Capacity (Std Hrs)", 
                "Start of Week Past Due (Std Hrs)", "End of Week Backlog/Buffer (Std Hrs)", "Overtime Hours"]
    if headers != expected:
        return False, f"Headers mismatch: {headers}"
        
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    if len(rows) < 1:
        return False, "No data rows found"
        
    for i in range(len(rows)-1):
        if abs(rows[i][5] - rows[i+1][4]) > 0.01:
            return False, f"Backlog discontinuity at week {rows[i][0]}: EoW {rows[i][5]} != SoW {rows[i+1][4]}"
            
    return True, f"Excel OK: {len(rows)} rows, headers match, continuity verified."

def verify_summary(path):
    with open(path, 'r') as f:
        lines = [l.strip() for l in f.readlines() if l.strip()]
    if len(lines) != 3:
        return False, f"Expected 3 lines, got {len(lines)}"
        
    summary_line = lines[2]
    if summary_line.startswith("Summary: "):
        summary_line = summary_line[9:]
        
    sentences = [s.strip() for s in re.split(r'[.!?]+', summary_line) if s.strip()]
    words = len(summary_line.split())
    
    checks = []
    if len(sentences) != 3:
        checks.append(f"Expected 3 sentences, got {len(sentences)}")
    if not (30 <= words <= 45):
        checks.append(f"Word count {words} outside 30-45 range")
        
    # Verify step-down references are explicitly mentioned
    has_week_ref = bool(re.search(r'(Week\s*\d+|N/A)', summary_line, re.IGNORECASE))
    if not has_week_ref:
        checks.append("Summary must explicitly mention step-down week numbers or N/A")
        
    if checks:
        return False, "; ".join(checks)
    return True, f"Summary OK: {len(sentences)} sentences, {words} words."

if __name__ == "__main__":
    excel_path = sys.argv[1] if len(sys.argv) > 1 else "/root/catch_up_plan.xlsx"
    summary_path = sys.argv[2] if len(sys.argv) > 2 else "/root/catch_up_summary.txt"
    
    ok1, msg1 = verify_excel(excel_path)
    ok2, msg2 = verify_summary(summary_path)
    print(msg1)
    print(msg2)
    sys.exit(0 if ok1 and ok2 else 1)
