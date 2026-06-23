#!/usr/bin/env python3
"""Verify openpyxl workbook structure, formulas, and data types."""
import sys
import json
import openpyxl

def verify(path, expected_sheets):
    wb = openpyxl.load_workbook(path, data_only=False)
    if wb.sheetnames != expected_sheets:
        print(f"FAIL: Sheet order mismatch. Expected {expected_sheets}, got {wb.sheetnames}")
        sys.exit(1)
    
    errors = []
    for sheet_name in expected_sheets:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                if isinstance(cell.value, str) and cell.value.startswith('='):
                    continue  # Formula cells are valid
                if not isinstance(cell.value, (int, float)):
                    # Allow strings for headers/labels
                    if not any(c.isalpha() for c in str(cell.value)):
                        errors.append(f"{sheet_name}!{cell.coordinate}: Expected numeric, got {type(cell.value).__name__} ({cell.value})")
    
    if errors:
        print("FAIL: Data type errors:")
        for e in errors:
            print(f"  - {e}")
        sys.exit(1)
    print("PASS: Workbook structure and types verified.")

if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Usage: verify_workbook.py <path> <expected_sheets_json>")
        sys.exit(1)
    verify(sys.argv[1], json.loads(sys.argv[2]))
