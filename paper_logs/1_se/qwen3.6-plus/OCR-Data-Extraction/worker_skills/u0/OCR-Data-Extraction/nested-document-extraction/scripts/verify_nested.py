#!/usr/bin/env python3
"""Verify nested document extraction outputs for schema, path format, and date normalization."""
import sys
import os
import re
from openpyxl import load_workbook

def verify(excel_path, expected_cols=None):
    if not os.path.exists(excel_path):
        print(f"FAIL: {excel_path} not found.")
        return False

    wb = load_workbook(excel_path)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        print("FAIL: Empty workbook.")
        return False

    headers = [str(h).strip() for h in rows[0]]
    data = rows[1:]

    if expected_cols and headers != expected_cols:
        print(f"FAIL: Column mismatch. Expected {expected_cols}, got {headers}")
        return False

    date_re = re.compile(r"^\d{4}-\d{2}-\d{2}$")
    errors = []
    for i, row in enumerate(data, 2):
        # Check relative path format
        if "/" not in str(row[1]):
            errors.append(f"Row {i}: relative_path missing directory separator.")
        # Check date format
        if not date_re.match(str(row[3])):
            errors.append(f"Row {i}: date '{row[3]}' not ISO YYYY-MM-DD.")
        # Check amount format
        try:
            amt = float(row[4])
            if f"{amt:.2f}" != str(row[4]):
                errors.append(f"Row {i}: amount '{row[4]}' not 2 decimal places.")
        except:
            errors.append(f"Row {i}: invalid amount.")

    if errors:
        print("FAIL: Validation errors:")
        for e in errors: print(f"  - {e}")
        return False

    print(f"Verified {len(data)} rows. Schema and formats OK.")
    return True

if __name__ == "__main__":
    cols = sys.argv[2].split(",") if len(sys.argv) > 2 else None
    verify(sys.argv[1], cols)
