#!/usr/bin/env python3
"""Verification script for document extraction outputs (orders, claims, etc.)."""
import sys
import os
import glob
import csv
from openpyxl import load_workbook

def verify_excel(excel_path, img_dir, ref_csv=None, expected_cols=None):
    if not os.path.exists(excel_path):
        print(f"FAIL: {excel_path} not found.")
        return False
        
    wb = load_workbook(excel_path)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=1, values_only=True))
    if not rows:
        print("FAIL: Empty workbook.")
        return False
        
    headers = [str(h).strip() for h in rows[0]]
    data_rows = rows[1:]
    
    if expected_cols:
        if headers != expected_cols:
            print(f"FAIL: Column mismatch. Expected {expected_cols}, got {headers}")
            return False
            
    images = sorted(glob.glob(os.path.join(img_dir, "*.jpg"))) + \
             sorted(glob.glob(os.path.join(img_dir, "*.png")))
             
    print(f"Images found: {len(images)}")
    print(f"Excel data rows: {len(data_rows)}")
    
    if len(data_rows) != len(images):
        print("FAIL: Row count mismatch!")
        return False
        
    # Check for unexpected nulls in key columns
    null_counts = {}
    for i, h in enumerate(headers):
        null_counts[h] = sum(1 for r in data_rows if r[i] is None or str(r[i]).strip() == "")
        
    print(f"Empty/Null counts per column: {null_counts}")
    
    if ref_csv and os.path.exists(ref_csv):
        with open(ref_csv, "r") as f:
            reader = csv.DictReader(f)
            ref_keys = set(row[reader.fieldnames[0]].strip() for row in reader)
        # Assume first data column is the key
        extracted_keys = set(str(r[0]).strip() for r in data_rows if r[0] is not None)
        unmatched = extracted_keys - ref_keys
        if unmatched:
            print(f"INFO: {len(unmatched)} keys not in reference list (expected if allowed).")
            
    print("Verification complete.")
    return True

if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Usage: verify_output.py <excel_path> <img_dir> [ref_csv] [col1,col2,...]")
        sys.exit(1)
    cols = sys.argv[4].split(",") if len(sys.argv) > 4 else None
    verify_excel(sys.argv[1], sys.argv[2], sys.argv[3] if len(sys.argv) > 3 else None, cols)
