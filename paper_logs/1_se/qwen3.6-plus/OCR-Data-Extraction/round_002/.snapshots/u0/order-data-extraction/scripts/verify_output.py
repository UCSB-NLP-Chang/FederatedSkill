#!/usr/bin/env python3
"""Quick verification script for order extraction outputs."""
import sys
import os
import glob
import csv
from openpyxl import load_workbook

def verify_excel(excel_path, img_dir, known_csv=None):
    wb = load_workbook(excel_path)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    
    images = sorted(glob.glob(os.path.join(img_dir, "*.jpg")))
    
    print(f"Images found: {len(images)}")
    print(f"Excel rows: {len(rows)}")
    
    if len(rows) != len(images):
        print("FAIL: Row count mismatch!")
        return False
        
    null_counts = {"order_id": 0, "date": 0, "total_amount": 0}
    for row in rows:
        if len(row) >= 4:
            if row[1] is None: null_counts["order_id"] += 1
            if row[2] is None: null_counts["date"] += 1
            if row[3] is None: null_counts["total_amount"] += 1
            
    print(f"Null counts: {null_counts}")
    
    if known_csv and os.path.exists(known_csv):
        with open(known_csv, "r") as f:
            reader = csv.reader(f)
            known_ids = set(row[0] for row in reader if row)
        extracted_ids = set(str(row[1]) for row in rows if row[1] is not None)
        unknown = extracted_ids - known_ids
        if unknown:
            print(f"WARNING: Unknown IDs found: {unknown}")
            
    print("Verification complete.")
    return True

if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Usage: verify_output.py <excel_path> <img_dir> [known_csv]")
        sys.exit(1)
    verify_excel(sys.argv[1], sys.argv[2], sys.argv[3] if len(sys.argv) > 3 else None)
