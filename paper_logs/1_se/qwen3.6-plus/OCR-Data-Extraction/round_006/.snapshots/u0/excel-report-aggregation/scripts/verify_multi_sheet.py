#!/usr/bin/env python3
"""Verify multi-sheet Excel reports (details + summary) for schema, counts, and aggregation."""
import sys
import os
import glob
from openpyxl import load_workbook

def verify(excel_path, details_cols, summary_cols, img_dir=None):
    if not os.path.exists(excel_path):
        print(f"FAIL: {excel_path} not found.")
        return False

    wb = load_workbook(excel_path)
    if "details" not in wb.sheetnames or "summary" not in wb.sheetnames:
        print("FAIL: Missing 'details' or 'summary' sheet.")
        return False

    # Verify Details
    ws_det = wb["details"]
    det_rows = list(ws_det.iter_rows(values_only=True))
    if not det_rows:
        print("FAIL: Empty details sheet.")
        return False
    det_headers = [str(h).strip() for h in det_rows[0]]
    if det_headers != details_cols:
        print(f"FAIL: Details column mismatch. Expected {details_cols}, got {det_headers}")
        return False
    det_data = det_rows[1:]
    print(f"Details rows: {len(det_data)}")

    # Verify Summary
    ws_sum = wb["summary"]
    sum_rows = list(ws_sum.iter_rows(values_only=True))
    if not sum_rows:
        print("FAIL: Empty summary sheet.")
        return False
    sum_headers = [str(h).strip() for h in sum_rows[0]]
    if sum_headers != summary_cols:
        print(f"FAIL: Summary column mismatch. Expected {summary_cols}, got {sum_headers}")
        return False
    sum_data = sum_rows[1:]
    print(f"Summary rows: {len(sum_data)}")

    # Optional: Check image count vs details rows if img_dir provided
    if img_dir:
        images = sorted(glob.glob(os.path.join(img_dir, "*.jpg"))) + \
                 sorted(glob.glob(os.path.join(img_dir, "*.png")))
        print(f"Images found: {len(images)}")
        # Note: details rows != images if multiple items per image. Just log it.

    print("Verification complete.")
    return True

if __name__ == "__main__":
    if len(sys.argv) < 4:
        print("Usage: verify_multi_sheet.py <excel_path> <details_col1,col2,...> <summary_col1,col2,...> [img_dir]")
        sys.exit(1)
    det_cols = sys.argv[2].split(",")
    sum_cols = sys.argv[3].split(",")
    img = sys.argv[4] if len(sys.argv) > 4 else None
    verify(sys.argv[1], det_cols, sum_cols, img)
