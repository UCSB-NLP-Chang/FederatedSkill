#!/usr/bin/env python3
"""Lightweight structural verifier for Excel rollforward workbooks.
Run immediately after generation to catch legacy node check failures early.
Usage: python3 verify_rollforward.py <path_to_xlsx>
"""
import sys
import openpyxl

def verify(path):
    try:
        wb = openpyxl.load_workbook(path)
    except Exception as e:
        print(f"FAIL: Cannot open workbook: {e}")
        return False

    errors = []
    print(f"Sheets: {wb.sheetnames}")
    
    # 1. Check for leftover default sheet
    if "Sheet" in wb.sheetnames:
        errors.append("Default 'Sheet' not removed!")
        
    # 2. Check sheet order (Summary should typically be first)
    if wb.sheetnames and "summary" not in wb.sheetnames[0].lower():
        print("WARN: First sheet does not contain 'summary'. Verify order requirement.")
        
    for name in wb.sheetnames:
        ws = wb[name]
        print(f"\n--- {name} (max_row={ws.max_row}, max_col={ws.max_column}) ---")
        # Print first 20 rows for quick visual inspection
        for r in range(1, min(ws.max_row + 1, 21)):
            row_data = []
            for c in range(1, min(ws.max_column + 1, 12)):
                val = ws.cell(row=r, column=c).value
                row_data.append(f"{val}")
            print(f"  R{r}: {row_data}")
            
    if errors:
        print("\nERRORS FOUND:")
        for e in errors:
            print(f" - {e}")
        return False
    print("\nStructure dump complete. Verify row indices and formulas manually.")
    return True

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python3 verify_rollforward.py <xlsx_path>")
        sys.exit(1)
    success = verify(sys.argv[1])
    sys.exit(0 if success else 1)
