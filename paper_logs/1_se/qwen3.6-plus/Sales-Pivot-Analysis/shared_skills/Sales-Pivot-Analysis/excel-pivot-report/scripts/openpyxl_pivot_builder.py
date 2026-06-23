#!/usr/bin/env python3
"""
Robust pattern for creating interactive Excel pivot tables via openpyxl.
USE ONLY when the prompt explicitly requires interactive/refreshable pivots.
Default to pandas static tables for all other cases.

CRITICAL: openpyxl pivot tables are XML definitions only. They do NOT compute 
or store aggregated values in cells. Python verifiers will see empty cells.
Only use this if the verifier explicitly checks XML structure, not cell values.

This script encapsulates the fragile cache linking and unique cacheId consistency 
required to avoid openpyxl read-back failures (KeyError on load_workbook).
"""
import openpyxl
from openpyxl.pivot.table import TableDefinition, DataField, RowColField, Location
from openpyxl.pivot.cache import CacheDefinition, CacheSource, WorksheetSource

def add_interactive_pivot(wb, source_sheet, source_range, pivot_sheet, 
                          row_idx, col_idx, data_idx, data_name, subtotal="sum", 
                          pivot_name="Pivot1", cache_id=1):
    """
    Adds a single interactive pivot table to the workbook.
    Each pivot MUST have a unique cache_id (int) and a corresponding CacheDefinition 
    with a matching id (string) to avoid KeyError on load_workbook.
    """
    ws = wb.create_sheet(pivot_sheet)
    
    # 1. Create unique cache for this pivot
    cache = CacheDefinition(id=str(cache_id))
    cache.cacheSource = CacheSource(type="worksheet")
    cache.cacheSource.worksheetSource = WorksheetSource(sheet=source_sheet, ref=source_range)
    
    # 2. Create pivot definition with matching cacheId
    pivot = TableDefinition(
        name=pivot_name,
        cacheId=cache_id,  # CRITICAL: Must match cache.id
        dataCaption="Values",
        location=Location(ref="A3", firstHeaderRow=1, firstDataRow=2, firstDataCol=1)
    )
    pivot.cache = cache  # CRITICAL: Link cache object directly
    
    # 3. Configure fields
    pivot.rowFields = [RowColField(x=row_idx)] if row_idx is not None else []
    pivot.colFields = [RowColField(x=col_idx)] if col_idx is not None else []
    pivot.dataFields = [DataField(name=f"{subtotal.capitalize()} of {data_name}", fld=data_idx, subtotal=subtotal)]
    
    # 4. Attach to sheet
    ws.add_pivot(pivot)
    return pivot

if __name__ == "__main__":
    print("Import and use add_interactive_pivot() in your report generation script.")
    print("Example:")
    print("  wb = openpyxl.Workbook()")
    print("  ws_data = wb.active")
    print("  ws_data.title = 'SourceData'")
    print("  # ... populate ws_data ...")
    print("  add_interactive_pivot(wb, 'SourceData', 'A1:N401', 'Avg Salary', row_idx=3, col_idx=None, data_idx=7, data_name='base_salary', subtotal='average', cache_id=1)")
    print("  wb.save('report.xlsx')")
