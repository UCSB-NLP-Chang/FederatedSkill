#!/usr/bin/env python3
"""Validate structure of a generated Excel workbook."""
import sys
import argparse
import openpyxl

def main():
    parser = argparse.ArgumentParser(description="Validate Excel workbook structure")
    parser.add_argument("workbook", help="Path to .xlsx file")
    parser.add_argument("--min-sheets", type=int, default=1)
    parser.add_argument("--min-named-ranges", type=int, default=0)
    parser.add_argument("--check-formula", help="Sheet!Cell to verify contains a formula")
    args = parser.parse_args()

    wb = openpyxl.load_workbook(args.workbook)
    errors = []

    if len(wb.sheetnames) < args.min_sheets:
        errors.append(f"Sheet count {len(wb.sheetnames)} < {args.min_sheets}")

    nr_count = len(list(wb.defined_names))
    if nr_count < args.min_named_ranges:
        errors.append(f"Named range count {nr_count} < {args.min_named_ranges}")

    if args.check_formula:
        sheet_name, cell = args.check_formula.split("!")
        ws = wb[sheet_name]
        val = ws[cell].value
        if not val or "=" not in str(val):
            errors.append(f"{sheet_name}!{cell} is not a formula")

    if errors:
        print("VALIDATION FAILED:")
        for e in errors:
            print(f"  - {e}")
        sys.exit(1)
    else:
        print("VALIDATION PASSED")
        sys.exit(0)

if __name__ == "__main__":
    main()
