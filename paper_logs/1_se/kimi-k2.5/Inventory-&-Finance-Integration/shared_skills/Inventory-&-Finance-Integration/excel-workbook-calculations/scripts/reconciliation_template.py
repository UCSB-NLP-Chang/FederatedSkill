#!/usr/bin/env python3
"""
Template for building reconciliation workbooks with linked formulas.
Usage: reconciliation, capacity planning, financial rollforwards.
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment


def build_reconciliation_workbook(
    output_path: str,
    detail_sheets: list,  # Each: {name, line_items_df, gl_balances, account_number}
    summary_config: dict  # {title, sections: [{name, detail_sheet_name, label}]}
) -> None:
    """
    Build reconciliation workbook with detail sheets and summary.
    
    Args:
        output_path: Path for output .xlsx file
        detail_sheets: List of detail sheet configs
        summary_config: Summary sheet configuration
    """
    wb = Workbook()
    wb.remove(wb.active)
    
    # Build detail sheets first
    for config in detail_sheets:
        ws = wb.create_sheet(title=config['name'])
        _build_detail_sheet(ws, config)
    
    # Build summary sheet (first position)
    summary_ws = wb.create_sheet(title=summary_config.get('title', 'Summary'), index=0)
    _build_summary_sheet(summary_ws, summary_config, detail_sheets)
    
    wb.save(output_path)


def _build_detail_sheet(ws, config):
    """Build single detail sheet with control rows."""
    df = config['line_items_df']
    gl_balances = config['gl_balances']  # dict: {'jan': float, 'feb': float, ...}
    months = list(gl_balances.keys())
    
    # Title
    ws['A1'] = f"{config['name']} - Reconciliation"
    ws['A1'].font = Font(bold=True, size=14)
    
    # Headers row 6
    ws['A6'] = 'Line Item'
    for col_idx, month in enumerate(months, start=2):
        ws.cell(row=6, column=col_idx, value=month.capitalize())
    
    # Data rows 7+
    for row_idx, (_, row) in enumerate(df.iterrows(), start=7):
        ws.cell(row=row_idx, column=1, value=row['name'])
        for col_idx, month in enumerate(months, start=2):
            value = row.get(f'{month}_ending_balance', 0)
            ws.cell(row=row_idx, column=col_idx, value=float(value))
    
    # Control rows (13-16 typical)
    control_start = 7 + len(df) + 1  # 1 row gap after data
    
    # Month Totals row
    ws.cell(row=control_start, column=1, value='Month Totals')
    for col_idx, month in enumerate(months, start=2):
        col_letter = chr(64 + col_idx)  # B=66, C=67, ...
        end_row = control_start - 2  # Last data row
        start_row = 7
        ws.cell(row=control_start, column=col_idx, 
                value=f'=SUM({col_letter}{start_row}:{col_letter}{end_row})')
    
    # Ending Balance row
    ws.cell(row=control_start + 1, column=1, value='Ending Balance')
    for col_idx, month in enumerate(months, start=2):
        col_letter = chr(64 + col_idx)
        ws.cell(row=control_start + 1, column=col_idx,
                value=f'={col_letter}{control_start}')
    
    # Variance row
    ws.cell(row=control_start + 2, column=1, value='Variance')
    for col_idx, month in enumerate(months, start=2):
        col_letter = chr(64 + col_idx)
        ws.cell(row=control_start + 2, column=col_idx,
                value=f'={col_letter}{control_start + 1}-{col_letter}{control_start + 3}')
    
    # GL Balance row (static values)
    ws.cell(row=control_start + 3, column=1, value='GL Balance')
    for col_idx, month in enumerate(months, start=2):
        ws.cell(row=control_start + 3, column=col_idx, 
                value=float(gl_balances[month]))


def _build_summary_sheet(ws, config, detail_sheets):
    """Build summary with cross-sheet references."""
    ws['A1'] = config.get('title', 'Reconciliation Summary')
    ws['A1'].font = Font(bold=True, size=14)
    
    current_row = 6
    variance_cells = []
    
    for section in config['sections']:
        # Section header
        ws.cell(row=current_row, column=1, value=section['label'])
        ws.cell(row=current_row, column=1).font = Font(bold=True)
        
        # Find the detail sheet control row positions
        detail_name = section['detail_sheet_name']
        # Assume standard positions: Month Totals=row 13, Ending Bal=14, Variance=15
        control_row_base = 13  # Configure as needed
        
        # Month Totals link
        ws.cell(row=current_row + 1, column=1, value='Month Totals')
        ws.cell(row=current_row + 1, column=2, 
                value=f"='{detail_name}'!O{control_row_base}")
        
        # Ending Balance link
        ws.cell(row=current_row + 2, column=1, value='Ending Balance')
        ws.cell(row=current_row + 2, column=2,
                value=f"='{detail_name}'!O{control_row_base + 1}")
        
        # Variance link
        ws.cell(row=current_row + 3, column=1, value='Variance')
        var_cell = f'B{current_row + 3}'
        ws.cell(row=current_row + 3, column=2,
                value=f"='{detail_name}'!O{control_row_base + 2}")
        variance_cells.append(var_cell)
        
        current_row += 5  # Gap between sections
    
    # Combined total
    ws.cell(row=current_row, column=1, value='Combined Total')
    ws.cell(row=current_row, column=1).font = Font(bold=True)
    if len(variance_cells) >= 2:
        ws.cell(row=current_row, column=2, 
                value=f'={variance_cells[0]}+{variance_cells[1]}')


if __name__ == '__main__':
    # Example usage
    detail_df = pd.DataFrame({
        'name': ['Vendor A', 'Vendor B'],
        'jan_ending_balance': [1000.0, 2000.0],
        'feb_ending_balance': [1100.0, 2200.0],
    })
    
    build_reconciliation_workbook(
        '/tmp/example_recon.xlsx',
        detail_sheets=[{
            'name': 'Detail Sheet',
            'line_items_df': detail_df,
            'gl_balances': {'jan': 3000.0, 'feb': 3300.0},
            'account_number': '8100'
        }],
        summary_config={
            'title': 'Example Reconciliation',
            'sections': [{'label': 'Section 1', 'detail_sheet_name': 'Detail Sheet'}]
        }
    )
    print("Created /tmp/example_recon.xlsx")
