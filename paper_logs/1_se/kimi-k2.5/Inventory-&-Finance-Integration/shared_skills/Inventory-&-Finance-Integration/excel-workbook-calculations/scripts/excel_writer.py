#!/usr/bin/env python3
"""
Template for creating calculated Excel workbooks with metadata headers.
"""

from datetime import date
from typing import Dict, Any, List
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows


def write_calculated_workbook(
    output_path: str,
    metadata: Dict[str, Any],
    sheets: List[Dict[str, Any]]
) -> None:
    """
    Create Excel workbook with metadata headers and calculated data sheets.
    
    Args:
        output_path: Path for output file
        metadata: Dict of field names to values (written as key-value rows)
        sheets: List of sheet configs, each with:
            - name: sheet name
            - df: DataFrame to write
            - start_row: row to start writing data (1-indexed, default 6)
    
    Example:
        metadata = {
            'AsOfDate': '2025-07-04',
            'PlanningHorizonEnd': '2025-07-31',
            'RemainingDaysInJuly': 27
        }
        sheets = [
            {'name': 'SKU_Results', 'df': df_results, 'start_row': 6},
            {'name': 'Filtered_View', 'df': df_filtered, 'start_row': 1}
        ]
        write_calculated_workbook('output.xlsx', metadata, sheets)
    """
    wb = Workbook()
    
    # Remove default sheet, create custom
    wb.remove(wb.active)
    
    for sheet_config in sheets:
        name = sheet_config['name']
        df = sheet_config['df']
        start_row = sheet_config.get('start_row', 6)
        write_metadata = sheet_config.get('write_metadata', start_row > 1)
        
        ws = wb.create_sheet(title=name)
        
        current_row = 1
        
        # Write metadata if requested
        if write_metadata and name == sheets[0]['name']:
            ws['A1'] = 'Field'
            ws['B1'] = 'Value'
            current_row = 2
            for key, value in metadata.items():
                ws.cell(row=current_row, column=1, value=key)
                # Convert dates to ISO strings, booleans to native
                if isinstance(value, date):
                    value = value.isoformat()
                ws.cell(row=current_row, column=2, value=value)
                current_row += 1
            current_row = start_row
        
        # Write DataFrame
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), current_row):
            for c_idx, value in enumerate(row, 1):
                # Convert numpy types to native Python for Excel
                if hasattr(value, 'item'):  # numpy scalar
                    value = value.item()
                elif isinstance(value, pd.Timestamp):
                    value = value.isoformat()
                ws.cell(row=r_idx, column=c_idx, value=value)
    
    wb.save(output_path)


if __name__ == '__main__':
    # Example usage
    df = pd.DataFrame({
        'Product': ['A', 'B'],
        'Cases': [100, 200],
        'Pallets': [2, 3]
    })
    
    write_calculated_workbook(
        '/tmp/example.xlsx',
        metadata={'AsOfDate': '2025-07-04', 'Period': 'July 2025'},
        sheets=[{'name': 'Data', 'df': df, 'start_row': 4}]
    )
    print("Created /tmp/example.xlsx")