#!/usr/bin/env python3
"""
Extract implied capacity constants from an existing plan Excel file.
Use this to verify capacity constants when a task provides an existing plan file.
"""

import sys
from openpyxl import load_workbook

def extract_constants(filepath, sheet_name='Plan'):
    """Extract unique (days_worked, weekly_capacity, overtime) combinations."""
    wb = load_workbook(filepath, data_only=True)
    
    if sheet_name not in wb.sheetnames:
        print(f"Error: Sheet '{sheet_name}' not found. Available: {wb.sheetnames}", file=sys.stderr)
        sys.exit(1)
    
    ws = wb[sheet_name]
    
    # Find header row
    header_row = None
    days_col = None
    capacity_col = None
    ot_col = None
    
    for idx, row in enumerate(ws.iter_rows(max_row=5, values_only=True), 1):
        row_lower = [str(c).lower() if c else '' for c in row]
        if 'days' in row_lower or 'days worked' in ' '.join(row_lower):
            header_row = idx
            for i, cell in enumerate(row):
                cell_str = str(cell).lower() if cell else ''
                if 'day' in cell_str:
                    days_col = i
                elif 'capacity' in cell_str and 'weekly' in cell_str:
                    capacity_col = i
                elif 'overtime' in cell_str:
                    ot_col = i
            break
    
    if days_col is None or capacity_col is None:
        print("Could not identify columns. Assuming: 2=Days, 4=Capacity, 7=Overtime", file=sys.stderr)
        days_col = 1
        capacity_col = 3
        ot_col = 6
    
    # Extract unique combinations
    constants = {}
    for row in ws.iter_rows(min_row=(header_row or 1) + 1, values_only=True):
        if len(row) <= max(days_col, capacity_col, ot_col):
            continue
        days = row[days_col]
        capacity = row[capacity_col]
        ot = row[ot_col] if ot_col and ot_col < len(row) else 0
        
        try:
            days = int(days)
            capacity = float(capacity)
            ot = float(ot) if ot is not None else 0
            if days not in constants:
                constants[days] = {'capacity': capacity, 'overtime': ot}
        except (ValueError, TypeError):
            continue
    
    # Calculate implied base rate
    print("Extracted Capacity Constants:")
    print("-" * 40)
    for days in sorted(constants.keys()):
        cap = constants[days]['capacity']
        ot = constants[days]['overtime']
        base_rate = cap / days if days > 0 else 0
        print(f"{days}-day: Capacity={cap}, Overtime={ot}, Base Rate={base_rate:.2f}")
    
    # Verify consistency
    base_rates = [constants[d]['capacity'] / d for d in constants if d > 0]
    if len(set(round(b, 1) for b in base_rates)) == 1:
        print(f"\nConsistent base rate: {base_rates[0]:.2f} hrs/day")
    else:
        print(f"\nWARNING: Inconsistent base rates detected: {[round(b, 2) for b in base_rates]}")
    
    # Suggest CAPACITY dict for Python
    print("\nSuggested Python constants:")
    print("CAPACITY = {")
    for days in sorted(constants.keys()):
        print(f"    {days}: {int(constants[days]['capacity'])},  # {days}-day")
    print("}")
    print("OVERTIME = {")
    for days in sorted(constants.keys()):
        print(f"    {days}: {int(constants[days]['overtime'])},  # {days}-day")
    print("}")

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: extract_capacity_constants.py <excel_file> [sheet_name]", file=sys.stderr)
        sys.exit(1)
    
    filepath = sys.argv[1]
    sheet = sys.argv[2] if len(sys.argv) > 2 else 'Plan'
    extract_constants(filepath, sheet)