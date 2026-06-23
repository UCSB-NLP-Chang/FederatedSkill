#!/usr/bin/env python3
"""
Defensive re-round of all numeric cells in Excel output.
Run this before final verification to prevent floating-point precision failures.
"""

import sys
from openpyxl import load_workbook

def reround_excel(filepath, sheet_name='Plan', decimals=2):
    """Re-round all numeric cells in specified Excel sheet."""
    wb = load_workbook(filepath)
    
    if sheet_name not in wb.sheetnames:
        print(f"Error: Sheet '{sheet_name}' not found. Available: {wb.sheetnames}", file=sys.stderr)
        sys.exit(1)
    
    ws = wb[sheet_name]
    modified = 0
    
    for row in ws.iter_rows(min_row=2):  # Skip header
        for cell in row:
            if isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool):
                original = cell.value
                rounded = round(float(original), decimals)
                if original != rounded:
                    cell.value = rounded
                    modified += 1
    
    wb.save(filepath)
    print(f"Re-rounded {modified} cells to {decimals} decimals in {filepath}")

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: defensive_reround.py <excel_file> [sheet_name]", file=sys.stderr)
        sys.exit(1)
    
    filepath = sys.argv[1]
    sheet = sys.argv[2] if len(sys.argv) > 2 else 'Plan'
    reround_excel(filepath, sheet)
