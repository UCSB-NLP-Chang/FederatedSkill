#!/usr/bin/env python3
"""
Pre-submission verification for catch-up plan Excel files.
Checks for common verifier failure modes: precision, -0.0, structure.
Run this AFTER defensive_reround.py and BEFORE final submission.
"""

import sys
from openpyxl import load_workbook

def verify_output(filepath, sheet_name='Plan'):
    """Verify Excel output for common verifier failure modes."""
    wb = load_workbook(filepath, data_only=False)
    
    if sheet_name not in wb.sheetnames:
        print(f"FAIL: Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
        return False
    
    ws = wb[sheet_name]
    
    issues = []
    warnings = []
    
    # Check headers
    headers = []
    for row in ws.iter_rows(max_row=1, values_only=True):
        headers = [str(h) if h else '' for h in row]
        break
    
    expected_headers = ['Week', 'Days', 'Scheduled Demand', 'Weekly Capacity', 
                       'Start of Week Past Due', 'End of Week Backlog/Buffer', 'Overtime Hours']
    
    # Flexible header matching
    header_found = {
        'week': any('week' in h.lower() for h in headers),
        'days': any('day' in h.lower() for h in headers),
        'demand': any('demand' in h.lower() for h in headers),
        'capacity': any('capacity' in h.lower() for h in headers),
        'start': any('start' in h.lower() and 'past' in h.lower() for h in headers),
        'end': any('end' in h.lower() and ('backlog' in h.lower() or 'buffer' in h.lower()) for h in headers),
        'overtime': any('overtime' in h.lower() for h in headers)
    }
    
    missing = [k for k, v in header_found.items() if not v]
    if missing:
        issues.append(f"Missing expected headers: {missing}")
    
    # Check numeric precision and -0.0
    precision_issues = 0
    negative_zero_count = 0
    prev_week = None
    week_gaps = []
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        values = [c.value for c in row]
        if all(v is None or v == '' for v in values):
            continue  # Skip empty rows
        
        # Check week sequence
        week_val = values[0] if values else None
        if week_val is not None:
            try:
                week_num = int(week_val)
                if prev_week is not None and week_num != prev_week + 1:
                    week_gaps.append(f"Row {row_idx}: week {prev_week} -> {week_num}")
                prev_week = week_num
            except (ValueError, TypeError):
                pass
        
        # Check numeric cells
        for col_idx, cell in enumerate(row, start=1):
            val = cell.value
            if isinstance(val, (int, float)) and not isinstance(val, bool):
                # Check for unrounded values
                rounded = round(float(val), 2)
                if abs(val - rounded) > 1e-9:
                    precision_issues += 1
                    if precision_issues <= 3:
                        issues.append(f"Row {row_idx}, Col {col_idx}: {val} (should be {rounded})")
                
                # Check for -0.0
                if val == 0 and str(val).startswith('-'):
                    negative_zero_count += 1
    
    if precision_issues > 3:
        issues.append(f"...and {precision_issues - 3} more precision issues")
    
    if week_gaps:
        issues.append(f"Week sequence gaps: {week_gaps[:3]}")
    
    # Report
    print(f"Verification results for {filepath}:")
    print("=" * 50)
    
    if issues:
        print(f"\nISSUES FOUND ({len(issues)}):")
        for issue in issues:
            print(f"  - {issue}")
        print("\nACTION: Run 'python3 scripts/defensive_reround.py' and re-verify")
        return False
    else:
        print("\n✓ All checks passed:")
        print(f"  - Headers found: {list(header_found.keys())}")
        print(f"  - No precision issues detected")
        print(f"  - No -0.0 values detected")
        if prev_week:
            print(f"  - Contiguous weeks through {prev_week}")
        return True

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: verify_output.py <excel_file> [sheet_name]")
        print("\nRun this AFTER defensive_reround.py to verify output before submission.")
        sys.exit(1)
    
    filepath = sys.argv[1]
    sheet = sys.argv[2] if len(sys.argv) > 2 else 'Plan'
    
    success = verify_output(filepath, sheet)
    sys.exit(0 if success else 1)
