#!/usr/bin/env python3
"""
Generate load planning workbook from source Excel data.

Usage:
    python3 generate_load_plan.py --source /path/to/input.xlsx --output /path/to/output.xlsx
    
Advanced (variant structures):
    python3 generate_load_plan.py --source input.xlsx --output out.xlsx \
        --stock-sheet "Current Stock" --inbound-sheet "Expected Arrivals" \
        --status-filter "Committed,Arranged"
"""

import argparse
import math
from datetime import datetime, timedelta
from openpyxl import load_workbook, Workbook

ALLOWED_STATUSES = {'committed', 'arranged', 'confirmed', 'approved'}
EXCLUDED_STATUSES = {'tentative', 'pending', 'draft', 'proposed', 'cancelled'}

def detect_structure(wb):
    """Auto-detect sheet roles and key parameters."""
    sheets = wb.sheetnames
    
    # Detect sheet types by name
    stock_sheet = None
    inbound_sheet = None
    config_sheet = None
    
    for name in sheets:
        lower = name.lower()
        if any(x in lower for x in ['stock', 'current', 'inventory', 'position']):
            stock_sheet = name
        elif any(x in lower for x in ['arrival', 'inbound', 'expected', 'scheduled', 'shipment']):
            inbound_sheet = name
        elif any(x in lower for x in ['config', 'guide', 'pallet', 'parameter']):
            config_sheet = name
    
    return {
        'stock_sheet': stock_sheet or sheets[0],
        'inbound_sheet': inbound_sheet or (sheets[1] if len(sheets) > 1 else sheets[0]),
        'config_sheet': config_sheet or sheets[-1]
    }

def find_dates(sheet):
    """Find AsOfDate and HorizonEnd in first few rows."""
    candidates = {}
    for row_idx in range(1, min(4, sheet.max_row + 1)):
        for col_idx in range(1, min(6, sheet.max_column + 1)):
            cell = sheet.cell(row=row_idx, column=col_idx)
            if isinstance(cell.value, datetime):
                label_cell = sheet.cell(row=row_idx, column=col_idx-1) if col_idx > 1 else None
                label = str(label_cell.value).lower() if label_cell and label_cell.value else ''
                
                if 'end' in label or 'horizon' in label or 'finish' in label:
                    candidates['horizon_end'] = cell.value
                elif 'asof' in label or 'snapshot' in label or 'start' in label or 'date' in label:
                    candidates['as_of_date'] = cell.value
                elif 'horizon_end' not in candidates and cell.value > datetime(2025, 1, 1):
                    candidates.setdefault('as_of_date', cell.value)
    
    # Fallback: assume first two dates found
    if 'as_of_date' not in candidates or 'horizon_end' not in candidates:
        dates = []
        for row_idx in range(1, min(4, sheet.max_row + 1)):
            for col_idx in range(1, min(6, sheet.max_column + 1)):
                cell = sheet.cell(row=row_idx, column=col_idx)
                if isinstance(cell.value, datetime):
                    dates.append(cell.value)
        if len(dates) >= 2:
            candidates['as_of_date'] = dates[0]
            candidates['horizon_end'] = dates[1]
    
    return candidates

def parse_stock_sheet(sheet):
    """Extract stock data with flexible header detection."""
    dates = find_dates(sheet)
    
    # Find header row
    header_row = 1
    for row_idx in range(1, min(6, sheet.max_row + 1)):
        row_vals = [sheet.cell(row=row_idx, column=c).value for c in range(1, min(8, sheet.max_column + 1))]
        row_str = ' '.join(str(v).lower() for v in row_vals if v)
        if any(x in row_str for x in ['sku', 'item', 'code', 'on hand', 'stock', 'daily', 'rate', 'units']):
            header_row = row_idx
            break
    
    # Map columns
    headers = [sheet.cell(row=header_row, column=c).value for c in range(1, sheet.max_column + 1)]
    header_lower = [str(h).lower() if h else '' for h in headers]
    
    item_col = next((i for i, h in enumerate(header_lower) if any(x in h for x in ['sku', 'item', 'code'])), 0)
    onhand_col = next((i for i, h in enumerate(header_lower) if any(x in h for x in ['on hand', 'floor', 'stock', 'units'])), 1)
    daily_col = next((i for i, h in enumerate(header_lower) if any(x in h for x in ['daily', 'rate', 'sales', 'demand'])), 2)
    
    stock_data = []
    for row_idx in range(header_row + 1, sheet.max_row + 1):
        item = sheet.cell(row=row_idx, column=item_col+1).value
        if not item:
            continue
        onhand = sheet.cell(row=row_idx, column=onhand_col+1).value or 0
        daily = sheet.cell(row=row_idx, column=daily_col+1).value or 0
        stock_data.append({'item': item, 'on_floor': onhand, 'daily_sales': daily})
    
    return {
        'as_of_date': dates.get('as_of_date'),
        'horizon_end': dates.get('horizon_end'),
        'stock': stock_data
    }

def parse_inbound_sheet(sheet, horizon_end):
    """Extract inbound data with status filtering."""
    # Find header row
    header_row = 1
    for row_idx in range(1, min(4, sheet.max_row + 1)):
        row_vals = [sheet.cell(row=row_idx, column=c).value for c in range(1, min(8, sheet.max_column + 1))]
        row_str = ' '.join(str(v).lower() for v in row_vals if v)
        if any(x in row_str for x in ['sku', 'item', 'arrival', 'date', 'expected', 'cases', 'units']):
            header_row = row_idx
            break
    
    headers = [sheet.cell(row=header_row, column=c).value for c in range(1, sheet.max_column + 1)]
    header_lower = [str(h).lower() if h else '' for h in headers]
    
    item_col = next((i for i, h in enumerate(header_lower) if any(x in h for x in ['sku', 'item', 'code'])), 0)
    date_col = next((i for i, h in enumerate(header_lower) if any(x in h for x in ['arrival', 'date', 'expected'])), 1)
    cases_col = next((i for i, h in enumerate(header_lower) if any(x in h for x in ['cases', 'units', 'qty', 'quantity'])), 2)
    status_col = next((i for i, h in enumerate(header_lower) if any(x in h for x in ['status', 'dock', 'confirm'])), -1)
    
    inbounds = []
    for row_idx in range(header_row + 1, sheet.max_row + 1):
        item = sheet.cell(row=row_idx, column=item_col+1).value
        if not item:
            continue
        
        date_val = sheet.cell(row=row_idx, column=date_col+1).value
        cases = sheet.cell(row=row_idx, column=cases_col+1).value or 0
        status = sheet.cell(row=row_idx, column=status_col+1).value if status_col >= 0 else None
        
        if not isinstance(date_val, datetime):
            continue
            
        arr_date = date_val.date() if hasattr(date_val, 'date') else date_val
        
        # Status filtering
        if status:
            status_str = str(status).lower().strip()
            if status_str in EXCLUDED_STATUSES:
                continue
            if status_str not in ALLOWED_STATUSES and not any(a in status_str for a in ALLOWED_STATUSES):
                continue  # Unknown status, skip to be safe
        
        if arr_date > (horizon_end.date() if hasattr(horizon_end, 'date') else horizon_end):
            continue
            
        inbounds.append({'item': item, 'arrival_date': arr_date, 'cases': cases})
    
    return inbounds

def parse_config_sheet(sheet):
    """Extract cases per pallet from config sheet."""
    for row_idx in range(1, min(5, sheet.max_row + 1)):
        for col_idx in range(1, min(5, sheet.max_column + 1)):
            cell = sheet.cell(row=row_idx, column=col_idx)
            if isinstance(cell.value, (int, float)) and 10 <= cell.value <= 1000:
                return int(cell.value)
    return 60  # Default

def calculate_metrics(stock_data, inbounds, as_of_date, horizon_end, cases_per_pallet):
    """Calculate planning metrics for each item."""
    planning_days = (horizon_end.date() if hasattr(horizon_end, 'date') else horizon_end) - \
                    (as_of_date.date() if hasattr(as_of_date, 'date') else as_of_date)
    planning_days = planning_days.days
    
    results = []
    
    # Index inbounds by item
    inbound_by_item = {}
    for ib in inbounds:
        inbound_by_item.setdefault(ib['item'], []).append(ib)
    
    for stock in stock_data:
        item = stock['item']
        on_floor = stock['on_floor']
        daily_sales = stock['daily_sales'] or 1  # Avoid div by zero
        
        current_doh = on_floor / daily_sales
        oos_date = (as_of_date.date() if hasattr(as_of_date, 'date') else as_of_date) + \
                   timedelta(days=int(current_doh))
        
        item_inbounds = inbound_by_item.get(item, [])
        inbound_cases = sum(ib['cases'] for ib in item_inbounds)
        
        delivered_doh = (on_floor + inbound_cases) / daily_sales
        remaining_demand = daily_sales * planning_days
        additional_needed = max(0, remaining_demand - on_floor - inbound_cases)
        pallets = math.ceil(additional_needed / cases_per_pallet) if additional_needed > 0 else 0
        
        # Check if any inbound arrives before OOS date
        earlier_required = True
        for ib in item_inbounds:
            if ib['arrival_date'] < oos_date:
                earlier_required = False
                break
        
        results.append({
            'item': item,
            'on_floor': on_floor,
            'daily_sales': stock['daily_sales'],
            'current_doh': round(current_doh, 2),
            'oos_date': oos_date.isoformat(),
            'inbound_cases': inbound_cases,
            'delivered_doh': round(delivered_doh, 2),
            'remaining_demand': remaining_demand,
            'additional_needed': int(additional_needed),
            'pallets': pallets,
            'delivery_date': oos_date.isoformat() if pallets > 0 else None,
            'earlier_required': earlier_required if pallets > 0 else False
        })
    
    return results, planning_days

def create_output_workbook(results, planning_days, as_of_date, horizon_end, output_path):
    """Create output workbook with SKU_Coverage and Pallet_Gap_List sheets."""
    wb = Workbook()
    
    # SKU_Coverage / Load_Detail sheet
    detail = wb.active
    detail.title = 'SKU_Coverage'
    
    # Metadata
    detail['A1'] = 'Field'
    detail['B1'] = 'Value'
    detail['A2'] = 'AsOfDate'
    detail['B2'] = as_of_date.date() if hasattr(as_of_date, 'date') else as_of_date
    detail['A3'] = 'HorizonEnd'
    detail['B3'] = horizon_end.date() if hasattr(horizon_end, 'date') else horizon_end
    detail['A4'] = 'PlanningDays'
    detail['B4'] = planning_days
    
    # Headers
    headers = ['SKU', 'Units_On_Hand', 'Daily_Rate_Units_Per_Day', 
               'Current_Days_On_Hand', 'Projected_OOS_Date', 'Inbound_Units_By_Horizon',
               'Delivered_Days_On_Hand', 'Remaining_Demand_Units', 'Additional_Units_Needed',
               'Pallets_Required', 'Required_Delivery_Date', 'Earlier_Delivery_Required']
    for col, h in enumerate(headers, 1):
        detail.cell(row=6, column=col, value=h)
    
    # Data
    for row_idx, r in enumerate(results, 7):
        detail.cell(row=row_idx, column=1, value=r['item'])
        detail.cell(row=row_idx, column=2, value=r['on_floor'])
        detail.cell(row=row_idx, column=3, value=r['daily_sales'])
        detail.cell(row=row_idx, column=4, value=r['current_doh'])
        detail.cell(row=row_idx, column=5, value=r['oos_date'])
        detail.cell(row=row_idx, column=6, value=r['inbound_cases'])
        detail.cell(row=row_idx, column=7, value=r['delivered_doh'])
        detail.cell(row=row_idx, column=8, value=r['remaining_demand'])
        detail.cell(row=row_idx, column=9, value=r['additional_needed'])
        detail.cell(row=row_idx, column=10, value=r['pallets'])
        detail.cell(row=row_idx, column=11, value=r['delivery_date'])
        detail.cell(row=row_idx, column=12, value=r['earlier_required'])
    
    # Pallet_Gap_List / Load_Action_Summary sheet
    summary = wb.create_sheet('Pallet_Gap_List')
    summary.append(['SKU', 'Required_Delivery_Date', 'Pallets_Required', 
                    'Additional_Units_Needed', 'Earlier_Delivery_Required'])
    
    for r in results:
        if r['pallets'] > 0:
            summary.append([r['item'], r['delivery_date'], r['pallets'],
                          r['additional_needed'], r['earlier_required']])
    
    wb.save(output_path)
    return output_path

def main():
    parser = argparse.ArgumentParser(description='Generate load planning workbook')
    parser.add_argument('--source', required=True, help='Path to source Excel file')
    parser.add_argument('--output', required=True, help='Path for output workbook')
    parser.add_argument('--stock-sheet', help='Override stock sheet name')
    parser.add_argument('--inbound-sheet', help='Override inbound sheet name')
    parser.add_argument('--config-sheet', help='Override config sheet name')
    args = parser.parse_args()
    
    wb = load_workbook(args.source)
    structure = detect_structure(wb)
    
    # Allow overrides
    stock_sheet = args.stock_sheet or structure['stock_sheet']
    inbound_sheet = args.inbound_sheet or structure['inbound_sheet']
    config_sheet = args.config_sheet or structure['config_sheet']
    
    # Parse data
    stock_info = parse_stock_sheet(wb[stock_sheet])
    as_of_date = stock_info['as_of_date']
    horizon_end = stock_info['horizon_end']
    stock_data = stock_info['stock']
    
    inbounds = parse_inbound_sheet(wb[inbound_sheet], horizon_end)
    cases_per_pallet = parse_config_sheet(wb[config_sheet])
    
    # Calculate and output
    results, planning_days = calculate_metrics(stock_data, inbounds, as_of_date, horizon_end, cases_per_pallet)
    create_output_workbook(results, planning_days, as_of_date, horizon_end, args.output)
    
    # Summary
    gaps = [r for r in results if r['pallets'] > 0]
    print(f'Workbook created: {args.output}')
    print(f'Planning period: {as_of_date.date() if hasattr(as_of_date, "date") else as_of_date} to {horizon_end.date() if hasattr(horizon_end, "date") else horizon_end} ({planning_days} days)')
    print(f'Items with pallet gaps: {len(gaps)}')
    for g in gaps:
        print(f"  {g['item']}: {g['pallets']} pallets needed by {g['delivery_date']}")

if __name__ == '__main__':
    main()