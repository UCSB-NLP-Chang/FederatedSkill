#!/usr/bin/env python3
"""
Template workbook preservation helper for OCR data extraction.

Preserves existing sheets (cover, instructions) while updating data sheets.
"""

import openpyxl
from openpyxl import load_workbook


def preserve_and_update(template_path, output_path, sheet_updates):
    """
    Load template workbook, preserve all sheets, update specified sheets.
    
    Args:
        template_path: Path to template Excel file
        output_path: Path to save updated workbook
        sheet_updates: Dict mapping sheet_name -> (headers, records)
                      records is list of dicts with keys matching headers
    
    Example:
        sheet_updates = {
            'bills': (
                ['scan_name', 'bill_date', 'amount_due'],
                [
                    {'scan_name': 'bill_001.jpg', 'bill_date': '2024-01-13', 'amount_due': '85.40'},
                    ...
                ]
            )
        }
    """
    wb = load_workbook(template_path)
    
    for sheet_name, (headers, records) in sheet_updates.items():
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not in template. Available: {wb.sheetnames}")
        
        ws = wb[sheet_name]
        
        # Clear existing data (keep header row)
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.value = None
        
        # Write new data
        for idx, record in enumerate(records, start=2):
            for col_idx, header in enumerate(headers, start=1):
                value = record.get(header, '')
                # Convert None to empty string
                if value is None:
                    value = ''
                ws.cell(row=idx, column=col_idx, value=value)
    
    wb.save(output_path)
    print(f"Saved to {output_path}")
    return output_path


def update_sheet_only(wb_path, sheet_name, headers, records, output_path=None):
    """
    Update single sheet in workbook, preserve all others.
    
    Args:
        wb_path: Path to workbook
        sheet_name: Name of sheet to update
        headers: List of column headers (must match existing if sheet has headers)
        records: List of dicts with data
        output_path: Where to save (defaults to wb_path)
    """
    output_path = output_path or wb_path
    
    wb = load_workbook(wb_path)
    
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found in {wb.sheetnames}")
    
    ws = wb[sheet_name]
    
    # Clear data rows
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.value = None
    
    # Write data
    for idx, record in enumerate(records, start=2):
        for col_idx, header in enumerate(headers, start=1):
            value = record.get(header, '')
            ws.cell(row=idx, column=col_idx, value=value if value is not None else '')
    
    wb.save(output_path)
    return output_path


if __name__ == '__main__':
    import argparse
    
    parser = argparse.ArgumentParser(description='Preserve template workbook while updating data')
    parser.add_argument('template', help='Template workbook path')
    parser.add_argument('output', help='Output path')
    parser.add_argument('--sheet', default='bills', help='Sheet to update')
    
    args = parser.parse_args()
    
    # Example usage with dummy data
    example_updates = {
        args.sheet: (
            ['scan_name', 'bill_date', 'amount_due'],
            [
                {'scan_name': 'example_001.jpg', 'bill_date': '2024-01-01', 'amount_due': '100.00'},
            ]
        )
    }
    
    preserve_and_update(args.template, args.output, example_updates)
