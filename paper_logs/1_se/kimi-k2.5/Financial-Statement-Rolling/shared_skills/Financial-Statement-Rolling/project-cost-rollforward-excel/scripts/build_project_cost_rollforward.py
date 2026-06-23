#!/usr/bin/env python3
"""
Build capitalized project cost rollforward workbook.

Usage:
    python build_project_cost_rollforward.py \
        --input project_cost_rollforward.json \
        --overrides schedule_overrides.csv \
        --gl gl_balances.json \
        --output output.xlsx \
        --company "Company Name" \
        --period "September 2025"
"""

import json
import csv
import argparse
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font


def parse_args():
    parser = argparse.ArgumentParser()
    parser.add_argument('--input', required=True)
    parser.add_argument('--overrides', required=True)
    parser.add_argument('--gl', required=True)
    parser.add_argument('--output', required=True)
    parser.add_argument('--company', required=True)
    parser.add_argument('--period', required=True)
    parser.add_argument('--period-totals-row', type=int, default=9)
    parser.add_argument('--ending-balance-row', type=int, default=10)
    parser.add_argument('--variance-row', type=int, default=11)
    parser.add_argument('--gl-balance-row', type=int, default=12)
    return parser.parse_args()


def load_overrides(path):
    """Load override CSV into dict keyed by row_id."""
    overrides = {}
    with open(path, 'r') as f:
        reader = csv.DictReader(f)
        for row in reader:
            row_id = row['row_id']
            overrides[row_id] = {k: v for k, v in row.items() if v}
    return overrides


def filter_items(items):
    """Keep highest revision active item per row_id."""
    by_row_id = defaultdict(list)
    for item in items:
        by_row_id[item['row_id']].append(item)
    
    result = []
    for row_id, versions in by_row_id.items():
        active = [v for v in versions if v.get('active', True)]
        if active:
            highest = max(active, key=lambda x: x['revision'])
            result.append(highest)
    return result


def apply_override(item, override):
    """Apply override values to item."""
    if not override:
        return item
    
    result = dict(item)
    
    # Notes override
    if override.get('notes_override'):
        result['memo'] = override['notes_override']
    
    # Copy months for modification
    result['months'] = {k: dict(v) for k, v in item['months'].items()}
    
    # Value overrides
    for month in ['jun', 'jul', 'aug', 'sep', 'oct', 'nov']:
        for field in ['adds', 'release', 'ending_balance']:
            key = f"{month}_{field}"
            if override.get(key):
                result['months'][month][field] = float(override[key])
    
    return result


def build_detail_sheet(wb, account, overrides, gl_values, args):
    """Build a detail sheet for one account."""
    ws = wb.create_sheet(account['sheet_name'])
    
    # Collect all items from all groups
    all_items = []
    for group in account.get('groups', []):
        all_items.extend(group.get('items', []))
    
    # Filter and apply overrides
    filtered = filter_items(all_items)
    processed = [apply_override(item, overrides.get(item['row_id'])) for item in filtered]
    
    # Sort by vendor name
    processed.sort(key=lambda x: x['vendor_name'])
    
    # Headers
    headers = ['Vendor', 'Beginning Balance', 'Jun Cap Adds', 'Jun Amortization', 
               'Jun Ending Balance', 'Jul Cap Adds', 'Jul Amortization', 'Jul Ending Balance',
               'Aug Cap Adds', 'Aug Amortization', 'Aug Ending Balance',
               'Sep Cap Adds', 'Sep Amortization', 'Sep Ending Balance',
               'Useful Life Months', 'Notes']
    ws.append(headers)
    
    # Data rows (starting row 6, so index 0 = row 6)
    data_start = 6
    for idx, item in enumerate(processed):
        row = [
            item['vendor_name'],
            item['opening_balance'],
            item['months']['jun']['adds'],
            item['months']['jun']['release'],
            item['months']['jun']['ending_balance'],
            item['months']['jul']['adds'],
            item['months']['jul']['release'],
            item['months']['jul']['ending_balance'],
            item['months']['aug']['adds'],
            item['months']['aug']['release'],
            item['months']['aug']['ending_balance'],
            item['months']['sep']['adds'],
            item['months']['sep']['release'],
            item['months']['sep']['ending_balance'],
            item['useful_life_months'],
            item['memo']
        ]
        ws.append(row)
    
    data_end = data_start + len(processed) - 1
    
    # Control rows
    pr = args.period_totals_row
    er = args.ending_balance_row
    vr = args.variance_row
    gr = args.gl_balance_row
    
    # Period Totals
    ws.cell(row=pr, column=1, value='Period Totals')
    for col in range(2, 15):  # B through N
        cell = ws.cell(row=pr, column=col, value=f'=SUM({cell_ref(col)}{data_start}:{cell_ref(col)}{data_end})')
    ws.cell(row=pr, column=15, value=f'=C{pr}+F{pr}+I{pr}+L{pr}')  # O column
    
    # Ending Balance
    ws.cell(row=er, column=1, value='Ending Balance')
    ws.cell(row=er, column=5, value=f'=B{er}+C{er}-D{er}')   # E: Jun
    ws.cell(row=er, column=8, value=f'=E{er}+F{er}-G{er}')   # H: Jul
    ws.cell(row=er, column=11, value=f'=H{er}+I{er}-J{er}')  # K: Aug
    ws.cell(row=er, column=14, value=f'=K{er}+L{er}-M{er}')  # N: Sep
    ws.cell(row=er, column=15, value=f'=D{er}+G{er}+J{er}+M{er}')  # O: total releases
    
    # Variance
    ws.cell(row=vr, column=1, value='Variance')
    ws.cell(row=vr, column=15, value=f'=O{gr}-N{gr}')
    
    # GL Balance
    ws.cell(row=gr, column=1, value='GL Balance')
    ws.cell(row=gr, column=5, value=gl_values.get('jun', 0))
    ws.cell(row=gr, column=8, value=gl_values.get('jul', 0))
    ws.cell(row=gr, column=11, value=gl_values.get('aug', 0))
    ws.cell(row=gr, column=14, value=gl_values.get('sep', 0))
    ws.cell(row=gr, column=15, value=f'=O{pr}-O{er}')
    
    return processed


def cell_ref(col_num):
    """Convert column number to letter (1=A, 2=B, etc.)."""
    result = ''
    while col_num > 0:
        col_num, remainder = divmod(col_num - 1, 26)
        result = chr(65 + remainder) + result
    return result


def build_summary(wb, accounts, company, period, args):
    """Build summary sheet with cross-references."""
    ws = wb.create_sheet('Project Cost Summary', 0)
    
    ws['A1'] = company
    ws['A2'] = 'Capitalized Project Cost Rollforward'
    ws['A3'] = f'Period Ending {period}'
    
    row = 6
    for account in accounts:
        ws.cell(row=row, column=1, value=account['sheet_name'])
        
        # Period Total Amortization
        ws.cell(row=row+1, column=1, value='Period Total Amortization:')
        ws.cell(row=row+1, column=2, value=f"='{account['sheet_name']}'!O{args.period_totals_row}")
        
        # Ending Balance
        ws.cell(row=row+2, column=1, value='Ending Balance:')
        ws.cell(row=row+2, column=2, value=f"='{account['sheet_name']}'!O{args.ending_balance_row}")
        
        # GL Balance
        ws.cell(row=row+3, column=1, value='GL Balance:')
        ws.cell(row=row+3, column=2, value=f"='{account['sheet_name']}'!O{args.gl_balance_row}")
        
        row += 5
    
    # Total
    gl_rows = [7 + i*5 for i in range(len(accounts))]  # Adjust based on spacing
    ws.cell(row=row, column=1, value='Total GL Balance:')
    # Build sum formula
    terms = [f'B{r+2}' for r in range(0, len(accounts)*5, 5)]  # Simplified
    ws.cell(row=row, column=2, value=f'=B9+B14')  # Adjust for actual layout


def main():
    args = parse_args()
    
    with open(args.input) as f:
        data = json.load(f)
    
    overrides = load_overrides(args.overrides)
    
    with open(args.gl) as f:
        gl_data = json.load(f)
    
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    accounts = data['accounts']
    
    # Build detail sheets first
    for account in accounts:
        gl_key = account['sheet_name'].lower().replace(' ', '_').replace('#', '')
        gl_values = gl_data.get(gl_key, {})
        build_detail_sheet(wb, account, overrides, gl_values, args)
    
    # Build summary
    build_summary(wb, accounts, args.company, args.period, args)
    
    wb.save(args.output)
    print(f'Workbook saved to {args.output}')


if __name__ == '__main__':
    main()
