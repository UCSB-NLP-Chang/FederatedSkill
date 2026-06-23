#!/usr/bin/env python3
"""
Verify Excel workbook structure against common requirements.
Usage: python3 verify_workbook.py <workbook_path> [--sheet SHEET_NAME]
"""

import argparse
import openpyxl
from datetime import datetime

def verify_sheet(ws, checks):
    """Run verification checks on a worksheet."""
    results = []
    
    if 'headers' in checks:
        for cell_ref, expected in checks['headers'].items():
            actual = ws[cell_ref].value
            results.append(f"  {cell_ref}: {actual} {'✓' if actual == expected else f'✗ (expected {expected})'}")
    
    if 'date_range' in checks:
        start = ws[checks['date_range']['start_cell']].value
        results.append(f"  First date ({checks['date_range']['start_cell']}): {start}")
        
    if 'column_types' in checks:
        row = checks['column_types'].get('sample_row', 10)
        for col, expected_type in checks['column_types']['cols'].items():
            cell = ws.cell(row=row, column=col)
            val = cell.value
            is_formula = isinstance(val, str) and val.startswith('=')
            is_number = isinstance(val, (int, float))
            
            if expected_type == 'formula':
                ok = is_formula
            elif expected_type == 'constant':
                ok = is_number
            else:
                ok = True
                
            results.append(f"  Col {col} row {row}: {type(val).__name__} {'✓' if ok else '✗'}")
    
    return results

def main():
    parser = argparse.ArgumentParser(description='Verify Excel workbook structure')
    parser.add_argument('workbook', help='Path to .xlsx file')
    parser.add_argument('--sheet', help='Specific sheet to verify')
    args = parser.parse_args()
    
    wb = openpyxl.load_workbook(args.workbook, data_only=False)
    
    print(f"Workbook: {args.workbook}")
    print(f"Sheets: {wb.sheetnames}")
    
    sheets_to_check = [args.sheet] if args.sheet else wb.sheetnames
    
    for sheet_name in sheets_to_check:
        if sheet_name not in wb.sheetnames:
            print(f"\n✗ Sheet '{sheet_name}' not found")
            continue
            
        ws = wb[sheet_name]
        print(f"\n=== {sheet_name} ===")
        print(f"Dimensions: {ws.dimensions}")
        
        # Basic structure check
        print(f"Max row: {ws.max_row}, Max col: {ws.max_column}")

if __name__ == '__main__':
    main()