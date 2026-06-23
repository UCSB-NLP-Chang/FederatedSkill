#!/usr/bin/env python3
"""
Payment request validation pipeline.
Cross-references claims against person registry and approvals.
Supports optional code crosswalk (external→internal codes).

Usage:
    python3 validate_claims.py <claims_pdf> <registry_xlsx> <approvals_csv> <output_json> [crosswalk_csv]

Field names are configurable via HEADERS dict below.
"""

import json
import sys
import re
from difflib import SequenceMatcher

# CONFIGURE: Adapt these field names to your data format
HEADERS = {
    # Registry (Excel) fields
    'registry_name_col': 'name',           # or 'employee_name', 'clinician_name', 'speaker_name'
    'registry_id_col': 'person_id',        # or 'employee_id', 'clinician_id'
    'registry_account_col': 'account',     # or 'bank_account', 'payout_account'
    
    # Approvals (CSV) fields  
    'approval_code_col': 'approval_code',  # or 'trip_id', 'shift_code_internal', 'engagement_id'
    'approval_amount_col': 'approved_amount',  # or 'approved_pay', 'honorarium_fee'
    'approval_person_id_col': 'person_id',     # or 'employee_id', 'clinician_id'
    
    # Crosswalk (CSV) fields
    'crosswalk_external_col': 'external_code', # or 'shift_ref'
    'crosswalk_internal_col': 'internal_code', # or 'shift_code_internal'
    
    # Claim extraction patterns (regex groups)
    'claim_name_pattern': r'(?:Name|Employee|Clinician|Speaker):\s*([^\n]+)',
    'claim_amount_pattern': r'(?:Amount|Claimed|Pay|Fee):\s*\$?([\d,]+\.?\d*)',
    'claim_account_pattern': r'(?:Account|Bank|Payout):\s*([^\n\s]+)',
    'claim_code_pattern': r'(?:Code|ID|Shift|Trip|Ref):\s*([^\n\s]+)',
}


def fuzzy_match(name, candidates, threshold=0.80):
    """Find best matching name from candidates."""
    # Normalize: lowercase, strip whitespace, remove common titles
    name_clean = re.sub(r'^(dr\.|prof\.|mr\.|ms\.|mrs\.)\s*', '', name.lower().strip())
    
    best_match = None
    best_score = 0
    
    for candidate in candidates:
        cand_clean = re.sub(r'^(dr\.|prof\.|mr\.|ms\.|mrs\.)\s*', '', candidate.lower().strip())
        score = SequenceMatcher(None, name_clean, cand_clean).ratio()
        if score > best_score and score >= threshold:
            best_score = score
            best_match = candidate
    
    return best_match


def load_registry(xlsx_path):
    """Load person registry from Excel."""
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    
    registry = {}
    headers = [cell.value for cell in ws[1]]
    
    name_col = HEADERS['registry_name_col']
    id_col = HEADERS['registry_id_col']
    acct_col = HEADERS['registry_account_col']
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        record = dict(zip(headers, row))
        
        # Try configured column name, then common alternatives
        name = record.get(name_col)
        if name is None:
            for alt in ['name', 'employee_name', 'clinician_name', 'speaker_name', 'full_name']:
                if alt in record:
                    name = record[alt]
                    break
        
        if name:
            person_id = record.get(id_col) or record.get('employee_id') or record.get('clinician_id') or record.get('id')
            account = record.get(acct_col) or record.get('bank_account') or record.get('payout_account') or record.get('account')
            registry[name] = {
                'person_id': person_id,
                'account': account
            }
    
    return registry


def load_approvals(csv_path):
    """Load approvals from CSV."""
    import csv
    approvals = {}
    
    with open(csv_path, 'r') as f:
        reader = csv.DictReader(f)
        for row in reader:
            # Try configured column, then common alternatives
            code = row.get(HEADERS['approval_code_col'])
            if code is None:
                for alt in ['trip_id', 'shift_code', 'shift_code_internal', 'engagement_id', 'approval_code', 'code']:
                    if alt in row:
                        code = row[alt]
                        break
            
            if code:
                amount_str = row.get(HEADERS['approval_amount_col']) or row.get('approved_amount') or row.get('approved_pay') or row.get('amount') or '0'
                person_id = row.get(HEADERS['approval_person_id_col']) or row.get('employee_id') or row.get('clinician_id') or row.get('person_id')
                approvals[code] = {
                    'approved_amount': float(str(amount_str).replace(',', '')),
                    'person_id': person_id
                }
    
    return approvals


def load_crosswalk(csv_path):
    """Load external→internal code mapping from CSV."""
    import csv
    crosswalk = {}
    
    with open(csv_path, 'r') as f:
        reader = csv.DictReader(f)
        for row in reader:
            # Try configured columns, then common alternatives
            external = row.get(HEADERS['crosswalk_external_col'])
            internal = row.get(HEADERS['crosswalk_internal_col'])
            
            if external is None:
                for alt in ['shift_ref', 'external_code', 'external', 'code']:
                    if alt in row:
                        external = row[alt]
                        break
            if internal is None:
                for alt in ['shift_code_internal', 'internal_code', 'internal', 'trip_id']:
                    if alt in row:
                        internal = row[alt]
                        break
            
            if external and internal:
                crosswalk[external] = internal
    
    return crosswalk


def parse_claims(pdf_text):
    """Extract claims from PDF text."""
    claims = []
    
    # Split by pages
    pages = re.split(r'Page\s+(\d+)', pdf_text)
    
    for i in range(1, len(pages), 2):
        page_num = int(pages[i])
        content = pages[i+1]
        
        # Extract fields using configured patterns
        name_match = re.search(HEADERS['claim_name_pattern'], content, re.I)
        amount_match = re.search(HEADERS['claim_amount_pattern'], content, re.I)
        account_match = re.search(HEADERS['claim_account_pattern'], content, re.I)
        code_match = re.search(HEADERS['claim_code_pattern'], content, re.I)
        
        if name_match:
            amount_str = amount_match.group(1).replace(',', '') if amount_match else None
            claims.append({
                'page': page_num,
                'person_name': name_match.group(1).strip(),
                'requested_amount': float(amount_str) if amount_str else None,
                'payment_account': account_match.group(1).strip() if account_match else None,
                'approval_code': code_match.group(1).strip() if code_match else None
            })
    
    return claims


def validate_claim(claim, registry, approvals, crosswalk=None):
    """Validate a single claim against all dimensions."""
    
    # Fuzzy match person name
    matched_name = fuzzy_match(claim['person_name'], registry.keys())
    
    if not matched_name:
        return {
            'claim_page_number': claim['page'],
            'person_name': claim['person_name'],
            'requested_amount': claim['requested_amount'],
            'payment_account': claim['payment_account'],
            'approval_code': claim['approval_code'],
            'reason': 'Unknown Person'
        }
    
    person = registry[matched_name]
    
    # Check payment account
    if claim['payment_account'] and claim['payment_account'] != person['account']:
        return {
            'claim_page_number': claim['page'],
            'person_name': claim['person_name'],
            'requested_amount': claim['requested_amount'],
            'payment_account': claim['payment_account'],
            'approval_code': claim['approval_code'],
            'reason': 'Account Mismatch'
        }
    
    # Resolve code via crosswalk if provided
    lookup_code = claim['approval_code']
    if crosswalk and lookup_code in crosswalk:
        lookup_code = crosswalk[lookup_code]
    
    # Check approval exists
    if lookup_code not in approvals:
        return {
            'claim_page_number': claim['page'],
            'person_name': claim['person_name'],
            'requested_amount': claim['requested_amount'],
            'payment_account': claim['payment_account'],
            'approval_code': claim['approval_code'],
            'reason': 'Invalid Approval Code'
        }
    
    approval = approvals[lookup_code]
    
    # Check amount matches
    if claim['requested_amount'] and abs(claim['requested_amount'] - approval['approved_amount']) > 0.01:
        return {
            'claim_page_number': claim['page'],
            'person_name': claim['person_name'],
            'requested_amount': claim['requested_amount'],
            'payment_account': claim['payment_account'],
            'approval_code': claim['approval_code'],
            'reason': 'Amount Mismatch'
        }
    
    # Check ownership
    if approval['person_id'] and person['person_id'] and approval['person_id'] != person['person_id']:
        return {
            'claim_page_number': claim['page'],
            'person_name': claim['person_name'],
            'requested_amount': claim['requested_amount'],
            'payment_account': claim['payment_account'],
            'approval_code': claim['approval_code'],
            'reason': 'Ownership Mismatch'
        }
    
    return None  # No alert - claim is valid


def main():
    if len(sys.argv) < 5:
        print("Usage: python3 validate_claims.py <claims_pdf> <registry_xlsx> <approvals_csv> <output_json> [crosswalk_csv]")
        sys.exit(1)
    
    claims_pdf = sys.argv[1]
    registry_xlsx = sys.argv[2]
    approvals_csv = sys.argv[3]
    output_json = sys.argv[4]
    crosswalk_csv = sys.argv[5] if len(sys.argv) > 5 else None
    
    # Load reference data
    registry = load_registry(registry_xlsx)
    approvals = load_approvals(approvals_csv)
    crosswalk = load_crosswalk(crosswalk_csv) if crosswalk_csv else None
    
    # Read PDF (assumes text extraction already done, or use pdfplumber if available)
    try:
        import pdfplumber
        with pdfplumber.open(claims_pdf) as pdf:
            pdf_text = ""
            for i, page in enumerate(pdf.pages, 1):
                text = page.extract_text() or ""
                pdf_text += f"\nPage {i}\n{text}\n"
    except ImportError:
        # Fallback: assume .txt version exists
        txt_path = claims_pdf.replace('.pdf', '.txt')
        with open(txt_path, 'r') as f:
            pdf_text = f.read()
    
    # Parse and validate
    claims = parse_claims(pdf_text)
    alerts = []
    
    for claim in claims:
        alert = validate_claim(claim, registry, approvals, crosswalk)
        if alert:
            alerts.append(alert)
    
    # Write output
    with open(output_json, 'w') as f:
        json.dump(alerts, f, indent=2)
    
    print(f"Validation complete. {len(alerts)} alerts written to {output_json}")


if __name__ == '__main__':
    main()
