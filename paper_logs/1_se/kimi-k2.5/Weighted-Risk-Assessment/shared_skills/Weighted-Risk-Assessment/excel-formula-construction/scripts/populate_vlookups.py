#!/usr/bin/env python3
"""
Populate VLOOKUP + MATCH formulas for multi-year data grids.

Usage: python3 populate_vlookups.py <input.xlsx> <output.xlsx> \
         --task-sheet Task --data-sheet Data \
         --lookup-col D --data-start-row 21 --data-end-row 38 \
         --year-headers-row 20 --year-data-start-col H --year-data-end-col L
"""

import openpyxl
import argparse
from openpyxl.utils import get_column_letter


def populate_vlookups(wb, task_sheet, data_sheet, config):
    """Populate VLOOKUP formulas for a multi-year data grid."""
    ws_task = wb[task_sheet]
    
    # Build absolute reference strings
    data_range = f"{data_sheet}!${config['lookup_col']}${config['data_start_row']}:${config['year_data_end_col']}${config['data_end_row']}"
    header_range = f"{data_sheet}!${config['year_data_start_col']}${config['year_headers_row']}:${config['year_data_end_col']}${config['year_headers_row']}"
    
    # Determine year columns
    start_col_idx = openpyxl.utils.column_index_from_string(config['year_data_start_col'])
    end_col_idx = openpyxl.utils.column_index_from_string(config['year_data_end_col'])
    year_cols = [get_column_letter(c) for c in range(start_col_idx, end_col_idx + 1)]
    
    formulas_written = 0
    
    for row in config['target_rows']:
        lookup_cell = f"{config['lookup_col']}{row}"
        for year_col in year_cols:
            year_header_cell = f"{year_col}{config['year_header_row']}"
            formula = f"=VLOOKUP({lookup_cell},{data_range},MATCH({year_header_cell},{header_range},0)+1,FALSE)"
            ws_task[f"{year_col}{row}"].value = formula
            formulas_written += 1
    
    return formulas_written


def populate_calculated_column(wb, sheet, config):
    """Populate calculated column formulas referencing looked-up values."""
    ws = wb[sheet]
    formulas_written = 0
    
    for row in config['target_rows']:
        for year_col in config['year_cols']:
            # Formula pattern: (success - failed) / capacity * 100
            success_cell = f"{year_col}{config['success_row_map'][row]}"
            failed_cell = f"{year_col}{config['failed_row_map'][row]}"
            capacity_cell = f"{year_col}{config['capacity_row_map'][row]}"
            
            formula = f"=({success_cell}-{failed_cell})/{capacity_cell}*100"
            ws[f"{year_col}{row}"].value = formula
            formulas_written += 1
    
    return formulas_written


def populate_statistics(wb, sheet, config):
    """Populate MIN, MAX, MEDIAN, AVERAGE, PERCENTILE formulas."""
    ws = wb[sheet]
    formulas_written = 0
    
    stats_config = [
        ('MIN', 'MIN', None),
        ('MAX', 'MAX', None),
        ('MEDIAN', 'MEDIAN', None),
        ('AVERAGE', 'AVERAGE', None),
        ('PERCENTILE.25', 'PERCENTILE.INC', 0.25),
        ('PERCENTILE.75', 'PERCENTILE.INC', 0.75),
    ]
    
    for stat_name, func, arg in stats_config:
        row = config['stat_rows'][stat_name]
        for year_col in config['year_cols']:
            data_range = f"{year_col}{config['data_start_row']}:{year_col}{config['data_end_row']}"
            if arg is not None:
                formula = f"={func}({data_range},{arg})"
            else:
                formula = f"={func}({data_range})"
            ws[f"{year_col}{row}"].value = formula
            formulas_written += 1
    
    return formulas_written


def populate_weighted_mean(wb, sheet, config):
    """Populate SUMPRODUCT weighted mean formulas."""
    ws = wb[sheet]
    formulas_written = 0
    
    for year_col in config['year_cols']:
        values_range = f"{year_col}{config['values_start_row']}:{year_col}{config['values_end_row']}"
        weights_range = f"{year_col}{config['weights_start_row']}:{year_col}{config['weights_end_row']}"
        
        formula = f"=SUMPRODUCT({values_range},{weights_range})/SUM({weights_range})"
        ws[f"{year_col}{config['output_row']}"].value = formula
        formulas_written += 1
    
    return formulas_written


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Populate Excel formulas')
    parser.add_argument('input', help='Input workbook path')
    parser.add_argument('output', help='Output workbook path')
    parser.add_argument('--task-sheet', default='Task')
    parser.add_argument('--data-sheet', default='Data')
    
    args = parser.parse_args()
    
    wb = openpyxl.load_workbook(args.input, data_only=False)
    
    # Example configuration - adjust for specific task
    print("This script requires task-specific configuration.")
    print("Copy and modify for your specific workbook structure.")
