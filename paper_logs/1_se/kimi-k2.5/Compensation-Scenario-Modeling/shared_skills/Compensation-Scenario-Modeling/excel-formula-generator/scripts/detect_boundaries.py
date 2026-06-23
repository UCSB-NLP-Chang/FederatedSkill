#!/usr/bin/env python3
"""
Detect actual data boundaries in Excel source files.
Use this instead of hardcoding row ranges which break when source structure changes.

Usage: python3 scripts/detect_boundaries.py <workbook_path> <sheet_name> [options]
"""

import argparse
import openpyxl
import re

def detect_data_rows(ws, name_col=2, value_col=3, start_row=1, 
                     header_pattern=None, exclude_patterns=None):
    """
    Find contiguous data rows with name/value pattern.
    
    Args:
        name_col: Column containing parameter names
        value_col: Column containing values (int, float, or convertible)
        start_row: Row to start scanning from
        header_pattern: Regex to identify header rows to skip
        exclude_patterns: List of regex patterns for rows to exclude (e.g., section headers)
    
    Returns:
        List of (row_num, name, value) tuples
    """
    if exclude_patterns is None:
        exclude_patterns = [r'^---', r'^\s*$', r'Parameter', r'^#']
    
    compiled_excludes = [re.compile(p) for p in exclude_patterns]
    compiled_header = re.compile(header_pattern) if header_pattern else None
    
    rows = []
    for row in range(start_row, ws.max_row + 1):
        name = ws.cell(row=row, column=name_col).value
        value = ws.cell(row=row, column=value_col).value
        
        # Skip empty rows
        if not name:
            continue
        
        name_str = str(name).strip()
        
        # Skip section headers and excluded patterns
        if any(p.match(name_str) for p in compiled_excludes):
            continue
        
        # Skip if no numeric value (indicates header/description row)
        if value is None:
            continue
        try:
            float(value)
        except (ValueError, TypeError):
            continue
        
        rows.append((row, name_str, value))
    
    return rows


def detect_section_boundaries(ws, section_marker_col=2):
    """
    Detect section boundaries marked by '--- Section Name ---' patterns.
    Returns dict of section_name -> (start_row, end_row)
    """
    sections = {}
    current_section = None
    section_start = None
    
    for row in range(1, ws.max_row + 1):
        cell_value = ws.cell(row=row, column=section_marker_col).value
        if not cell_value:
            continue
        
        value_str = str(cell_value).strip()
        
        # Detect section header: '--- Section Name ---'
        match = re.match(r'^---\s*(.+?)\s*---$', value_str)
        if match:
            if current_section and section_start:
                sections[current_section] = (section_start, row - 1)
            current_section = match.group(1)
            section_start = row + 1
    
    # Close last section
    if current_section and section_start:
        sections[current_section] = (section_start, ws.max_row)
    
    return sections


def main():
    parser = argparse.ArgumentParser(description='Detect data boundaries in Excel sheets')
    parser.add_argument('workbook', help='Path to workbook')
    parser.add_argument('sheet', help='Sheet name to analyze')
    parser.add_argument('--name-col', type=int, default=2, help='Column with names (default: 2=B)')
    parser.add_argument('--value-col', type=int, default=3, help='Column with values (default: 3=C)')
    parser.add_argument('--start-row', type=int, default=1, help='Row to start from')
    parser.add_argument('--sections', action='store_true', help='Detect section boundaries')
    
    args = parser.parse_args()
    
    wb = openpyxl.load_workbook(args.workbook, data_only=True)
    ws = wb[args.sheet]
    
    print(f"Analyzing: {args.workbook} [{args.sheet}]")
    print(f"Dimensions: {ws.dimensions}")
    print(f"Max row: {ws.max_row}, Max col: {ws.max_column}")
    print()
    
    if args.sections:
        sections = detect_section_boundaries(ws, args.name_col)
        print("Detected sections:")
        for name, (start, end) in sections.items():
            print(f"  '{name}': rows {start}-{end}")
        print()
    
    rows = detect_data_rows(ws, args.name_col, args.value_col, args.start_row)
    
    print(f"Found {len(rows)} data rows:")
    for row_num, name, value in rows[:20]:
        print(f"  Row {row_num}: {name[:40]:40} = {value}")
    
    if len(rows) > 20:
        print(f"  ... and {len(rows) - 20} more")
    
    # Suggest code
    print(f"\nSuggested extraction code:")
    print(f"  data_rows = detect_data_rows(ws, {args.name_col}, {args.value_col}, {args.start_row})")
    print(f"  # Or hardcoded range (less robust): range({rows[0][0] if rows else 'START'}, {rows[-1][0]+1 if rows else 'END'})")


if __name__ == '__main__':
    main()
