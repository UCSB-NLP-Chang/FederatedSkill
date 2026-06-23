#!/usr/bin/env python3
"""
Quick structural verifier for openpyxl-generated workbooks.
Run this when no formal test file exists to catch common errors.
Usage: python3 scripts/quick_verifier.py <workbook_path> [--checks sheet_order,formulas,named_ranges]
"""

import argparse
import sys
import openpyxl
from openpyxl.workbook.defined_name import DefinedName

def verify_sheet_order(wb, expected):
    """Verify sheets appear in exact order expected."""
    actual = wb.sheetnames
    if actual == expected:
        print(f"  ✓ Sheet order matches: {actual}")
        return True
    print(f"  ✗ Sheet order mismatch")
    print(f"    Expected: {expected}")
    print(f"    Actual:   {actual}")
    return False

def verify_named_ranges(wb, min_count=0, required=None):
    """Verify named ranges exist and point to valid cells."""
    defined = list(wb.defined_names.values())
    print(f"  Named ranges: {len(defined)} total")
    
    if len(defined) < min_count:
        print(f"  ✗ Expected at least {min_count} named ranges")
        return False
    
    if required:
        missing = [r for r in required if r not in wb.defined_names]
        if missing:
            print(f"  ✗ Missing required ranges: {missing}")
            return False
        print(f"  ✓ All {len(required)} required ranges present")
    
    # Check for broken references
    broken = []
    for name in defined:
        try:
            dest = name.destinations
            for sheet, cell in dest:
                if sheet not in wb:
                    broken.append(f"{name.name} -> {sheet}!{cell}")
        except Exception as e:
            broken.append(f"{name.name} (error: {e})")
    
    if broken:
        print(f"  ✗ Broken references: {broken[:5]}")
        return False
    
    print(f"  ✓ All named ranges valid")
    return True

def verify_formulas(ws, sample_cells, expected_prefix='='):
    """Verify sample cells contain expected formulas."""
    errors = []
    for coord, expected_contains in sample_cells.items():
        cell = ws[coord]
        val = cell.value
        if val is None:
            errors.append(f"{coord}: empty")
        elif not isinstance(val, str) or not val.startswith(expected_prefix):
            errors.append(f"{coord}: expected formula starting with {expected_prefix}, got: {val[:50]}")
        elif expected_contains and expected_contains not in val:
            errors.append(f"{coord}: expected to contain '{expected_contains}', got: {val}")
    
    if errors:
        print(f"  ✗ Formula errors: {errors[:3]}")
        return False
    print(f"  ✓ Sample formulas valid")
    return True

def verify_totals_row(ws, totals_row, data_start_row, col_range):
    """Verify totals row uses SUM over expected range."""
    errors = []
    for col in col_range:
        coord = f"{col}{totals_row}"
        cell = ws[coord]
        expected_formula = f"=SUM({col}{data_start_row}:{col}{totals_row-1})"
        if cell.value != expected_formula:
            errors.append(f"{coord}: expected {expected_formula}, got {cell.value}")
    
    if errors:
        print(f"  ✗ Totals formula errors: {errors[:3]}")
        return False
    print(f"  ✓ Totals row formulas correct")
    return True

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('workbook')
    parser.add_argument('--sheets', help='Comma-separated expected sheet names')
    parser.add_argument('--min-ranges', type=int, default=0)
    parser.add_argument('--check-formulas', help='JSON: {"Sheet": {"A1": "expected_substring"}}')
    args = parser.parse_args()
    
    wb = openpyxl.load_workbook(args.workbook, data_only=False)
    all_ok = True
    
    print(f"\nVerifying: {args.workbook}")
    
    if args.sheets:
        expected = args.sheets.split(',')
        all_ok &= verify_sheet_order(wb, expected)
    
    if args.min_ranges > 0:
        all_ok &= verify_named_ranges(wb, min_count=args.min_ranges)
    
    if args.check_formulas:
        import json
        checks = json.loads(args.check_formulas)
        for sheet, cells in checks.items():
            print(f"\nChecking {sheet}:")
            all_ok &= verify_formulas(wb[sheet], cells)
    
    print()
    sys.exit(0 if all_ok else 1)

if __name__ == '__main__':
    main()
