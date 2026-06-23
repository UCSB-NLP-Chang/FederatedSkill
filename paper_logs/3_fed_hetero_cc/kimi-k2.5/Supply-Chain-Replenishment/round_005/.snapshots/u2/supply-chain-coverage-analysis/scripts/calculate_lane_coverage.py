#!/usr/bin/env python3
"""
Produce lane restock gap calculator.
Usage: python3 calculate_lane_coverage.py <lane_snapshot.xlsx> <arrival_board.xlsx> <output.xlsx>
"""
import sys
import math
from datetime import datetime, date, timedelta
from openpyxl import load_workbook, Workbook

def normalize_date(val):
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    return None

def parse_lane_snapshot(wb):
    """Extract lane/SKU inventory from section-based Lane Snapshot."""
    ws = wb['Lane Snapshot']
    
    # Metadata from row 1
    as_of = normalize_date(ws['B1'].value)
    horizon = normalize_date(ws['D1'].value)
    planning_days = (horizon - as_of).days
    
    # Parse lane sections
    inventory = []  # list of (lane, sku, cases, pull)
    current_lane = None
    in_data_section = False
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        cell_a = row[0]
        
        if cell_a is None:
            in_data_section = False
            continue
        
        val = str(cell_a).strip()
        
        if val.startswith('Lane:'):
            current_lane = val.replace('Lane:', '').strip()
            in_data_section = False
        elif val == 'SKU' and current_lane:
            in_data_section = True  # next row starts data
        elif in_data_section and current_lane:
            # Data row: SKU, Cases, Daily Pull
            sku = cell_a
            if sku:
                cases = row[1] or 0
                pull = row[2] or 0
                inventory.append((current_lane, sku, cases, pull))
    
    return as_of, horizon, planning_days, inventory

def parse_arrivals(wb, horizon):
    """Extract and filter arrivals, return dict of (lane, sku) -> list of (eta, cases)."""
    ws = wb['Arrival Board']
    arrivals = {}
    
    # Find header row
    header_idx = 1
    for idx, row in enumerate(ws.iter_rows(values_only=True), 1):
        if row[0] == 'Lane' and row[1] == 'SKU':
            header_idx = idx
            break
    
    for row in ws.iter_rows(min_row=header_idx + 1, values_only=True):
        lane = row[0]
        sku = row[1]
        eta_raw = row[2]
        cases = row[3] or 0
        status = row[4]
        
        if lane is None or sku is None:
            continue
        if status not in ('Ready', 'Docked'):
            continue
        
        eta = normalize_date(eta_raw)
        if eta is None or eta > horizon:
            continue
        
        key = (lane, sku)
        if key not in arrivals:
            arrivals[key] = []
        arrivals[key].append((eta, cases))
    
    return arrivals

def calculate_gaps(as_of, horizon, planning_days, inventory, arrivals):
    """Calculate coverage metrics for all lane/SKU pairs."""
    PALLET_SIZE = 54
    coverage = []
    gaps = []
    
    for lane, sku, cases_on_hand, daily_pull in inventory:
        # Skip if no consumption
        if daily_pull <= 0:
            continue
        
        # Current state
        current_doh = cases_on_hand / daily_pull
        oos_date = as_of + timedelta(days=math.floor(current_doh))
        
        # Inbound arrivals
        key = (lane, sku)
        inbound_list = arrivals.get(key, [])
        inbound_cases = sum(c for _, c in inbound_list)
        earliest_eta = min((e for e, _ in inbound_list), default=None)
        
        # Delivered state
        delivered_doh = (cases_on_hand + inbound_cases) / daily_pull
        remaining_demand = daily_pull * planning_days
        additional = max(0, remaining_demand - cases_on_hand - inbound_cases)
        pallets = math.ceil(additional / PALLET_SIZE) if additional > 0 else 0
        
        # Delivery requirements
        req_delivery = oos_date if additional > 0 else None
        earlier_req = False
        if req_delivery is not None:
            earlier_req = earliest_eta is None or earliest_eta > req_delivery
        
        coverage.append({
            'lane': lane, 'sku': sku, 'cases_on_hand': cases_on_hand,
            'daily_pull': daily_pull, 'current_doh': current_doh,
            'oos_date': oos_date, 'inbound_cases': inbound_cases,
            'delivered_doh': delivered_doh, 'remaining_demand': remaining_demand,
            'additional': additional, 'pallets': pallets,
            'req_delivery': req_delivery, 'earlier_req': earlier_req
        })
        
        if additional > 0:
            gaps.append({
                'lane': lane, 'sku': sku, 'req_delivery': req_delivery,
                'pallets': pallets, 'additional': additional,
                'earlier_req': earlier_req
            })
    
    return coverage, gaps

def write_output(output_path, as_of, horizon, planning_days, coverage, gaps):
    """Write Lane_Coverage and Restock_Actions sheets."""
    wb = Workbook()
    
    # Lane_Coverage sheet
    ws1 = wb.active
    ws1.title = 'Lane_Coverage'
    ws1.append(['Field', 'Value'])
    ws1.append(['AsOfDate', as_of])
    ws1.append(['HorizonEnd', horizon])
    ws1.append(['PlanningDays', planning_days])
    ws1.append([])
    ws1.append(['Lane', 'SKU', 'Cases_On_Hand', 'Daily_Pull_Cases_Per_Day',
                'Current_Days_On_Hand', 'Projected_OOS_Date', 'Inbound_Cases_By_Horizon',
                'Delivered_Days_On_Hand', 'Remaining_Demand_Cases', 'Additional_Cases_Needed',
                'Pallets_Required', 'Required_Delivery_Date', 'Earlier_Delivery_Required'])
    
    for c in coverage:
        ws1.append([c['lane'], c['sku'], c['cases_on_hand'], c['daily_pull'],
                    c['current_doh'], str(c['oos_date']), c['inbound_cases'],
                    c['delivered_doh'], c['remaining_demand'], c['additional'],
                    c['pallets'], str(c['req_delivery']) if c['req_delivery'] else None,
                    c['earlier_req']])
    
    # Restock_Actions sheet
    ws2 = wb.create_sheet('Restock_Actions')
    ws2.append(['Lane', 'SKU', 'Required_Delivery_Date', 'Pallets_Required',
                'Additional_Cases_Needed', 'Earlier_Delivery_Required'])
    
    for g in gaps:
        ws2.append([g['lane'], g['sku'], str(g['req_delivery']),
                    g['pallets'], g['additional'], g['earlier_req']])
    
    wb.save(output_path)
    print(f"Saved {output_path}")

def main(lane_path, arrival_path, output_path):
    lane_wb = load_workbook(lane_path, data_only=True)
    arrival_wb = load_workbook(arrival_path, data_only=True)
    
    as_of, horizon, planning_days, inventory = parse_lane_snapshot(lane_wb)
    arrivals = parse_arrivals(arrival_wb, horizon)
    coverage, gaps = calculate_gaps(as_of, horizon, planning_days, inventory, arrivals)
    write_output(output_path, as_of, horizon, planning_days, coverage, gaps)

if __name__ == '__main__':
    if len(sys.argv) != 4:
        print("Usage: python3 calculate_lane_coverage.py <lane_snapshot.xlsx> <arrival_board.xlsx> <output.xlsx>")
        sys.exit(1)
    main(sys.argv[1], sys.argv[2], sys.argv[3])