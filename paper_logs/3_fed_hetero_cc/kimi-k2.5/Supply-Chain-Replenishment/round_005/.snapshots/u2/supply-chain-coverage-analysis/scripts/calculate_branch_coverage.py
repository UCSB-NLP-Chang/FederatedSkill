#!/usr/bin/env python3
"""
Branch inventory coverage calculator for healthcare/pharmacy contexts.
Handles deduplication by Transfer ID and Confirmed-only status filtering.

Usage: python3 calculate_branch_coverage.py <branch_stock.xlsx> <planned_transfers.xlsx> <output.xlsx>
"""
import sys
import math
from datetime import timedelta
import pandas as pd
from openpyxl import Workbook

UNITS_PER_PALLET = 50

def normalize_date(val):
    """Convert pandas Timestamp or datetime to date."""
    if hasattr(val, 'date'):
        return val.date()
    return val

def deduplicate_transfers(df):
    """Keep row with maximum Transfer Date per Transfer ID."""
    df = df.copy()
    df['Transfer Date'] = pd.to_datetime(df['Transfer Date']).dt.date
    idx = df.groupby('Transfer ID')['Transfer Date'].idxmax()
    return df.loc[idx].reset_index(drop=True)

def main(branch_path, transfers_path, output_path):
    # Read Branch Stock with raw structure
    inventory_raw = pd.read_excel(branch_path, sheet_name='Branch Stock', header=None)
    
    # Extract metadata
    as_of = normalize_date(pd.to_datetime(inventory_raw.iloc[0, 1]))
    horizon = normalize_date(pd.to_datetime(inventory_raw.iloc[0, 3]))
    planning_days = (horizon - as_of).days
    
    # Parse inventory data (starts row 4, headers at row 3)
    inventory_df = inventory_raw.iloc[3:].copy()
    inventory_df.columns = ['Branch', 'Item', 'Units', 'Daily Use']
    inventory_df = inventory_df.dropna(subset=['Branch', 'Item'])
    inventory_df['Units'] = pd.to_numeric(inventory_df['Units'], errors='coerce').fillna(0)
    inventory_df['Daily Use'] = pd.to_numeric(inventory_df['Daily Use'], errors='coerce').fillna(0)
    
    # Read and process transfers
    transfers_df = pd.read_excel(transfers_path, sheet_name='Planned Transfers')
    transfers_deduped = deduplicate_transfers(transfers_df)
    transfers_confirmed = transfers_deduped[transfers_deduped['Status'] == 'Confirmed'].copy()
    transfers_confirmed['Units Planned'] = pd.to_numeric(transfers_confirmed['Units Planned'], errors='coerce').fillna(0)
    
    # Calculate coverage for each branch/item
    coverage_rows = []
    gap_rows = []
    
    for _, inv_row in inventory_df.iterrows():
        branch = inv_row['Branch']
        item = inv_row['Item']
        units_on_hand = inv_row['Units']
        daily_use = inv_row['Daily Use']
        
        # Calculate inbound units
        mask = (
            (transfers_confirmed['Branch'] == branch) &
            (transfers_confirmed['Item'] == item) &
            (transfers_confirmed['Transfer Date'] <= horizon)
        )
        matching = transfers_confirmed[mask]
        inbound_units = matching['Units Planned'].sum()
        earliest_eta = matching['Transfer Date'].min() if len(matching) > 0 else None
        
        # Coverage calculations
        if daily_use > 0:
            current_doh = units_on_hand / daily_use
            projected_oos = as_of + timedelta(days=math.floor(current_doh))
            delivered_doh = (units_on_hand + inbound_units) / daily_use
            remaining_demand = daily_use * planning_days
        else:
            current_doh = float('inf')
            projected_oos = as_of
            delivered_doh = float('inf')
            remaining_demand = 0
        
        additional_needed = max(0, remaining_demand - units_on_hand - inbound_units)
        pallets_required = math.ceil(additional_needed / UNITS_PER_PALLET) if additional_needed > 0 else 0
        required_delivery = projected_oos if additional_needed > 0 else None
        
        earlier_required = False
        if required_delivery:
            earlier_required = earliest_eta is None or earliest_eta > required_delivery
        
        coverage_rows.append([
            branch, item, units_on_hand, daily_use,
            current_doh, projected_oos, inbound_units,
            delivered_doh, remaining_demand, additional_needed,
            pallets_required, required_delivery, earlier_required
        ])
        
        if additional_needed > 0:
            gap_rows.append([branch, item, required_delivery, pallets_required, additional_needed, earlier_required])
    
    # Write output
    wb = Workbook()
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Branch_Item_Coverage sheet
    ws1 = wb.create_sheet('Branch_Item_Coverage')
    ws1['A1'], ws1['B1'] = 'Field', 'Value'
    ws1['A2'], ws1['B2'] = 'AsOfDate', as_of
    ws1['A3'], ws1['B3'] = 'HorizonEnd', horizon
    ws1['A4'], ws1['B4'] = 'PlanningDays', planning_days
    
    headers = ['Branch', 'Item', 'Units_On_Hand', 'Daily_Use_Units_Per_Day',
               'Current_Days_On_Hand', 'Projected_OOS_Date', 'Inbound_Units_By_Horizon',
               'Delivered_Days_On_Hand', 'Remaining_Demand_Units', 'Additional_Units_Needed',
               'Pallets_Required', 'Required_Delivery_Date', 'Earlier_Delivery_Required']
    
    for col_idx, h in enumerate(headers, 1):
        ws1.cell(row=6, column=col_idx, value=h)
    
    for row_idx, row_data in enumerate(coverage_rows, 7):
        for col_idx, val in enumerate(row_data, 1):
            ws1.cell(row=row_idx, column=col_idx, value=val)
    
    # Transfer_Gap_List sheet
    ws2 = wb.create_sheet('Transfer_Gap_List')
    ws2.append(['Branch', 'Item', 'Required_Delivery_Date', 'Pallets_Required',
                'Additional_Units_Needed', 'Earlier_Delivery_Required'])
    for row in gap_rows:
        ws2.append(row)
    
    wb.save(output_path)
    print(f"Saved {output_path}")

if __name__ == '__main__':
    if len(sys.argv) != 4:
        print("Usage: python3 calculate_branch_coverage.py <branch_stock.xlsx> <planned_transfers.xlsx> <output.xlsx>")
        sys.exit(1)
    main(sys.argv[1], sys.argv[2], sys.argv[3])
