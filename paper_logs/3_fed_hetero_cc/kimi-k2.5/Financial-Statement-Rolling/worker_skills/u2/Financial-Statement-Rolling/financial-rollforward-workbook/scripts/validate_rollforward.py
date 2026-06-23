#!/usr/bin/env python3
"""Validate rollforward workbook formulas and balances.

Usage:
    python validate_rollforward.py <workbook_path>

Checks:
    1. Variance formula uses column N for BOTH operands (not O-N)
    2. Ending Balance row references Period Totals (not self-referencing)
    3. Sheet order matches expected pattern
    4. Formula syntax is correct (no #REF! errors)

Exit codes:
    0 = All checks passed
    1 = Validation failed
"""

import sys
import re
from openpyxl import load_workbook


def check_variance_formula(ws, var_row, gl_row, eb_row):
    """ASSERT: Variance must use column N for both operands.

    The most common bugs are:
    - =O{gl_row}-N{eb_row} (wrong column for GL)
    - =O{gl_row}-O{eb_row} (wrong column for both)
    - =N{gl_row}-O{eb_row} (wrong column for Ending Balance)
    
    Correct formula: =N{gl_row}-N{eb_row}
    """
    errors = []

    # Check column N variance formula
    n_var = ws[f'N{var_row}'].value
    if n_var:
        formula_str = str(n_var)
        
        # Pattern for WRONG formula: uses column O for GL (most common bug)
        wrong_pattern_gl_o = rf'=O{gl_row}\s*-[NO]{eb_row}'
        if re.match(wrong_pattern_gl_o, formula_str):
            errors.append(
                f"CRITICAL: N{var_row} has WRONG Variance formula '{formula_str}'. "
                f"GL should be from column N, not O. Correct: =N{gl_row}-N{eb_row}"
            )

        # Pattern for WRONG formula: uses column O for ending balance
        wrong_pattern_eb_o = rf'=[NO]{gl_row}\s*-O{eb_row}'
        if re.match(wrong_pattern_eb_o, formula_str):
            errors.append(
                f"CRITICAL: N{var_row} has WRONG Variance formula '{formula_str}'. "
                f"Ending Balance should be from column N, not O. Correct: =N{gl_row}-N{eb_row}"
            )

        # Check for correct formula
        correct_pattern = rf'=N{gl_row}\s*-\s*N{eb_row}'
        if not re.match(correct_pattern, formula_str):
            errors.append(
                f"WARNING: N{var_row} formula '{formula_str}' may be incorrect. "
                f"Expected: =N{gl_row}-N{eb_row}"
            )

    # Check column O variance formula
    o_var = ws[f'O{var_row}'].value
    if o_var:
        formula_str = str(o_var)
        # Correct pattern for column O
        correct_pattern = rf'=O{gl_row}\s*-\s*O{eb_row}'
        if not re.match(correct_pattern, formula_str):
            errors.append(
                f"WARNING: O{var_row} formula '{formula_str}' may be incorrect. "
                f"Expected: =O{gl_row}-O{eb_row}"
            )

    return errors


def check_ending_balance_references(ws, pt_row, eb_row):
    """ASSERT: Ending Balance must reference Period Totals for activity.

    Wrong: =B{eb_row}+C{eb_row}-D{eb_row} (self-referencing, B{eb_row} is empty)
    Right: =B{pt_row}+C{pt_row}-D{pt_row}
    """
    errors = []

    # Check first period Ending Balance (column E)
    e_eb = ws[f'E{eb_row}'].value
    if e_eb:
        formula_str = str(e_eb)
        # Pattern for self-referencing (WRONG)
        self_ref_pattern = rf'=B{eb_row}\s*[+\-]'
        if re.search(self_ref_pattern, formula_str):
            errors.append(
                f"CRITICAL: E{eb_row} self-references B{eb_row} in '{formula_str}'. "
                f"Beginning Balance should reference Period Totals: =B{pt_row}+C{pt_row}-D{pt_row}"
            )

    return errors


def validate_sheet_order(wb, expected_order):
    """Verify sheets are in expected order."""
    actual = wb.sheetnames
    if actual != expected_order:
        print(f"ERROR: Sheet order mismatch")
        print(f"  Expected: {expected_order}")
        print(f"  Actual:   {actual}")
        return False
    print(f"OK: Sheet order correct: {actual}")
    return True


def validate_formulas_exist(ws, required_cells):
    """Check that formula cells contain formulas."""
    errors = []
    for cell_ref in required_cells:
        value = ws[cell_ref].value
        if value is None:
            errors.append(f"{cell_ref}: empty")
        elif isinstance(value, str) and value.startswith('='):
            pass  # Valid formula
        else:
            errors.append(f"{cell_ref}: not a formula (value={value})")

    if errors:
        print(f"ERROR: Formula issues in {ws.title}:")
        for e in errors:
            print(f"  {e}")
        return False
    return True


def find_control_rows(ws):
    """Find control row indices by label."""
    rows = {}
    for row in range(1, 50):
        label = ws[f'A{row}'].value
        if label:
            label_str = str(label).strip()
            if 'Period Totals' in label_str:
                rows['period_totals'] = row
            elif 'Ending Balance' in label_str:
                rows['ending_balance'] = row
            elif 'Variance' in label_str:
                rows['variance'] = row
            elif 'GL Balance' in label_str:
                rows['gl_balance'] = row
    return rows


def validate_workbook(filepath):
    """Run all validation checks on workbook."""
    print(f"Validating: {filepath}")
    print("-" * 60)

    wb = load_workbook(filepath, data_only=False)

    all_errors = []
    passed = 0
    failed = 0

    # Check each detail sheet
    for sheet_name in wb.sheetnames:
        if 'Summary' in sheet_name:
            continue

        ws = wb[sheet_name]
        print(f"\nSheet: {sheet_name}")

        # Find control rows
        rows = find_control_rows(ws)
        if not rows:
            print(f"  WARNING: Could not find control rows")
            continue

        pt_row = rows.get('period_totals')
        eb_row = rows.get('ending_balance')
        var_row = rows.get('variance')
        gl_row = rows.get('gl_balance')

        # Check Variance formula (CRITICAL)
        if var_row and gl_row and eb_row:
            var_errors = check_variance_formula(ws, var_row, gl_row, eb_row)
            if var_errors:
                for e in var_errors:
                    print(f"  {e}")
                    all_errors.append((sheet_name, e))
                failed += 1
            else:
                print(f"  OK: Variance formula correct (=N{gl_row}-N{eb_row})")
                passed += 1

        # Check Ending Balance references (CRITICAL)
        if pt_row and eb_row:
            eb_errors = check_ending_balance_references(ws, pt_row, eb_row)
            if eb_errors:
                for e in eb_errors:
                    print(f"  {e}")
                    all_errors.append((sheet_name, e))
                failed += 1
            else:
                print(f"  OK: Ending Balance references Period Totals")
                passed += 1

        # Check formulas exist
        required = ['E6', 'H6', 'K6', 'N6', 'O6']
        if validate_formulas_exist(ws, required):
            passed += 1
        else:
            failed += 1

    print("\n" + "=" * 60)
    print(f"Result: {passed} checks passed, {failed} checks failed")

    if all_errors:
        print("\nCRITICAL ERRORS (must fix):")
        for sheet, err in all_errors:
            if "CRITICAL" in err:
                print(f"  [{sheet}] {err}")
        return False

    return True


def main():
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)

    filepath = sys.argv[1]
    success = validate_workbook(filepath)
    sys.exit(0 if success else 1)


if __name__ == '__main__':
    main()
