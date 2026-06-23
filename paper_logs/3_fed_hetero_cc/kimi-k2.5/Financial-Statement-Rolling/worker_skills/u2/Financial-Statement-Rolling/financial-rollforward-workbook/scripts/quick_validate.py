#!/usr/bin/env python3
"""Quick pre-submission validation for rollforward workbooks.

Run this BEFORE the full test suite to catch the most common fatal bugs.

Usage:
    python scripts/quick_validate.py <workbook_path>

Checks (fast, targeted):
    1. Variance formula uses column N for BOTH operands
    2. Ending Balance references Period Totals (not self)
    3. Summary GL Balance links use column N (not O)
    4. No #REF! errors in formula cells

Exit codes:
    0 = All quick checks passed
    1 = One or more checks failed
"""

import sys
import re
from openpyxl import load_workbook


def find_control_rows(ws):
    """Find control row indices by scanning column A labels."""
    rows = {}
    for row in range(1, min(ws.max_row + 1, 50)):
        label = ws.cell(row=row, column=1).value
        if label:
            label_str = str(label).strip()
            if 'Period Totals' in label_str or 'Month Totals' in label_str:
                rows['period_totals'] = row
            elif 'Ending Balance' in label_str:
                rows['ending_balance'] = row
            elif 'Variance' in label_str:
                rows['variance'] = row
            elif 'GL Balance' in label_str:
                rows['gl_balance'] = row
    return rows


def check_variance(ws, var_row, gl_row, eb_row):
    """CRITICAL: Variance must be =N{gl}-N{eb}, NOT =O{gl}-N{eb} or =O{gl}-O{eb}."""
    cell_val = ws.cell(row=var_row, column=14).value  # Column N
    if not cell_val or not isinstance(cell_val, str):
        return False, f"N{var_row} is empty or not a formula"
    formula = cell_val.strip()
    expected = f'=N{gl_row}-N{eb_row}'
    if formula == expected:
        return True, f"Variance correct: {formula}"
    # Check for common wrong patterns
    if formula.startswith('=O'):
        return False, f"Variance WRONG: uses column O for GL. Got '{formula}', expected '{expected}'"
    if '-O' in formula:
        return False, f"Variance WRONG: uses column O for Ending Balance. Got '{formula}', expected '{expected}'"
    return False, f"Variance unexpected: '{formula}', expected '{expected}'"


def check_ending_balance(ws, pt_row, eb_row):
    """CRITICAL: Ending Balance must reference Period Totals, not self."""
    cell_val = ws.cell(row=eb_row, column=5).value  # Column E
    if not cell_val or not isinstance(cell_val, str):
        return False, f"E{eb_row} is empty or not a formula"
    formula = cell_val.strip()
    # Check for self-reference (B{eb_row} instead of B{pt_row})
    if f'B{eb_row}' in formula:
        return False, f"Ending Balance WRONG: self-references B{eb_row}. Got '{formula}', should reference B{pt_row}"
    if f'B{pt_row}' in formula:
        return True, f"Ending Balance correct: references Period Totals"
    return False, f"Ending Balance unexpected: '{formula}'"


def check_summary_gl_links(ws):
    """CRITICAL: Summary sheet GL Balance links must use column N, not O."""
    errors = []
    for row in range(1, min(ws.max_row + 1, 50)):
        for col in range(1, min(ws.max_column + 1, 20)):
            val = ws.cell(row=row, column=col).value
            if val and isinstance(val, str) and val.startswith('='):
                # Check if this is a GL Balance link (contains 'GL' in nearby label or formula pattern)
                label = ws.cell(row=row, column=1).value
                if label and isinstance(label, str) and 'GL' in label.upper() and 'Balance' in label:
                    # This row's formula should use column N
                    if '!O' in val:
                        errors.append(f"Row {row}: GL Balance link uses column O '{val}' (should be N)")
                    elif '!N' not in val:
                        errors.append(f"Row {row}: GL Balance link missing column N '{val}'")
                # Also check any formula that looks like GL Balance link pattern
                # Pattern: ='Sheet Name'!O{row} where row number suggests GL row (high numbers)
                if '!O' in val and re.search(r'!O\d+', val):
                    # Check if this might be a GL link by looking at context
                    # GL rows are typically 4 rows above data_end, so higher numbers
                    gl_row_match = re.search(r'!O(\d+)', val)
                    if gl_row_match:
                        row_num = int(gl_row_match.group(1))
                        # GL rows are typically in the 40-60 range for 5-10 partner sheets
                        if row_num >= 10 and '!O' in val:
                            # Flag potential GL link using O instead of N
                            errors.append(f"Row {row}: Potential GL Balance link uses O '{val}' - verify if this is GL Balance")

    if errors:
        return False, errors
    return True, []


def check_ref_errors(ws):
    """Check for #REF! errors in any cell."""
    for row in range(1, min(ws.max_row + 1, 30)):
        for col in range(1, min(ws.max_column + 1, 20)):
            val = ws.cell(row=row, column=col).value
            if val and isinstance(val, str) and '#REF' in val:
                return False, f"#REF! error at {chr(64+col)}{row}"
    return True, "No #REF! errors"


def main(filepath):
    print(f"Quick validation: {filepath}")
    print("=" * 50)

    try:
        wb = load_workbook(filepath, data_only=False)
    except Exception as e:
        print(f"FATAL: Cannot load workbook: {e}")
        return 1

    all_passed = True

    # Check detail sheets
    for sheet_name in wb.sheetnames:
        if 'Summary' in sheet_name:
            continue
        ws = wb[sheet_name]
        print(f"\nSheet: {sheet_name}")

        rows = find_control_rows(ws)
        if not rows:
            print("  SKIP: No control rows found")
            continue

        pt = rows.get('period_totals')
        eb = rows.get('ending_balance')
        var = rows.get('variance')
        gl = rows.get('gl_balance')

        # Check Variance
        if var and gl and eb:
            ok, msg = check_variance(ws, var, gl, eb)
            status = "PASS" if ok else "FAIL"
            print(f"  [{status}] Variance: {msg}")
            all_passed = all_passed and ok

        # Check Ending Balance
        if pt and eb:
            ok, msg = check_ending_balance(ws, pt, eb)
            status = "PASS" if ok else "FAIL"
            print(f"  [{status}] Ending Balance: {msg}")
            all_passed = all_passed and ok

        # Check REF errors
        ok, msg = check_ref_errors(ws)
        status = "PASS" if ok else "FAIL"
        print(f"  [{status}] REF errors: {msg}")
        all_passed = all_passed and ok

    # Check Summary sheet GL links
    if 'Summary' in wb.sheetnames:
        print(f"\nSummary Sheet:")
        ws = wb['Summary']
        ok, errors = check_summary_gl_links(ws)
        if ok:
            print(f"  [PASS] GL Balance links: All use column N correctly")
        else:
            print(f"  [FAIL] GL Balance links:")
            for e in errors:
                print(f"    {e}")
            all_passed = all_passed and ok
        # Also check REF errors in summary
        ok, msg = check_ref_errors(ws)
        status = "PASS" if ok else "FAIL"
        print(f"  [{status}] REF errors: {msg}")
        all_passed = all_passed and ok

    print("\n" + "=" * 50)
    if all_passed:
        print("ALL QUICK CHECKS PASSED - proceed to full test suite")
        return 0
    else:
        print("QUICK CHECKS FAILED - fix before running test suite")
        return 1


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)
    sys.exit(main(sys.argv[1]))