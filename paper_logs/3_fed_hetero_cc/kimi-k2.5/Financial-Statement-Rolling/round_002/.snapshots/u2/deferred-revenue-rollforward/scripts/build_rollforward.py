#!/usr/bin/env python3
"""Template for building deferred revenue rollforward workbooks.

Adapt paths, account names, and output filename to the specific task.
Do NOT use hardcoded financial values for cells that should contain formulas.

Usage:
    python3 build_rollforward.py
"""

import csv
import json
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


def read_csv(path):
    with open(path, newline="") as f:
        return list(csv.DictReader(f))


def read_gl(path):
    with open(path) as f:
        return json.load(f)


MONETARY_FMT = "#,##0.00"


def build_detail_sheet(ws, data, gl_values, months_order):
    """Populate a detail sheet with line items and control rows.

    Args:
        ws: openpyxl Worksheet
        data: list of dicts from CSV
        gl_values: dict of {month: balance} from GL source
        months_order: list like ['may', 'jun', 'jul', 'aug']
    """
    HEADER_ROW = 5
    DATA_START = 6
    n_items = len(data)

    # Headers
    headers = ["Customer", "Beginning Balance"]
    for m in months_order:
        title = m.capitalize()
        headers += [f"{title} Billings", f"{title} Recognition", f"{title} Ending Balance"]
    headers += ["Contract Months", "Notes", "Revenue Code"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=HEADER_ROW, column=col, value=h)

    # Line items
    for i, row in enumerate(data):
        r = DATA_START + i
        ws.cell(row=r, column=1, value=row["entity"])
        ws.cell(row=r, column=2, value=float(row["beginning_balance"]))
        ws.cell(row=r, column=2).number_format = MONETARY_FMT
        col_idx = 3
        for m in months_order:
            ws.cell(row=r, column=col_idx, value=float(row[f"{m}_adds"]))
            ws.cell(row=r, column=col_idx).number_format = MONETARY_FMT
            ws.cell(row=r, column=col_idx + 1, value=float(row[f"{m}_release"]))
            ws.cell(row=r, column=col_idx + 1).number_format = MONETARY_FMT
            ws.cell(row=r, column=col_idx + 2, value=float(row[f"{m}_ending_balance"]))
            ws.cell(row=r, column=col_idx + 2).number_format = MONETARY_FMT
            col_idx += 3
        ws.cell(row=r, column=col_idx, value=int(row["term_months"]))
        ws.cell(row=r, column=col_idx + 1, value=row["comments"])
        ws.cell(row=r, column=col_idx + 2, value=row["account_number"])

    last_data_row = DATA_START + n_items - 1
    totals_row = last_data_row + 1
    ending_row = totals_row + 1
    variance_row = ending_row + 1
    gl_row = variance_row + 1

    # Period Totals — SUM formulas
    ws.cell(row=totals_row, column=1, value="Period Totals")
    for c in range(2, 15):
        col_letter = get_column_letter(c)
        cell = ws.cell(row=totals_row, column=c,
                       value=f"=SUM({col_letter}{DATA_START}:{col_letter}{last_data_row})")
        cell.number_format = MONETARY_FMT
    # Total billings in O column
    billing_cols = [3, 6, 9, 12]
    total_bill_formula = "+".join(
        f"{get_column_letter(c)}{totals_row}" for c in billing_cols
    )
    ws.cell(row=totals_row, column=15, value=f"={total_bill_formula}").number_format = MONETARY_FMT

    # Ending Balance — rollforward formulas (CORRECT pattern)
    ws.cell(row=ending_row, column=1, value="Ending Balance")
    # B_ending MUST reference B_totals (Beginning Balance from Period Totals)
    ws.cell(row=ending_row, column=2, value=f"=B{totals_row}").number_format = MONETARY_FMT
    # May Ending: BegBal(from totals) + MayBillings - MayRecognition
    ws.cell(row=ending_row, column=5,
            value=f"=B{ending_row}+C{totals_row}-D{totals_row}").number_format = MONETARY_FMT
    # Jun Ending: MayEnd + JunBill - JunRec
    ws.cell(row=ending_row, column=8,
            value=f"=E{ending_row}+F{totals_row}-G{totals_row}").number_format = MONETARY_FMT
    # Jul Ending
    ws.cell(row=ending_row, column=11,
            value=f"=H{ending_row}+I{totals_row}-J{totals_row}").number_format = MONETARY_FMT
    # Aug Ending
    ws.cell(row=ending_row, column=14,
            value=f"=K{ending_row}+L{totals_row}-M{totals_row}").number_format = MONETARY_FMT
    # Total recognition
    rec_cols = [4, 7, 10, 13]
    total_rec_formula = "+".join(
        f"{get_column_letter(c)}{totals_row}" for c in rec_cols
    )
    ws.cell(row=ending_row, column=15, value=f"={total_rec_formula}").number_format = MONETARY_FMT

    # Variance = GL Balance - Ending Balance
    ws.cell(row=variance_row, column=1, value="Variance")
    ws.cell(row=variance_row, column=14,
            value=f"=N{gl_row}-N{ending_row}").number_format = MONETARY_FMT

    # GL Balance — hardcoded values from source JSON
    ws.cell(row=gl_row, column=1, value="GL Balance")
    gl_col_map = {0: 5, 1: 8, 2: 11, 3: 14}
    for i, m in enumerate(months_order):
        ws.cell(row=gl_row, column=gl_col_map[i], value=gl_values[m]).number_format = MONETARY_FMT

    return {
        "totals_row": totals_row,
        "ending_row": ending_row,
        "variance_row": variance_row,
        "gl_row": gl_row,
    }


def build_summary_sheet(ws, detail_sheets_info, months_order):
    """Build the summary sheet with links to detail sheets.

    IMPORTANT: Use single quotes around sheet names containing spaces or special chars.
    """
    ws.cell(row=1, column=1, value="Company Name")
    ws.cell(row=2, column=1, value="Deferred Revenue Analysis")
    ws.cell(row=3, column=1, value="Period Ending: August 2025")

    row = 5
    for sheet_name, info in detail_sheets_info.items():
        ws.cell(row=row, column=1, value=sheet_name)
        row += 1
        ws.cell(row=row, column=1, value="Period Totals")
        ws.cell(row=row, column=2,
                value=f"='{sheet_name}'!N{info['totals_row']}").number_format = MONETARY_FMT
        row += 1
        ws.cell(row=row, column=1, value="Ending Balance")
        ws.cell(row=row, column=2,
                value=f"='{sheet_name}'!N{info['ending_row']}").number_format = MONETARY_FMT
        row += 1
        ws.cell(row=row, column=1, value="Variance")
        ws.cell(row=row, column=2,
                value=f"='{sheet_name}'!N{info['variance_row']}").number_format = MONETARY_FMT
        row += 2


def main():
    saas_csv = "/root/saas_deferred_revenue_schedule.csv"
    services_csv = "/root/services_deferred_revenue_schedule.csv"
    gl_json = "/root/gl_balances.json"
    output_path = "/root/Deferred_Revenue.xlsx"
    months = ["may", "jun", "jul", "aug"]

    gl = read_gl(gl_json)
    saas_data = read_csv(saas_csv)
    services_data = read_csv(services_csv)

    wb = Workbook()
    wb.remove(wb.active)

    summary = wb.create_sheet("Deferred Summary", 0)
    saas_ws = wb.create_sheet("SaaS Rev #2300", 1)
    services_ws = wb.create_sheet("Services Rev #2310", 2)

    saas_info = build_detail_sheet(saas_ws, saas_data, gl["saas_rev_2300"], months)
    services_info = build_detail_sheet(services_ws, services_data, gl["services_rev_2310"], months)

    build_summary_sheet(summary, {
        "SaaS Rev #2300": saas_info,
        "Services Rev #2310": services_info,
    }, months)

    wb.save(output_path)
    print(f"Workbook saved to {output_path}")


if __name__ == "__main__":
    main()
