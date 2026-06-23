#!/usr/bin/env python3
"""Verify rollforward workbook structure and formulas.

Usage:
    python verify_workbook.py <workbook_path>

Checks:
    1. Sheet order (Summary first)
    2. Control rows exist (Period Totals, Ending Balance, Variance, GL Balance)
    3. Variance formula uses column N for GL Balance (CRITICAL)
    4. Ending Balance references Period Totals, not self
    5. Cross-sheet references use proper quoting
"""

import sys
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def check_sheet_order(wb):
    """Verify Summary sheet is first."""
    sheets = wb.sheetnames
    if not sheets:
        return False, "No sheets found"
    if sheets[0] != 'Summary':
        return False, f"First sheet should be 'Summary', got '{sheets[0]}'"
    return True, f"Sheet order OK: {sheets}"


def find_control_rows(ws):
    """Find control row positions by label."""
    rows = {}
    for row in range(1, ws.max_row + 1):
        label = ws.cell(row=row, column=1).value
        if label in ['Period Totals', 'Ending Balance', 'Variance', 'GL Balance']:
            rows[label] = row
    return rows


def check_variance_formula(ws, control_rows):
    """CRITICAL: Variance must use column N for GL Balance."""
    if 'Variance' not in control_rows or 'GL Balance' not in control_rows or 'Ending Balance' not in control_rows:
        return False, "Missing control rows"

    var_row = control_rows['Variance']
    formula = ws.cell(row=var_row, column=15).value  # Column O

    if not formula or not isinstance(formula, str) or not formula.startswith('='):
        return False, f"Variance row {var_row}: not a formula '{formula}'"

    gl_row = control_rows['GL Balance']
    eb_row = control_rows['Ending Balance']

    # Expected pattern: =N{gl_row}-N{eb_row}
    expected = f'=N{gl_row}-N{eb_row}'

    # Check for WRONG pattern: =O-N
    if 'O' in formula and '-N' in formula:
        return False, f"Variance row {var_row}: WRONG formula '{formula}' - uses column O for GL (should be N)"

    # Check it references column N for GL
    if f'N{gl_row}' not in formula:
        return False, f"Variance row {var_row}: formula '{formula}' should reference N{gl_row} for GL Balance"

    return True, f"Variance formula OK: {formula}"


def check_ending_balance_formula(ws, control_rows):
    """Ending Balance Beginning Balance should reference Period Totals."""
    if 'Ending Balance' not in control_rows or 'Period Totals' not in control_rows:
        return False, "Missing control rows"

    eb_row = control_rows['Ending Balance']
    pt_row = control_rows['Period Totals']

    # Check first month ending (column E) references Period Totals for activity
    e_formula = ws.cell(row=eb_row, column=5).value

    if not e_formula or not isinstance(e_formula, str):
        return False, f"EB row {eb_row} column E: not a formula"

    # Should reference B{pt_row} for beginning balance (from Period Totals)
    if f'B{pt_row}' in e_formula:
        return True, f"Ending Balance OK: references Period Totals B{pt_row}"

    # Check if self-referencing (WRONG)
    if f'B{eb_row}' in e_formula:
        return False, f"EB row {eb_row}: WRONG - self-references B{eb_row}, should reference Period Totals B{pt_row}"

    return True, f"Ending Balance formula: {e_formula}"


def check_cross_sheet_refs(wb):
    """Verify cross-sheet references use proper quoting."""
    errors = []
    if 'Summary' not in wb.sheetnames:
        return False, "No Summary sheet"

    summary = wb['Summary']
    for row in range(1, summary.max_row + 1):
        for col in range(1, summary.max_column + 1):
            cell = summary.cell(row=row, column=col)
            if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                # Check for sheet reference without quotes when name has space
                if '!' in cell.value and "'" not in cell.value:
                    # Extract sheet name
                    parts = cell.value.split('!')
                    if len(parts) >= 1:
                        sheet_ref = parts[0].lstrip('=')
                        if ' ' in sheet_ref or '#' in sheet_ref:
                            errors.append(f"Summary {cell.coordinate}: '{cell.value}' needs quotes around '{sheet_ref}'")

    if errors:
        return False, "Cross-sheet ref errors: " + "; ".join(errors)
    return True, "Cross-sheet references OK"


def main(filepath):
    print(f"Verifying: {filepath}")
    print("-" * 50)

    try:
        wb = load_workbook(filepath, data_only=False)
    except Exception as e:
        print(f"ERROR: Cannot load workbook: {e}")
        return 1

    all_ok = True

    # Check sheet order
    ok, msg = check_sheet_order(wb)
    print(f"[{'OK' if ok else 'FAIL'}] {msg}")
    all_ok = all_ok and ok

    # Check each detail sheet
    for sheet_name in wb.sheetnames:
        if 'Summary' in sheet_name:
            continue

        ws = wb[sheet_name]
        print(f"\nSheet: {sheet_name}")

        control_rows = find_control_rows(ws)

        # Check control rows exist
        for label in ['Period Totals', 'Ending Balance', 'Variance', 'GL Balance']:
            if label not in control_rows:
                print(f"[FAIL] Missing control row: {label}")
                all_ok = False

        # Check Variance formula (CRITICAL)
        ok, msg = check_variance_formula(ws, control_rows)
        print(f"[{'OK' if ok else 'FAIL'}] {msg}")
        all_ok = all_ok and ok

        # Check Ending Balance formula
        ok, msg = check_ending_balance_formula(ws, control_rows)
        print(f"[{'OK' if ok else 'FAIL'}] {msg}")
        all_ok = all_ok and ok

    # Check cross-sheet refs
    print("\n[Cross-sheet References]")
    ok, msg = check_cross_sheet_refs(wb)
    print(f"[{'OK' if ok else 'FAIL'}] {msg}")
    all_ok = all_ok and ok

    print("-" * 50)
    if all_ok:
        print("PASS: All checks passed")
        return 0
    else:
        print("FAIL: Some checks failed")
        return 1


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)
    sys.exit(main(sys.argv[1]))
