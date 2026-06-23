#!/usr/bin/env python3
"""Verify production catch-up plan Excel file structure."""
import argparse
import sys
import openpyxl


def verify_plan(filepath, expected_weeks, sheet_name='Plan', expected_columns=None):
    """Verify Excel plan matches expected structure using openpyxl."""
    errors = []

    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb[sheet_name]
    except (KeyError, ValueError) as e:
        print(f"FAIL: Could not read sheet '{sheet_name}': {e}")
        return False

    # Extract headers from row 1
    headers = [cell.value for cell in ws[1]]

    # Count data rows (excluding header)
    data_rows = list(ws.iter_rows(min_row=2, values_only=True))
    row_count = len([r for r in data_rows if r[0] is not None])

    # Check row count
    if row_count != expected_weeks:
        errors.append(f"Row count: got {row_count}, expected {expected_weeks}")

    # Check columns
    if expected_columns:
        if headers != expected_columns:
            errors.append(f"Column mismatch:\n  Got: {headers}\n  Expected: {expected_columns}")

    # Check week sequence
    weeks = [row[0] for row in data_rows if row[0] is not None]
    if weeks:
        expected_seq = list(range(int(weeks[0]), int(weeks[0]) + expected_weeks))
        if weeks != expected_seq:
            missing = set(expected_seq) - set(weeks)
            extra = set(weeks) - set(expected_seq)
            if missing:
                errors.append(f"Missing weeks: {sorted(missing)}")
            if extra:
                errors.append(f"Extra weeks: {sorted(extra)}")

    # Check state carry if both columns present
    # Find column indices for state tracking
    end_col_idx = None
    start_col_idx = None
    for i, h in enumerate(headers):
        if h and 'End of Week Backlog' in str(h):
            end_col_idx = i
        if h and 'Start of Week Past Due' in str(h):
            start_col_idx = i

    if end_col_idx is not None and start_col_idx is not None:
        mismatches = []
        for i in range(len(data_rows) - 1):
            end_val = data_rows[i][end_col_idx]
            next_start_val = data_rows[i + 1][start_col_idx]
            if end_val is None or next_start_val is None:
                continue
            if end_val > 0 and abs(end_val - next_start_val) > 0.01:
                mismatches.append(f"Row {i+1}: end={end_val:.2f}, next_start={next_start_val:.2f}")
            elif end_val <= 0 and next_start_val != 0:
                mismatches.append(f"Row {i+1}: end={end_val:.2f} (<=0), next_start={next_start_val:.2f} (should be 0)")
        if mismatches:
            errors.append(f"State carry errors:\n  " + "\n  ".join(mismatches[:5]))

    if errors:
        print("FAIL:")
        for e in errors:
            print(f"  - {e}")
        return False

    print(f"PASS: {filepath} validated ({expected_weeks} rows, sheet '{sheet_name}')")
    return True


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Verify production plan Excel file')
    parser.add_argument('filepath', help='Path to Excel file')
    parser.add_argument('--weeks', type=int, required=True, help='Expected number of weeks/rows')
    parser.add_argument('--sheet', default='Plan', help='Sheet name to verify')
    parser.add_argument('--columns', nargs='+', help='Expected column headers')

    args = parser.parse_args()
    success = verify_plan(args.filepath, args.weeks, args.sheet, args.columns)
    sys.exit(0 if success else 1)