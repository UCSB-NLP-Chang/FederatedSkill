#!/usr/bin/env python3
"""
Template for iterative capacity/backlog planning from Excel data.

CRITICAL: Read the task prompt first and fill in EXACT values below.
Do not run this script with default/example values.
"""
import openpyxl
import sys

def run_planning(input_path, output_plan_path, output_summary_path):
    # ============================================================
    # STEP 1: EXTRACT THESE VALUES FROM THE TASK PROMPT
    # ============================================================

    # From task prompt: "initial backlog = X" or "initial condition = Y"
    # If "initial condition", calculate: INITIAL_BACKLOG = Y - first_week_demand
    INITIAL_BACKLOG = None  # MUST fill in from task

    # From task prompt: period range (e.g., "weeks 3-51" or "phases 8-56")
    START_PERIOD = None  # MUST fill in from task
    END_PERIOD = None    # MUST fill in from task

    # From task prompt: capacity table or rate + OT formula
    # Example formats:
    #   Direct: {6: {"total": 152, "ot": 20}, 5: {"total": 120, "ot": 10}, 4: {"total": 88, "ot": 0}}
    #   Rate-based: rate=22, ot_formula=lambda d: 10*(d-4) if d>4 else 0
    CAPACITY_RULES = None  # MUST fill in from task - see references/capacity-derivation.md

    # From task prompt: threshold for 4-day selection in normal mode
    # Often equal to 4-day capacity, but verify
    DEMAND_THRESHOLD_4DAY = None  # MUST fill in from task

    # From task prompt: backlog threshold for catch-up mode (usually 0.01)
    BACKLOG_THRESHOLD = 0.01

    # From task prompt: exact column headers for output
    OUTPUT_HEADERS = None  # MUST fill in from task

    # From task prompt: exact sheet name for output
    OUTPUT_SHEET_NAME = "Plan"  # Verify case matches task

    # ============================================================
    # VALIDATION: Ensure all required values are set
    # ============================================================
    required = {
        'INITIAL_BACKLOG': INITIAL_BACKLOG,
        'START_PERIOD': START_PERIOD,
        'END_PERIOD': END_PERIOD,
        'CAPACITY_RULES': CAPACITY_RULES,
        'DEMAND_THRESHOLD_4DAY': DEMAND_THRESHOLD_4DAY,
        'OUTPUT_HEADERS': OUTPUT_HEADERS
    }
    missing = [k for k, v in required.items() if v is None]
    if missing:
        print(f"ERROR: Required values not set: {missing}")
        print("Read the task prompt and fill in these values before running.")
        sys.exit(1)

    # ============================================================
    # STEP 2: READ INPUT DATA
    # ============================================================
    wb_in = openpyxl.load_workbook(input_path, data_only=True)

    # Detect sheet structure - adjust as needed for your task
    # Common patterns: see references/excel-patterns.md
    sheet_name = wb_in.sheetnames[0]  # Or specific name from task
    ws_in = wb_in[sheet_name]

    # Extract demand data - adjust parsing for your Excel layout
    # Pattern 1: Column-based (week, demand pairs in rows)
    # Pattern 2: Row-based (labels in col 0, data in cols 1-N)
    demands = {}  # period -> demand

    # TODO: Implement extraction based on your task's Excel layout
    # See references/excel-patterns.md for common patterns

    # Example for column-based layout:
    # for row in ws_in.iter_rows(values_only=True):
    #     if row and row[0] is not None:
    #         try:
    #             period = int(row[0])
    #             demand = float(row[1]) if row[1] is not None else 0.0
    #             demands[period] = demand
    #         except (ValueError, TypeError):
    #             pass

    # Check for duplicates
    # df = pd.DataFrame(...)  # If using pandas
    # df = df.drop_duplicates(subset=['Period'], keep='first')

    # ============================================================
    # STEP 3: ITERATIVE PLANNING CALCULATION
    # ============================================================
    periods = list(range(START_PERIOD, END_PERIOD + 1))
    plan_rows = []

    calc_start = INITIAL_BACKLOG
    first_week_5_days = None
    first_week_4_days = None

    for period in periods:
        demand = demands.get(period, 0.0)

        # Reported start (for display and mode decision)
        reported_start = max(0.0, calc_start)

        # DETERMINISTIC POLICY - do not modify
        if reported_start > BACKLOG_THRESHOLD:
            # Catch-up mode: try 5 days first, then 6
            chosen_days = 6  # default
            for days in [5, 6]:
                capacity = CAPACITY_RULES[days]['total']
                if calc_start + demand - capacity <= 0:
                    chosen_days = days
                    break
        else:
            # Normal mode: demand-based only
            if demand <= DEMAND_THRESHOLD_4DAY:
                chosen_days = 4
            else:
                chosen_days = 5

        # Track first occurrences
        if chosen_days == 5 and first_week_5_days is None:
            first_week_5_days = period
        if chosen_days == 4 and first_week_4_days is None:
            first_week_4_days = period

        # Calculate outputs
        capacity = CAPACITY_RULES[chosen_days]['total']
        overtime = CAPACITY_RULES[chosen_days]['ot']
        calc_end = calc_start + demand - capacity

        # Build row - adjust column order to match OUTPUT_HEADERS
        row = [
            period,
            chosen_days,
            demand,
            capacity,
            reported_start,
            calc_end,  # Mathematical value (can be negative)
            overtime
        ]
        plan_rows.append(row)

        # Carry state to next period
        calc_start = calc_end

    # ============================================================
    # STEP 4: WRITE EXCEL OUTPUT
    # ============================================================
    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = OUTPUT_SHEET_NAME

    ws_out.append(OUTPUT_HEADERS)
    for row in plan_rows:
        ws_out.append(row)

    wb_out.save(output_plan_path)
    print(f"Plan written to {output_plan_path} ({len(plan_rows)} rows)")

    # ============================================================
    # STEP 5: WRITE SUMMARY
    # ============================================================
    summary_lines = [
        f"First_Week_5_Days: {first_week_5_days if first_week_5_days else 'N/A'}",
        f"First_Week_4_Days: {first_week_4_days if first_week_4_days else 'N/A'}",
        # Add task-specific summary text here
    ]

    with open(output_summary_path, 'w') as f:
        f.write('\n'.join(summary_lines) + '\n')

    print(f"Summary written to {output_summary_path}")

    # ============================================================
    # STEP 6: SELF-VERIFICATION
    # ============================================================
    print("\n--- Verification ---")
    print(f"Periods: {periods[0]} to {periods[-1]} ({len(periods)} periods)")
    print(f"First 5-day week: {first_week_5_days}")
    print(f"First 4-day week: {first_week_4_days}")

    # Verify state carry
    errors = []
    for i in range(len(plan_rows) - 1):
        this_end = plan_rows[i][5]  # calc_end column
        next_start_reported = plan_rows[i+1][4]  # reported_start column
        # reported_start should be max(0, calc_start) where calc_start = this_end
        expected_next_reported = max(0.0, this_end)
        if abs(next_start_reported - expected_next_reported) > 0.001:
            errors.append(f"Period {plan_rows[i][0]}: end={this_end:.4f}, next_start={next_start_reported:.4f}, expected={expected_next_reported:.4f}")

    if errors:
        print("WARNING: State carry errors detected:")
        for e in errors[:5]:
            print(f"  {e}")
    else:
        print("State carry: OK")

if __name__ == "__main__":
    if len(sys.argv) != 4:
        print("Usage: python capacity_planner_template.py <input.xlsx> <output_plan.xlsx> <output_summary.txt>")
        print("\nIMPORTANT: Edit the script to fill in required values from your task prompt before running.")
        sys.exit(1)
    run_planning(sys.argv[1], sys.argv[2], sys.argv[3])