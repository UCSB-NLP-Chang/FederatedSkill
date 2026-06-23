#!/usr/bin/env python3
"""
BLOCKING validation script for Excel formula correctness.
Exit code 1 = BLOCKING: agent must fix before proceeding.
Exit code 0 = PASS: formulas validated.
"""

import sys
import openpyxl
import re

DEPRECATED_FUNCTIONS = [
    "QUARTILE",
    "PERCENTILE",
]

# Patterns that MUST have $ signs
REQUIRED_ABSOLUTE_PATTERNS = [
    # INDEX lookup ranges must be fully absolute
    r"INDEX\([^,]+,\s*MATCH\([^,]+,\s*([^,]+)",  # second MATCH arg = row key range
    r"INDEX\([^,]+,\s*MATCH\([^,]+,\s*[^,]+,\s*0\),\s*MATCH\([^,]+,\s*([^,]+)",  # third MATCH arg = col header range
]

def check_deprecated_functions(formula: str) -> list:
    """Check for deprecated function usage."""
    issues = []
    for func in DEPRECATED_FUNCTIONS:
        # Match function name not followed by .INC or .EXC
        pattern = rf"={func}\([^.]"
        if re.search(pattern, formula, re.IGNORECASE):
            issues.append(f"DEPRECATED: Use {func}.INC instead of {func}")
    return issues

def check_missing_dollar_in_lookup_ranges(formula: str) -> list:
    """Check INDEX/MATCH lookup ranges have $ signs."""
    issues = []

    # Extract the INDEX data range (first argument)
    idx_match = re.search(r"INDEX\(([^,]+)", formula)
    if idx_match:
        data_range = idx_match.group(1)
        # Data range should have $ in both row and column
        if not re.search(r"\$[A-Z]+\$\d+", data_range):
            issues.append(f"MISSING $: INDEX data range '{data_range}' should be fully absolute (e.g., Data!$H$21:$L$38)")

    # Extract MATCH lookup ranges
    for match_call in re.findall(r"MATCH\([^,]+,\s*([^,]+)", formula):
        # MATCH lookup range should have $ signs
        if not re.search(r"\$", match_call):
            issues.append(f"MISSING $: MATCH range '{match_call}' needs absolute references")

    return issues

def check_statistics_range_dollars(formula: str) -> list:
    """Check MIN/MAX/MEDIAN/AVERAGE/SUMPRODUCT ranges have row-absolute $."""
    issues = []

    stat_funcs = ["MIN", "MAX", "MEDIAN", "AVERAGE", "SUMPRODUCT", "SUM", "PERCENTILE.INC", "QUARTILE.INC"]

    for func in stat_funcs:
        pattern = rf"={func}\(([^)]+)\)"
        matches = re.findall(pattern, formula, re.IGNORECASE)
        for match in matches:
            # Check ranges in the argument
            ranges = re.findall(r"[A-Z]+\d+:[A-Z]+\d+", match)
            for rng in ranges:
                # Row numbers should have $ (row-absolute)
                if not re.search(r"\$\d+", rng):
                    issues.append(f"MISSING $: {func} range '{rng}' should be row-absolute (e.g., H$35:H$40)")

    return issues

def validate_workbook(path: str, target_sheet: str = None) -> list:
    """Validate all formulas in workbook or specific sheet."""
    wb = openpyxl.load_workbook(path, data_only=False)
    all_issues = []

    sheets_to_check = [target_sheet] if target_sheet else wb.sheetnames

    for sheet_name in sheets_to_check:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    formula = cell.value

                    # Check deprecated functions
                    issues = check_deprecated_functions(formula)
                    all_issues.extend([f"{sheet_name}:{cell.coordinate}: {i}" for i in issues])

                    # Check missing $ in lookup ranges
                    issues = check_missing_dollar_in_lookup_ranges(formula)
                    all_issues.extend([f"{sheet_name}:{cell.coordinate}: {i}" for i in issues])

                    # Check statistics range dollars
                    issues = check_statistics_range_dollars(formula)
                    all_issues.extend([f"{sheet_name}:{cell.coordinate}: {i}" for i in issues])

    wb.close()
    return all_issues

def main():
    if len(sys.argv) < 2:
        print("Usage: python3 validate_formulas.py <workbook_path> [target_sheet]")
        print("BLOCKING: This script exits with code 1 on validation failures.")
        sys.exit(1)

    path = sys.argv[1]
    target_sheet = sys.argv[2] if len(sys.argv) > 2 else None

    print(f"=== VALIDATING FORMULAS IN {path} ===")

    issues = validate_workbook(path, target_sheet)

    if issues:
        print("\n*** BLOCKING ERRORS DETECTED ***")
        for issue in issues:
            print(f"  {issue}")
        print("\nFIX THESE ISSUES BEFORE SAVING. Script exiting with code 1.")
        sys.exit(1)
    else:
        print("\n=== VALIDATION PASSED ===")
        print("No deprecated functions or missing $ signs detected.")
        sys.exit(0)

if __name__ == "__main__":
    main()