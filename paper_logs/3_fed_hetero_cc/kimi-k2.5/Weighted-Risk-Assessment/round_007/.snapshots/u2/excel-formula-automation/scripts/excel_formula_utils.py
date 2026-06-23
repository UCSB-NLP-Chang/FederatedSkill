#!/usr/bin/env python3
"""Utility functions for Excel formula automation with openpyxl."""

import openpyxl
from openpyxl.utils import get_column_letter
from typing import Optional, List


def inspect_workbook(path: str) -> dict:
    """Load and return key structural information about a workbook."""
    wb = openpyxl.load_workbook(path, data_only=False)
    info = {
        "sheets": wb.sheetnames,
        "active_sheet": wb.active.title,
    }
    wb.close()
    return info


def inspect_sheet(wb: openpyxl.Workbook, sheet_name: str, max_row: int = 50, max_col: int = 15) -> List[List]:
    """Return cell values for a sheet up to max_row/max_col for inspection."""
    ws = wb[sheet_name]
    data = []
    for row in range(1, min(max_row + 1, ws.max_row + 1)):
        row_data = []
        for col in range(1, min(max_col + 1, ws.max_column + 1)):
            cell = ws.cell(row=row, column=col)
            row_data.append(cell.value)
        data.append(row_data)
    return data


def get_cell_fill(ws: openpyxl.Worksheet, row: int, col: int) -> Optional[str]:
    """Get the fill color of a cell as hex string, or None if no fill."""
    cell = ws.cell(row=row, column=col)
    if cell.fill and cell.fill.fgColor:
        return cell.fill.fgColor.rgb
    return None


def count_formulas(ws: openpyxl.Worksheet) -> int:
    """Count cells containing formulas in a worksheet."""
    return sum(1 for row in ws.iter_rows() for cell in row
               if isinstance(cell.value, str) and cell.value.startswith('='))


def find_header_row(ws: openpyxl.Worksheet, search_cols: list, max_row: int = 30) -> int:
    """
    Find the header row by searching for known header values in specified columns.
    Returns row number (1-indexed) or 0 if not found.
    """
    for row in range(1, min(max_row + 1, ws.max_row + 1)):
        for col in search_cols:
            cell = ws.cell(row=row, column=col)
            if cell.value and isinstance(cell.value, str):
                # Common header patterns
                if any(h in cell.value.upper() for h in ["YEAR", "CODE", "ID", "NAME", "SERIES"]):
                    return row
    return 0


def get_column_values(ws: openpyxl.Worksheet, col: int, start_row: int, end_row: int) -> list:
    """Get all values in a column range."""
    return [ws.cell(row=r, column=col).value for r in range(start_row, end_row + 1)]


def create_index_match_formula(
    data_range: str,
    row_key_cell: str,
    row_key_range: str,
    col_key_cell: str,
    col_key_range: str,
) -> str:
    """
    Generate an INDEX/MATCH/MATCH formula for 2D lookup.
    IMPORTANT: All ranges should use $ for absolute references.
    """
    return f"=INDEX({data_range},MATCH({row_key_cell},{row_key_range},0),MATCH({col_key_cell},{col_key_range},0))"