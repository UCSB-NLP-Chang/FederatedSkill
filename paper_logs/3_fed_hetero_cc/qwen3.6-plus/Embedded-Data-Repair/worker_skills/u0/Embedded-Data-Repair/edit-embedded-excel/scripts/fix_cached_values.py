#!/usr/bin/env python3
"""
Fix cached formula values in an Excel file after openpyxl edits.

openpyxl preserves formulas but may leave cached <v> values empty or stale.
This script patches the worksheet XML to set correct cached values.

Usage:
    # Auto-detect and fix ROUND formulas (recommended)
    python3 fix_cached_values.py <input.xlsx> <output.xlsx> --auto

    # Fix specific cells manually
    python3 fix_cached_values.py <input.xlsx> <output.xlsx> G4=2 E5=0.5

    # Detect empty cached values only
    python3 fix_cached_values.py <input.xlsx> --detect-only

Example:
    python3 fix_cached_values.py /tmp/embedded.xlsx /tmp/fixed.xlsx --auto
"""

import argparse
import zipfile
import re
import io


def detect_empty_cached_values(xlsx_path):
    """
    Detect formula cells with empty cached values.

    Returns:
        empty_cells: List of cell refs with empty <v /> tags
        formula_cells: List of (cell_ref, formula) tuples for all formula cells
    """
    empty_cells = []
    formula_cells = []

    with zipfile.ZipFile(xlsx_path, 'r') as z:
        for name in z.namelist():
            if name.startswith('xl/worksheets/sheet'):
                xml = z.read(name).decode('utf-8')
                
                # Find formula cells with empty cached values
                # Handles both self-closing <v /> and empty <v></v>
                # Also handles optional attributes on <v> tag like <v t="str">
                empty_matches = re.findall(
                    r'<(?:s:)?c r="([A-Z]+\d+)"[^>]*>.*?<(?:s:)?f>[^<]+</(?:s:)?f>.*?<(?:s:)?v[^>]*\s*/?>\s*</(?:s:)?v>.*?</(?:s:)?c>',
                    xml, re.DOTALL
                )
                # Filter to only truly empty values
                for cell_ref in empty_matches:
                    # Verify the cached value is actually empty
                    v_match = re.search(
                        rf'<(?:s:)?c r="{cell_ref}"[^>]*>.*?<(?:s:)?f>[^<]+</(?:s:)?f>.*?<(?:s:)?v[^>]*>([^<]*)</(?:s:)?v>',
                        xml, re.DOTALL
                    )
                    if v_match and not v_match.group(1).strip():
                        empty_cells.append(cell_ref)
                
                # Find all formula cells with their formulas (namespace-aware)
                all_formulas = re.findall(
                    r'<(?:s:)?c r="([A-Z]+\d+)"[^>]*>.*?<(?:s:)?f>([^<]+)</(?:s:)?f>',
                    xml, re.DOTALL
                )
                formula_cells.extend(all_formulas)

    return empty_cells, formula_cells


def parse_round_formula(formula):
    """
    Parse ROUND(1/X, N) formula and return (source_cell, precision).

    Supports:
    - ROUND(1/A1, 4)
    - ROUND(1/$A$1, 2)
    """
    match = re.match(r'ROUND\(1/\$?([A-Z]+)\$?(\d+),\s*(\d+)\)', formula)
    if match:
        return f"{match.group(1)}{match.group(2)}", int(match.group(3))
    return None, None


def auto_fix_round_formulas(input_xlsx, output_xlsx):
    """
    Auto-detect ROUND(1/X, N) formulas and fix their cached values.

    This is the recommended approach for reciprocal rate matrices.
    """
    try:
        import openpyxl
    except ImportError:
        raise RuntimeError("openpyxl required for --auto mode. Run: pip install openpyxl")

    # Load workbook to get cell values
    wb = openpyxl.load_workbook(input_xlsx)

    # Detect formulas and empty cells
    empty_cells, formula_cells = detect_empty_cached_values(input_xlsx)

    if not empty_cells:
        print("No empty cached values found.")
        return False

    print(f"Empty cached values detected: {empty_cells}")

    # Process each sheet
    all_cell_values = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        for cell_ref, formula in formula_cells:
            if cell_ref in empty_cells:
                source_cell, precision = parse_round_formula(formula)
                if source_cell:
                    source_value = ws[source_cell].value
                    if source_value is not None and isinstance(source_value, (int, float)):
                        if source_value == 0:
                            print(f"Warning: {source_cell} is zero, skipping {cell_ref}")
                            continue
                        calculated = round(1 / source_value, precision)
                        all_cell_values[cell_ref] = calculated
                        print(f"  {cell_ref}: formula={formula}, source={source_cell}={source_value}, cached={calculated}")

    if all_cell_values:
        fix_cached_values(input_xlsx, output_xlsx, all_cell_values)
        return True
    else:
        print("No auto-fixable ROUND formulas found.")
        return False


def fix_cached_values(input_xlsx, output_xlsx, cell_values):
    """
    Fix cached values for formula cells in an Excel file.

    Args:
        input_xlsx: Path to input Excel file
        output_xlsx: Path to output Excel file
        cell_values: Dict mapping cell refs to computed values
    """
    # Read all entries from input
    with zipfile.ZipFile(input_xlsx, 'r') as zin:
        entries = {name: zin.read(name) for name in zin.namelist()}

    # Fix each worksheet
    for sheet_name in [n for n in entries if n.startswith('xl/worksheets/sheet')]:
        xml = entries[sheet_name].decode('utf-8')
        modified = False

        for cell_ref, value in cell_values.items():
            # Pattern handles both self-closing <v /> and regular <v>...</v>
            # Also handles optional attributes on <v> tag
            pattern = rf'(<(?:s:)?c r="{cell_ref}"[^>]*>.*?<(?:s:)?f>[^<]*</(?:s:)?f>.*?<(?:s:)?v[^>]*>)[^<]*(</(?:s:)?v>.*?</(?:s:)?c>)'
            if re.search(pattern, xml, re.DOTALL):
                replacement = rf'\g<1>{value}\g<2>'
                xml = re.sub(pattern, replacement, xml, flags=re.DOTALL)
                modified = True
                print(f"Fixed cached value for {cell_ref} -> {value}")

        if modified:
            entries[sheet_name] = xml.encode('utf-8')

    # Write output
    with zipfile.ZipFile(output_xlsx, 'w', zipfile.ZIP_DEFLATED) as zout:
        for name, data in entries.items():
            zout.writestr(name, data)

    print(f"Saved to: {output_xlsx}")


def main():
    parser = argparse.ArgumentParser(
        description='Fix cached formula values in Excel file',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""Examples:
  # Auto-detect and fix ROUND formulas (recommended)
  python3 fix_cached_values.py input.xlsx output.xlsx --auto

  # Fix specific cells manually
  python3 fix_cached_values.py input.xlsx output.xlsx G4=2 E5=0.5

  # Detect empty cached values
  python3 fix_cached_values.py input.xlsx --detect-only
"""
    )
    parser.add_argument('input_xlsx', help='Input Excel file')
    parser.add_argument('output_xlsx', nargs='?', help='Output Excel file')
    parser.add_argument('cells', nargs='*', help='Cell=value pairs (e.g., G4=2)')
    parser.add_argument('--detect-only', action='store_true',
                        help='Only detect empty cached values, do not fix')
    parser.add_argument('--auto', action='store_true',
                        help='Auto-detect ROUND(1/X, N) formulas and calculate values')

    args = parser.parse_args()

    if args.detect_only:
        empty, formulas = detect_empty_cached_values(args.input_xlsx)
        if empty:
            print(f"Formula cells with empty cached values: {empty}")
            print(f"All formulas: {formulas}")
        else:
            print("No empty cached values found.")
        return

    if not args.output_xlsx:
        parser.error("output_xlsx required when not using --detect-only")

    if args.auto:
        auto_fix_round_formulas(args.input_xlsx, args.output_xlsx)
        return

    if not args.cells:
        parser.error("cell=value pairs required when not using --auto or --detect-only")

    # Parse cell=value pairs
    cell_values = {}
    for cv in args.cells:
        if '=' not in cv:
            parser.error(f"Invalid format '{cv}'. Use CELL=VALUE format (e.g., G4=2)")
        cell, value = cv.split('=', 1)
        # Try to convert to number
        try:
            value = float(value)
            if value == int(value):
                value = int(value)
        except ValueError:
            pass  # Keep as string
        cell_values[cell.upper()] = value

    fix_cached_values(args.input_xlsx, args.output_xlsx, cell_values)


if __name__ == '__main__':
    main()
