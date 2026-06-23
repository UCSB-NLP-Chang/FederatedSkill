#!/usr/bin/env python3
"""
Deterministic inventory coverage calculator.
Reads a multi-sheet Excel workbook, filters bookings, calculates gaps, and writes output.
Usage: python3 calculate_coverage.py <input.xlsx> <output.xlsx>
"""
import sys
import math
from datetime import datetime, date, timedelta
import openpyxl

def normalize_date(val):
    """Convert datetime or date to date. Returns None if invalid."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    return None

def main(input_path, output_path):
    wb = openpyxl.load_workbook(input_path, data_only=True)

    # 1. Extract Metadata
    meta = wb['Rack Snapshot']
    as_of = normalize_date(meta['B1'].value)
    horizon = normalize_date(meta['D1'].value)
    planning_days = (horizon - as_of).days

    pallet_size = wb['Pallet Defaults']['A2'].value

    # 2. Parse SKUs
    skus = {}
    for row in meta.iter_rows(min_row=4, values_only=True):
        sku = row[0]
        if sku:
            skus[sku] = {'on_rack': row[1] or 0, 'pull': row[2] or 0}

    # 3. Parse & Filter Bookings
    bookings = {}
    for row in wb['Booking Feed'].iter_rows(min_row=2, values_only=True):
        sku, eta_raw, cases, state = row[0], row[1], row[2], row[3]
        if not sku or state in ('Tentative', 'Hold', 'Cancelled'):
            continue
        eta = normalize_date(eta_raw)
        if eta is None or eta > horizon:
            continue
        bookings.setdefault(sku, []).append({'eta': eta, 'cases': cases or 0})

    # 4. Calculate Coverage
    coverage_rows = []
    gap_rows = []

    for sku, data in skus.items():
        on_rack = data['on_rack']
        pull = data['pull']
        days_on_hand = on_rack / pull if pull else float('inf')
        oos_date = as_of + timedelta(days=days_on_hand)

        sku_bookings = bookings.get(sku, [])
        booked_cases = sum(b['cases'] for b in sku_bookings)
        earliest_eta = min((b['eta'] for b in sku_bookings), default=None)

        delivered_days = (on_rack + booked_cases) / pull if pull else float('inf')
        remaining_demand = pull * planning_days
        additional = max(0, remaining_demand - (on_rack + booked_cases))
        pallets = math.ceil(additional / pallet_size) if additional > 0 else 0

        req_delivery = oos_date if additional > 0 else None
        earlier_req = False
        if req_delivery:
            if earliest_eta is None or earliest_eta > req_delivery:
                earlier_req = True

        coverage_rows.append([
            sku, on_rack, pull, days_on_hand, oos_date,
            booked_cases, delivered_days, remaining_demand,
            additional, pallets, req_delivery, earlier_req
        ])

        if additional > 0:
            gap_rows.append([sku, req_delivery, pallets, additional, earlier_req])

    # 5. Write Output
    out = openpyxl.Workbook()

    # Rack_Coverage
    ws1 = out.active
    ws1.title = 'Rack_Coverage'
    ws1.append(['Field', 'Value'])
    ws1.append(['AsOfDate', as_of])
    ws1.append(['HorizonEnd', horizon])
    ws1.append(['PlanningDays', planning_days])
    ws1.append([])
    ws1.append(['SKU_Ref', 'Cases_On_Rack', 'Avg_Daily_Pull_Cases', 'Current_Days_On_Hand',
                'Projected_OOS_Date', 'Booked_Cases_By_Horizon', 'Delivered_Days_On_Hand',
                'Remaining_Demand_Cases', 'Additional_Cases_Needed', 'Pallets_Required',
                'Required_Delivery_Date', 'Earlier_Delivery_Required'])
    for r in coverage_rows:
        ws1.append(r)

    # Commit_Gap_Actions
    ws2 = out.create_sheet('Commit_Gap_Actions')
    ws2.append(['SKU_Ref', 'Required_Delivery_Date', 'Pallets_Required', 'Additional_Cases_Needed', 'Earlier_Delivery_Required'])
    for r in gap_rows:
        ws2.append(r)

    out.save(output_path)
    print(f"Saved {output_path}")

if __name__ == '__main__':
    if len(sys.argv) != 3:
        print("Usage: python3 calculate_coverage.py <input.xlsx> <output.xlsx>")
        sys.exit(1)
    main(sys.argv[1], sys.argv[2])