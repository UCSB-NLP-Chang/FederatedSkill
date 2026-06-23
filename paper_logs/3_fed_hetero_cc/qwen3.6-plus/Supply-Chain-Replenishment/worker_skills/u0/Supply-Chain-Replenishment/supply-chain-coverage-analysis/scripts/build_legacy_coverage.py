#!/usr/bin/env python3
"""Legacy ratio-based coverage builder.
Usage: python3 build_legacy_coverage.py <inventory.xlsx> <output.xlsx>

Handles workbooks with 'Current Inventory', 'Incoming Shipments', and 'Ratio' sheets.
Reads load size from Ratio sheet, uses data_only=True for formula cells,
and outputs Rounding_Applied + Earliest_Scheduled_Inbound_Date columns.
"""
import sys
import openpyxl
import math
from datetime import datetime, date, timedelta


def parse_date(val):
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    try:
        return datetime.strptime(str(val).strip(), '%Y-%m-%d').date()
    except (ValueError, TypeError):
        return None


def main():
    if len(sys.argv) != 3:
        print("Usage: python3 build_legacy_coverage.py <inventory.xlsx> <output.xlsx>")
        sys.exit(1)

    input_path, output_path = sys.argv[1:]

    wb = openpyxl.load_workbook(input_path, data_only=True)

    inv_ws = wb['Current Inventory']
    ship_ws = wb['Incoming Shipments']
    ratio_ws = wb['Ratio']

    # Metadata
    as_of_date = parse_date(inv_ws['B1'].value)
    horizon_end = parse_date(inv_ws['D1'].value)
    planning_days = (horizon_end - as_of_date).days

    # Load size from Ratio sheet
    load_size = int(ratio_ws['A2'].value)

    # Stock
    stock_rows = []
    for row in inv_ws.iter_rows(min_row=4, values_only=True):
        sku, on_hand, daily_demand = row[0], row[1], row[2]
        if sku is None:
            break
        stock_rows.append({
            'SKU': str(sku).strip(),
            'On_Hand': on_hand if on_hand is not None else 0,
            'Daily_Demand': daily_demand if daily_demand is not None else 0
        })

    # Inbound shipments
    inbound = {}
    earliest_dates = {}
    for row in ship_ws.iter_rows(min_row=2, values_only=True):
        sku, delivery_date, pallets, cases = row[0], row[1], row[2], row[3]
        if sku is None:
            continue
        sku = str(sku).strip()
        eta = parse_date(delivery_date)
        if eta is None:
            continue
        units = cases if cases is not None else 0

        if sku not in inbound:
            inbound[sku] = 0
            earliest_dates[sku] = eta
        inbound[sku] += units
        if eta < earliest_dates[sku]:
            earliest_dates[sku] = eta

    # Coverage calculation
    coverage = []
    for s in stock_rows:
        sku = s['SKU']
        on_hand = s['On_Hand']
        daily = s['Daily_Demand']

        days_on_hand = on_hand / daily
        oos_date = as_of_date + timedelta(days=days_on_hand)

        inbound_units = inbound.get(sku, 0)
        delivered_days = inbound_units / daily
        remaining_demand = daily * planning_days
        additional_needed = max(0, remaining_demand - on_hand - inbound_units)
        pallets = math.ceil(additional_needed / load_size) if additional_needed > 0 else 0

        req_date = oos_date if oos_date <= horizon_end else horizon_end

        earliest_inbound = earliest_dates.get(sku)
        earlier_req = False
        if pallets > 0:
            if inbound_units == 0:
                earlier_req = True
            elif earliest_inbound and earliest_inbound > req_date:
                earlier_req = True

        rounding_applied = (additional_needed % load_size != 0) if additional_needed > 0 else False

        coverage.append({
            'Product_SKU': sku,
            'Current_Cases': on_hand,
            'Daily_Rate_Cases_Per_Day': daily,
            'Current_DOH': days_on_hand,
            'Projected_OOS_Date': oos_date,
            'Inbound_Cases_By_Horizon': inbound_units,
            'Delivered_DOH_To_Horizon': delivered_days,
            'Remaining_Demand_Cases': remaining_demand,
            'Additional_Cases_Needed': additional_needed,
            'Pallets_Required_Rounded_Up': pallets,
            'Required_Delivery_Date': req_date,
            'Rounding_Applied': rounding_applied,
            'Earlier_Delivery_Required': earlier_req,
            'Earliest_Scheduled_Inbound_Date': earliest_inbound
        })

    # Output
    out_wb = openpyxl.Workbook()

    ws_cov = out_wb.active
    ws_cov.title = 'SKU_Results'
    ws_cov.append(['Field', 'Value'])
    ws_cov.append(['AsOfDate', as_of_date])
    ws_cov.append(['PlanningHorizonEnd', horizon_end])
    ws_cov.append(['RemainingDaysInJuly', planning_days])
    ws_cov.append([])

    headers = ['Product_SKU', 'Current_Cases', 'Daily_Rate_Cases_Per_Day',
               'Current_DOH', 'Projected_OOS_Date', 'Inbound_Cases_By_Horizon',
               'Delivered_DOH_To_Horizon', 'Remaining_Demand_Cases',
               'Additional_Cases_Needed', 'Pallets_Required_Rounded_Up',
               'Required_Delivery_Date', 'Rounding_Applied',
               'Earlier_Delivery_Required', 'Earliest_Scheduled_Inbound_Date']
    ws_cov.append(headers)
    for c in coverage:
        ws_cov.append([c[h] for h in headers])

    ws_gap = out_wb.create_sheet('Additional_Shipments_Needed')
    gap_headers = ['Product_SKU', 'Required_Delivery_Date',
                   'Pallets_Required_Rounded_Up', 'Additional_Cases_Needed',
                   'Rounding_Applied', 'Earlier_Delivery_Required']
    ws_gap.append(gap_headers)
    for c in coverage:
        if c['Pallets_Required_Rounded_Up'] > 0:
            ws_gap.append([c[h] for h in gap_headers])

    out_wb.save(output_path)
    print(f"Done: {output_path}")


if __name__ == '__main__':
    main()
