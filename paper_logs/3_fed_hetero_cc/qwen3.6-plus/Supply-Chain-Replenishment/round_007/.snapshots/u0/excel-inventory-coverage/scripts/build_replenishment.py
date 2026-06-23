#!/usr/bin/env python3
"""Replenishment gap analysis builder.
Usage: python3 build_replenishment.py <stock.xlsx> <feed.xlsx> <alias.xlsx> <output.xlsx>
"""
import sys
import openpyxl
import math
from datetime import datetime, date, timedelta

PALLET_SIZE = 36

def parse_date(val):
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    try:
        return datetime.strptime(str(val).strip(), '%Y-%m-%d').date()
    except (ValueError, TypeError):
        return None

def main():
    if len(sys.argv) != 5:
        print("Usage: python3 build_replenishment.py <stock.xlsx> <feed.xlsx> <alias.xlsx> <output.xlsx>")
        sys.exit(1)

    stock_path, feed_path, alias_path, output_path = sys.argv[1:]

    stock_wb = openpyxl.load_workbook(stock_path, data_only=True)
    feed_wb = openpyxl.load_workbook(feed_path, data_only=True)
    alias_wb = openpyxl.load_workbook(alias_path, data_only=True)

    stock_ws = stock_wb['Zone Snapshot']
    feed_ws = feed_wb['Zone Feed']
    alias_ws = alias_wb['Alias Map']

    # Metadata
    as_of_date = parse_date(stock_ws['B1'].value)
    horizon_end = parse_date(stock_ws['D1'].value)
    planning_days = (horizon_end - as_of_date).days

    # Stock
    stock_rows = []
    for row in stock_ws.iter_rows(min_row=4, values_only=True):
        zone, sku, on_hand, daily_demand = row
        if zone is None:
            break
        stock_rows.append({'Zone': zone, 'SKU': sku, 'On_Hand': on_hand, 'Daily_Demand': daily_demand})

    # Alias Map
    alias_map = {}
    for row in alias_ws.iter_rows(min_row=2, values_only=True):
        alias_val, canonical = row
        if alias_val is None:
            break
        alias_map[alias_val] = canonical

    # Feed Processing
    raw_feed = []
    for row in feed_ws.iter_rows(min_row=2, values_only=True):
        rec_type, dispatch_ref, revision, zone_alias, sku_code, eta, units, release_state = row
        if rec_type != 'DELIVERY':
            continue
        if sku_code is None or str(sku_code).strip() == '':
            continue
        eta_date = parse_date(eta)
        if eta_date is None:
            continue
        raw_feed.append({
            'Dispatch_Ref': dispatch_ref,
            'Revision': revision if revision is not None else 0,
            'Zone_Alias': zone_alias,
            'SKU_Code': sku_code,
            'ETA': eta_date,
            'Units': units,
            'Release_State': release_state,
        })

    # Dedup by Dispatch_Ref (keep highest revision)
    deduped = {}
    for r in raw_feed:
        key = r['Dispatch_Ref']
        if key not in deduped or r['Revision'] > deduped[key]['Revision']:
            deduped[key] = r

    # Filter & Map
    qualifying = []
    for r in deduped.values():
        if r['Release_State'] not in ('Released', 'Staged'):
            continue
        canonical_zone = alias_map.get(r['Zone_Alias'])
        if canonical_zone is None:
            continue
        if r['ETA'] > horizon_end:
            continue
        qualifying.append({**r, 'Canonical_Zone': canonical_zone})

    # Coverage Calculation
    coverage = []
    for s in stock_rows:
        zone, sku = s['Zone'], s['SKU']
        on_hand = s['On_Hand']
        daily = s['Daily_Demand']

        days_on_hand = on_hand / daily
        oos_date = as_of_date + timedelta(days=days_on_hand)

        inbound_units = sum(r['Units'] for r in qualifying if r['Canonical_Zone'] == zone and r['SKU_Code'] == sku)
        delivered_days = inbound_units / daily
        remaining_demand = daily * planning_days
        additional_needed = max(0, remaining_demand - on_hand - inbound_units)
        pallets = math.ceil(additional_needed / PALLET_SIZE) if additional_needed > 0 else 0

        req_date = oos_date if oos_date <= horizon_end else horizon_end

        earliest_inbound = min((r['ETA'] for r in qualifying if r['Canonical_Zone'] == zone and r['SKU_Code'] == sku), default=None)
        earlier_req = False
        if pallets > 0:
            if inbound_units == 0:
                earlier_req = True
            elif earliest_inbound and earliest_inbound > req_date:
                earlier_req = True

        coverage.append({
            'Zone': zone, 'SKU': sku,
            'Units_On_Hand': on_hand, 'Daily_Demand_Units_Per_Day': daily,
            'Current_Days_On_Hand': days_on_hand, 'Projected_OOS_Date': oos_date,
            'Inbound_Units_By_Horizon': inbound_units, 'Delivered_Days_On_Hand': delivered_days,
            'Remaining_Demand_Units': remaining_demand, 'Additional_Units_Needed': additional_needed,
            'Pallets_Required': pallets, 'Required_Delivery_Date': req_date,
            'Earlier_Delivery_Required': earlier_req
        })

    # Output
    out_wb = openpyxl.Workbook()

    ws_cov = out_wb.active
    ws_cov.title = 'Zone_Coverage'
    ws_cov.append(['Field', 'Value'])
    ws_cov.append(['AsOfDate', as_of_date])
    ws_cov.append(['HorizonEnd', horizon_end])
    ws_cov.append(['PlanningDays', planning_days])
    ws_cov.append([])
    headers = ['Zone', 'SKU', 'Units_On_Hand', 'Daily_Demand_Units_Per_Day', 'Current_Days_On_Hand',
               'Projected_OOS_Date', 'Inbound_Units_By_Horizon', 'Delivered_Days_On_Hand',
               'Remaining_Demand_Units', 'Additional_Units_Needed', 'Pallets_Required',
               'Required_Delivery_Date', 'Earlier_Delivery_Required']
    ws_cov.append(headers)
    for c in coverage:
        ws_cov.append([c[h] for h in headers])

    ws_gap = out_wb.create_sheet('Dispatch_Gap_List')
    gap_headers = ['Zone', 'SKU', 'Required_Delivery_Date', 'Pallets_Required', 'Additional_Units_Needed', 'Earlier_Delivery_Required']
    ws_gap.append(gap_headers)
    for c in coverage:
        if c['Pallets_Required'] > 0:
            ws_gap.append([c[h] for h in gap_headers])

    out_wb.save(output_path)
    print(f"Done: {output_path}")

if __name__ == '__main__':
    main()
