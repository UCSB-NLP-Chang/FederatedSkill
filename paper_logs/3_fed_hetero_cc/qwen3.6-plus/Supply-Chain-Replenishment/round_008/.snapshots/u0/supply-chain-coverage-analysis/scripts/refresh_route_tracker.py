#!/usr/bin/env python3
"""Route dispatch tracker refresh.
Usage: python3 refresh_route_tracker.py <route_snapshot.xlsx> <queue_export.xlsx> <template.xlsx> <output.xlsx>
"""
import sys
import openpyxl
import math
from datetime import datetime, date, timedelta


def parse_date(val):
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    try:
        return datetime.strptime(str(val).strip(), '%Y-%m-%d').date()
    except (ValueError, TypeError):
        return None


def main():
    if len(sys.argv) != 5:
        print("Usage: python3 refresh_route_tracker.py <route_snapshot.xlsx> <queue_export.xlsx> <template.xlsx> <output.xlsx>")
        sys.exit(1)

    snapshot_path, queue_path, template_path, output_path = sys.argv[1:]

    # Load template workbook (preserve structure)
    out_wb = openpyxl.load_workbook(template_path)

    # Load read-only copies for input data
    snap_wb = openpyxl.load_workbook(snapshot_path, data_only=True)
    queue_wb = openpyxl.load_workbook(queue_path, data_only=True)

    snap_ws = snap_wb['Route Snapshot']
    queue_ws = queue_wb['Queue Export']

    # 1. Parse Route Snapshot metadata
    as_of_date = parse_date(snap_ws['B1'].value)
    horizon_end = parse_date(snap_ws['D1'].value)
    planning_days = (horizon_end - as_of_date).days

    # 2. Parse Route Snapshot stock (section-based)
    stock_rows = []
    current_route = None
    for row in snap_ws.iter_rows(min_row=3, values_only=True):
        cell_a = row[0]
        if cell_a and isinstance(cell_a, str) and cell_a.startswith('Route '):
            current_route = cell_a.replace('Route ', '').strip()
            continue
        if current_route and cell_a is not None:
            sku = str(cell_a).strip()
            on_hand = row[1] if row[1] is not None else 0
            daily_demand = row[2] if row[2] is not None else 0
            stock_rows.append({
                'Route': current_route, 'SKU': sku,
                'On_Hand': on_hand, 'Daily_Demand': daily_demand
            })

    # 3. Load Pack Matrix and Route Alias Map from template
    pack_ws = out_wb['Pack Matrix']
    pack_matrix = {}
    for row in pack_ws.iter_rows(min_row=2, values_only=True):
        route, sku, cases_per_load = row
        if route and sku and cases_per_load is not None:
            pack_matrix[(str(route).strip(), str(sku).strip())] = cases_per_load

    alias_ws = out_wb['Route Alias Map']
    alias_map = {}
    for row in alias_ws.iter_rows(min_row=2, values_only=True):
        alias_val, canonical = row
        if alias_val and canonical:
            alias_map[str(alias_val).strip()] = str(canonical).strip()

    # 4. Process Queue Export
    raw_feed = []
    for row in queue_ws.iter_rows(min_row=2, values_only=True):
        row_type = row[0]
        if row_type != 'DISPATCH':
            continue
        dispatch_ref = row[1]
        revision = row[2] if row[2] is not None else 0
        route_alias = row[3]
        sku_code = row[4]
        ship_date = row[5]
        units = row[6]
        queue_state = row[7]

        if sku_code is None or str(sku_code).strip() == '':
            continue
        eta_date = parse_date(ship_date)
        if eta_date is None:
            continue

        raw_feed.append({
            'Dispatch_Ref': dispatch_ref,
            'Revision': revision,
            'Route_Alias': route_alias,
            'SKU_Code': str(sku_code).strip(),
            'ETA': eta_date,
            'Units': units if units is not None else 0,
            'Queue_State': queue_state,
        })

    # Dedup by Dispatch_Ref (keep highest revision)
    deduped = {}
    for r in raw_feed:
        key = r['Dispatch_Ref']
        if key not in deduped or r['Revision'] > deduped[key]['Revision']:
            deduped[key] = r

    # Filter qualifying rows
    qualifying = []
    for r in deduped.values():
        if r['Queue_State'] not in ('Released', 'Approved'):
            continue
        canonical_route = alias_map.get(r['Route_Alias'])
        if canonical_route is None:
            continue
        if r['ETA'] > horizon_end:
            continue
        qualifying.append({**r, 'Canonical_Route': canonical_route})

    # 5. Calculate coverage
    coverage = []
    for s in stock_rows:
        route, sku = s['Route'], s['SKU']
        on_hand = s['On_Hand']
        daily = s['Daily_Demand']

        days_on_hand = on_hand / daily
        oos_date = as_of_date + timedelta(days=days_on_hand)

        inbound_units = sum(
            r['Units'] for r in qualifying
            if r['Canonical_Route'] == route and r['SKU_Code'] == sku
        )
        delivered_days = inbound_units / daily
        remaining_demand = daily * planning_days
        additional_needed = max(0, remaining_demand - on_hand - inbound_units)

        # Variable load size from Pack Matrix
        cases_per_load = pack_matrix.get((route, sku), 36)
        loads = math.ceil(additional_needed / cases_per_load) if additional_needed > 0 else 0

        req_date = oos_date if oos_date <= horizon_end else horizon_end

        earliest_inbound = min(
            (r['ETA'] for r in qualifying
             if r['Canonical_Route'] == route and r['SKU_Code'] == sku),
            default=None
        )
        earlier_req = False
        if loads > 0:
            if inbound_units == 0:
                earlier_req = True
            elif earliest_inbound and earliest_inbound > req_date:
                earlier_req = True

        coverage.append({
            'Route': route, 'SKU': sku,
            'Units_On_Hand': on_hand, 'Daily_Demand_Units_Per_Day': daily,
            'Current_Days_On_Hand': days_on_hand, 'Projected_OOS_Date': oos_date,
            'Inbound_Units_By_Horizon': inbound_units,
            'Delivered_Days_On_Hand': delivered_days,
            'Remaining_Demand_Units': remaining_demand,
            'Additional_Units_Needed': additional_needed,
            'Loads_Required': loads,
            'Required_Delivery_Date': req_date,
            'Earlier_Delivery_Required': earlier_req
        })

    # 6. Refresh Coverage_Detail sheet
    if 'Coverage_Detail' in out_wb.sheetnames:
        del out_wb['Coverage_Detail']
    ws_cov = out_wb.create_sheet('Coverage_Detail')
    ws_cov.append(['Field', 'Value'])
    ws_cov.append(['AsOfDate', as_of_date])
    ws_cov.append(['HorizonEnd', horizon_end])
    ws_cov.append(['PlanningDays', planning_days])
    ws_cov.append([])
    cov_headers = ['Route', 'SKU', 'Units_On_Hand', 'Daily_Demand_Units_Per_Day',
                   'Current_Days_On_Hand', 'Projected_OOS_Date',
                   'Inbound_Units_By_Horizon', 'Delivered_Days_On_Hand',
                   'Remaining_Demand_Units', 'Additional_Units_Needed',
                   'Loads_Required', 'Required_Delivery_Date',
                   'Earlier_Delivery_Required']
    ws_cov.append(cov_headers)
    for c in coverage:
        ws_cov.append([c[h] for h in cov_headers])

    # 7. Refresh Dispatch_Plan sheet
    if 'Dispatch_Plan' in out_wb.sheetnames:
        del out_wb['Dispatch_Plan']
    ws_plan = out_wb.create_sheet('Dispatch_Plan')
    plan_headers = ['Route', 'SKU', 'Required_Delivery_Date', 'Loads_Required',
                    'Additional_Units_Needed', 'Earlier_Delivery_Required']
    ws_plan.append(plan_headers)
    for c in coverage:
        if c['Loads_Required'] > 0:
            ws_plan.append([c[h] for h in plan_headers])

    out_wb.save(output_path)
    print(f"Done: {output_path}")


if __name__ == '__main__':
    main()
