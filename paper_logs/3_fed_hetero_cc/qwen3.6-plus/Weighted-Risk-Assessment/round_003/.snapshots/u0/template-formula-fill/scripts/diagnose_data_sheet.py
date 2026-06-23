#!/usr/bin/env python3
"""Diagnose Data sheet structure to verify header rows and data bounds.

Usage: python3 scripts/diagnose_data_sheet.py <workbook>

Prints the structure of the Data sheet to help identify:
- Which row contains year headers (usually row 4)
- Which rows contain entity codes
- Data range bounds
"""
import sys
import openpyxl


def diagnose_data_sheet(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    if 'Data' not in wb.sheetnames:
        print("ERROR: No 'Data' sheet found")
        return
    
    ws = wb['Data']
    print(f"Data sheet dimensions: {ws.dimensions}")
    print(f"Max row: {ws.max_row}, Max col: {ws.max_column}")
    print()
    
    # Check candidate header rows
    print("=== Candidate Header Rows ===")
    for row_idx in range(1, min(10, ws.max_row + 1)):
        row_values = [c.value for c in ws[row_idx]]
        # Check if row looks like years (contains 2020, 2021, etc.)
        has_years = any(str(v).startswith('20') and str(v).isdigit() for v in row_values if v)
        marker = " <-- LIKELY HEADERS" if has_years else ""
        print(f"Row {row_idx}: {row_values[:10]}{marker}")
    print()
    
    # Check column D for entity codes
    print("=== Column D (Entity Codes) ===")
    for row_idx in range(1, min(25, ws.max_row + 1)):
        val = ws.cell(row=row_idx, column=4).value
        if val:
            print(f"Row {row_idx}: {val}")
    print("...")
    for row_idx in range(max(1, ws.max_row - 5), ws.max_row + 1):
        val = ws.cell(row=row_idx, column=4).value
        if val:
            print(f"Row {row_idx}: {val}")
    print()
    
    # Count unique entities
    entities = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=4, max_col=4):
        val = row[0].value
        if val and val not in entities:
            entities.append(val)
    print(f"=== Entity Count ===")
    print(f"Total unique entities: {len(entities)}")
    print(f"Entities: {entities[:10]}..." if len(entities) > 10 else f"Entities: {entities}")
    
    # Check row 20 specifically (common error location)
    if ws.max_row >= 20:
        print()
        print("=== Row 20 Detail (check if mistaken for headers) ===")
        row_20 = [c.value for c in ws[20]]
        print(f"Row 20: {row_20[:10]}")
        print("Note: Row 20 is often part of data range, not headers (headers usually row 4)")


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print(f"Usage: {sys.argv[0]} <workbook.xlsx>")
        sys.exit(1)
    diagnose_data_sheet(sys.argv[1])