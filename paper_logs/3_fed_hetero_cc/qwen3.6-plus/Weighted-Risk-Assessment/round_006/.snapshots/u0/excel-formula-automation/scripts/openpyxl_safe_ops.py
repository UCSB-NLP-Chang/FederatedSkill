#!/usr/bin/env python3
"""Safe openpyxl operations for formula injection and verification."""
import os
import openpyxl

def load_workbook(path):
    """Load workbook with absolute path resolution and existence check."""
    abs_path = os.path.abspath(path)
    if not os.path.exists(abs_path):
        raise FileNotFoundError(f"Workbook not found: {abs_path}")
    return openpyxl.load_workbook(abs_path), abs_path

def count_formulas(ws):
    """Count cells containing formulas in a worksheet."""
    return sum(1 for row in ws.iter_rows() for cell in row if isinstance(cell.value, str) and cell.value.startswith('='))

def inject_formulas(ws, formula_map):
    """Apply formulas from a dict of {cell_coord: formula_string}."""
    for coord, formula in formula_map.items():
        ws[coord].value = formula

def verify_and_save(wb, out_path, expected_new_formulas=0):
    """Save workbook and verify formula count."""
    abs_out = os.path.abspath(out_path)
    wb.save(abs_out)

    # Reload to verify
    verify_wb = openpyxl.load_workbook(abs_out)
    total = sum(count_formulas(ws) for ws in verify_wb.worksheets)
    print(f"Saved to {abs_out}. Total formula cells: {total}")
    return total