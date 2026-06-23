#!/usr/bin/env python3
"""
Pre-flight validation for Data sheet integrity.
Exit code 1 = BLOCKING: data corruption or missing values detected.
Exit code 0 = PASS: data sheet appears valid for formula building.
"""

import sys
import openpyxl
from typing import List

def validate_data_sheet(path: str, sheet_name: str = "Data") -> List[str]:
    """Validate Data sheet contains expected structure and no corruption."""
    issues = []

    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as e:
        return [f"Cannot load workbook: {e}"]

    if sheet_name not in wb.sheetnames:
        return [f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}"]

    ws = wb[sheet_name]

    # Check for #REF! errors in first 50 rows of first 20 columns
    ref_errors = []
    for row in range(1, min(51, ws.max_row + 1)):
        for col in range(1, min(21, ws.max_column + 1)):
            cell = ws.cell(row=row, column=col)
            if cell.value == "#REF!" or (isinstance(cell.value, str) and "#REF!" in cell.value):
                ref_errors.append(f"{cell.coordinate}")

    if ref_errors:
        issues.append(f"#REF! errors detected at: {', '.join(ref_errors[:5])}")
        issues.append("  -> Source data is corrupted. Escalate to repair Data sheet first.")

    # Check for series code column (common column D) - verify it contains strings
    series_col = 4  # Column D
    sample_series = []
    for row in range(21, min(39, ws.max_row + 1)):
        val = ws.cell(row=row, column=series_col).value
        if val is not None:
            sample_series.append((row, val, type(val).__name__))

    if not sample_series:
        issues.append(f"No values found in column {series_col} (rows 21-38) - expected series codes")
    else:
        # Check for type consistency
        types = set(t for _, _, t in sample_series)
        if len(types) > 1:
            issues.append(f"Mixed types in series code column: {types}")
            issues.append(f"  Sample: {sample_series[:3]}")
        elif "str" not in types and "int" not in types:
            issues.append(f"Unexpected type in series codes: {types}")
            issues.append(f"  May cause MATCH failures. Expected string or int codes.")

    # Check for year headers in common header rows (row 4 or row 21)
    header_candidates = [4, 21, 1]
    found_headers = False
    for header_row in header_candidates:
        headers = []
        for col in range(8, min(13, ws.max_column + 1)):  # H through L
            val = ws.cell(row=header_row, column=col).value
            if val is not None:
                headers.append(str(val))

        # Look for year patterns (2018, 2019, etc.)
        year_like = [h for h in headers if h.isdigit() and len(h) == 4]
        if len(year_like) >= 3:
            found_headers = True
            break

    if not found_headers:
        issues.append("Could not find year headers in expected locations (rows 1, 4, 21, columns H-L)")
        issues.append("  -> Verify Data sheet structure before building formulas")

    # Check data range population (H21:L38 common pattern)
    empty_cells = 0
    total_cells = 0
    for row in range(21, min(39, ws.max_row + 1)):
        for col in range(8, min(13, ws.max_column + 1)):
            total_cells += 1
            if ws.cell(row=row, column=col).value is None:
                empty_cells += 1

    if total_cells > 0 and empty_cells / total_cells > 0.5:
        issues.append(f"Data range appears mostly empty ({empty_cells}/{total_cells} cells None)")
        issues.append("  -> Verify correct data range before building INDEX/MATCH")

    wb.close()
    return issues

def main():
    if len(sys.argv) < 2:
        print("Usage: python3 validate_data_sheet.py <workbook_path> [sheet_name]")
        sys.exit(1)

    path = sys.argv[1]
    sheet_name = sys.argv[2] if len(sys.argv) > 2 else "Data"

    print(f"=== VALIDATING DATA SHEET '{sheet_name}' in {path} ===")

    issues = validate_data_sheet(path, sheet_name)

    if issues:
        print("\n*** BLOCKING DATA ISSUES DETECTED ***")
        for issue in issues:
            print(f"  {issue}")
        print("\nFIX DATA SHEET ISSUES BEFORE BUILDING FORMULAS. Exiting with code 1.")
        sys.exit(1)
    else:
        print("\n=== DATA VALIDATION PASSED ===")
        print("Series codes found, no #REF! errors, year headers located.")
        sys.exit(0)

if __name__ == "__main__":
    main()