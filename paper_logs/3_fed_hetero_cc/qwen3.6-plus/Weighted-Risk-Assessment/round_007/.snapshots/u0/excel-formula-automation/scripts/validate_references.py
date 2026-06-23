#!/usr/bin/env python3
"""
Blocking validation for Excel formula reference locking and MATCH mode.
Exits with non-zero status if required $ signs missing or MATCH mode omitted.
"""
import sys
import re
import openpyxl

# Patterns that MUST have $ signs (row or column locked)
REQUIRED_LOCKING = {
    # Lookup ranges in INDEX/MATCH should be fully absolute
    "lookup_range": re.compile(r'INDEX\s*\(\s*([^,]+)', re.IGNORECASE),
    # Statistics ranges should have row-absolute
    "stats_range": re.compile(r'(?:MIN|MAX|MEDIAN|AVERAGE|PERCENTILE\.INC|QUARTILE\.INC)\s*\(\s*([^)]+)\)', re.IGNORECASE),
    # SUMPRODUCT ranges should have row-absolute
    "sumproduct_range": re.compile(r'SUMPRODUCT\s*\(\s*([^,]+)', re.IGNORECASE),
}

# Pattern to detect MATCH without mode argument
MATCH_PATTERN = re.compile(r'MATCH\s*\(\s*([^,]+)\s*,\s*([^,)]+)(?:\s*\)|\s*,)', re.IGNORECASE)

def check_range_locking(range_str, context):
    """Check if a range has appropriate $ locking."""
    issues = []
    range_str = range_str.strip().strip('"').strip("'")

    # Skip if not a range reference (e.g., a named range or single cell)
    if ':' not in range_str:
        return issues

    # Extract sheet name if present
    if '!' in range_str:
        sheet_part, range_part = range_str.split('!', 1)
        range_part = range_part.strip()
    else:
        range_part = range_str

    # Check for $ signs in range
    # For lookup ranges: both row and column should be locked
    if context == "lookup_range":
        # Should have $ before both row numbers and column letters
        if not re.search(r'\$[A-Z]+\$\d+.*:\$[A-Z]+\$\d+', range_part):
            issues.append(f"Lookup range '{range_str}' should be fully absolute (e.g., $A$1:$B$10)")

    # For stats/sumproduct: row numbers should be locked
    elif context in ("stats_range", "sumproduct_range"):
        # Extract the range part and check row locking
        # Should have $ before row numbers
        parts = range_part.split(':')
        for part in parts:
            part = part.strip()
            # Check if row number is locked (has $ before the number)
            if re.search(r'[A-Z]+(\d+)', part):
                if not re.search(r'[A-Z]+\$\d+', part):
                    issues.append(f"Range '{range_str}' should have row-absolute references (e.g., H$35:H$40)")
                    break

    return issues

def check_match_mode(formula: str) -> list:
    """Check if MATCH functions have explicit mode=0 argument."""
    issues = []

    # Find all MATCH calls
    for match in re.finditer(r'MATCH\s*\([^)]+\)', formula, re.IGNORECASE):
        match_call = match.group(0)
        # Check if it has 3 arguments (value, range, mode)
        args = match_call.split(',')
        if len(args) < 3:
            # Missing mode argument - defaults to 1 (sorted), which fails on unsorted
            issues.append(f"MATCH missing mode argument: '{match_call}' should have ',0)' for exact match")
        else:
            # Check if mode is 0
            mode_arg = args[2].strip().rstrip(')')
            if mode_arg != '0':
                issues.append(f"MATCH mode should be 0 (exact): '{match_call}' has mode '{mode_arg}'")

    return issues

def validate_workbook(path):
    """Validate all formulas in a workbook for proper reference locking and MATCH mode."""
    try:
        wb = openpyxl.load_workbook(path, data_only=False)
    except Exception as e:
        print(f"ERROR: Could not load workbook: {e}", file=sys.stderr)
        return False

    all_issues = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    formula = cell.value

                    # Check reference locking
                    for context, pattern in REQUIRED_LOCKING.items():
                        matches = pattern.findall(formula)
                        for match in matches:
                            issues = check_range_locking(match, context)
                            for issue in issues:
                                all_issues.append(f"[{sheet_name}!{cell.coordinate}] {issue}")

                    # Check MATCH mode
                    match_issues = check_match_mode(formula)
                    for issue in match_issues:
                        all_issues.append(f"[{sheet_name}!{cell.coordinate}] {issue}")

    if all_issues:
        print("VALIDATION FAILED:", file=sys.stderr)
        for issue in all_issues:
            print(f"  - {issue}", file=sys.stderr)
        return False

    print("VALIDATION PASSED - All reference locking and MATCH modes correct")
    return True

def inspect_workbook(path):
    """Print workbook structure for inspection."""
    try:
        wb = openpyxl.load_workbook(path, data_only=False)
    except Exception as e:
        print(f"ERROR: Could not load workbook: {e}", file=sys.stderr)
        return

    print(f"Workbook: {path}")
    print(f"Sheets: {wb.sheetnames}")
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n[{sheet_name}]")
        print(f"  Dimensions: {ws.dimensions}")
        # Print first few rows
        for row_idx, row in enumerate(ws.iter_rows(max_row=10), 1):
            row_vals = [str(cell.value)[:20] if cell.value else "" for cell in row[:5]]
            print(f"  Row {row_idx}: {row_vals}")

if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(description="Validate Excel formula reference locking and MATCH mode")
    parser.add_argument("path", nargs="?", help="Path to Excel workbook")
    parser.add_argument("--inspect", action="store_true", help="Inspect workbook structure")
    args = parser.parse_args()

    if args.inspect and args.path:
        inspect_workbook(args.path)
    elif args.path:
        if not validate_workbook(args.path):
            sys.exit(1)
    else:
        parser.print_help()