#!/usr/bin/env python3
"""Skeleton script for building rollforward workbooks.

MANDATORY: Adapt this script for each task. Do NOT build from scratch.

Usage:
    python scripts/build_rollforward.py --base <csv> --gl <json> --output <xlsx>
"""

import argparse
import csv
import json
from openpyxl import Workbook
from openpyxl.styles import Font


def load_csv(path):
    with open(path) as f:
        return list(csv.DictReader(f))


def load_json(path):
    with open(path) as f:
        return json.load(f)


def build_detail_sheet(ws, line_items, gl_balances, account_name):
    """Build a detail sheet with headers, data, and control rows.

    Control row order: Period Totals → Ending Balance → Variance → GL Balance
    """
    # Headers (row 5)
    headers = ['Partner', 'Beginning Balance', 'Jul Accruals', 'Jul Utilization', 'Jul Ending',
               'Aug Accruals', 'Aug Utilization', 'Aug Ending',
               'Sep Accruals', 'Sep Utilization', 'Sep Ending',
               'Oct Accruals', 'Oct Utilization', 'Oct Ending', 'Reserve Months']
    for col, header in enumerate(headers, 1):
        ws.cell(row=5, column=col, value=header)
        ws.cell(row=5, column=col).font = Font(bold=True)

    # Data rows (row 6+)
    first_data_row = 6
    for i, item in enumerate(line_items):
        row = first_data_row + i
        ws.cell(row=row, column=1, value=item.get('partner', item.get('entity', '')))
        ws.cell(row=row, column=2, value=float(item.get('beginning_balance', 0)))
        ws.cell(row=row, column=3, value=float(item.get('jul_accruals', 0)))
        ws.cell(row=row, column=4, value=float(item.get('jul_utilization', 0)))
        # Jul Ending formula
        ws.cell(row=row, column=5, value=f'=B{row}+C{row}-D{row}')
        ws.cell(row=row, column=6, value=float(item.get('aug_accruals', 0)))
        ws.cell(row=row, column=7, value=float(item.get('aug_utilization', 0)))
        # Aug Ending formula
        ws.cell(row=row, column=8, value=f'=E{row}+F{row}-G{row}')
        ws.cell(row=row, column=9, value=float(item.get('sep_accruals', 0)))
        ws.cell(row=row, column=10, value=float(item.get('sep_utilization', 0)))
        # Sep Ending formula
        ws.cell(row=row, column=11, value=f'=H{row}+I{row}-J{row}')
        ws.cell(row=row, column=12, value=float(item.get('oct_accruals', 0)))
        ws.cell(row=row, column=13, value=float(item.get('oct_utilization', 0)))
        # Oct Ending formula
        ws.cell(row=row, column=14, value=f'=K{row}+L{row}-M{row}')
        # Reserve formula
        ws.cell(row=row, column=15, value=f'=C{row}+F{row}+I{row}+L{row}')

    last_data_row = first_data_row + len(line_items) - 1

    # Control rows
    totals_row = last_data_row + 1
    ending_row = last_data_row + 2
    variance_row = last_data_row + 3
    gl_row = last_data_row + 4

    # Period Totals
    ws.cell(row=totals_row, column=1, value='Period Totals')
    for col in range(2, 15):  # B through N
        col_letter = chr(64 + col)
        ws.cell(row=totals_row, column=col, value=f'=SUM({col_letter}{first_data_row}:{col_letter}{last_data_row})')
    ws.cell(row=totals_row, column=15, value=f'=C{totals_row}+F{totals_row}+I{totals_row}+L{totals_row}')

    # Ending Balance (references Period Totals for Beg, prior Ending for subsequent)
    ws.cell(row=ending_row, column=1, value='Ending Balance')
    # Jul Ending: Beg from Period Totals
    ws.cell(row=ending_row, column=5, value=f'=B{totals_row}+C{totals_row}-D{totals_row}')
    # Aug Ending: prior Ending + activity
    ws.cell(row=ending_row, column=8, value=f'=E{ending_row}+F{totals_row}-G{totals_row}')
    # Sep Ending
    ws.cell(row=ending_row, column=11, value=f'=H{ending_row}+I{totals_row}-J{totals_row}')
    # Oct Ending
    ws.cell(row=ending_row, column=14, value=f'=K{ending_row}+L{totals_row}-M{totals_row}')
    # Reserve
    ws.cell(row=ending_row, column=15, value=f'=D{ending_row}+G{ending_row}+J{ending_row}+M{ending_row}')

    # Variance (CRITICAL: both operands use column N, NOT O)
    ws.cell(row=variance_row, column=1, value='Variance')
    ws.cell(row=variance_row, column=15, value=f'=N{gl_row}-N{ending_row}')  # N - N, NOT O - N

    # GL Balance (static values from JSON)
    ws.cell(row=gl_row, column=1, value='GL Balance')
    ws.cell(row=gl_row, column=5, value=gl_balances.get('jul', 0))
    ws.cell(row=gl_row, column=8, value=gl_balances.get('aug', 0))
    ws.cell(row=gl_row, column=11, value=gl_balances.get('sep', 0))
    ws.cell(row=gl_row, column=14, value=gl_balances.get('oct', 0))
    ws.cell(row=gl_row, column=15, value=f'=O{totals_row}-O{ending_row}')

    return totals_row, ending_row, gl_row


def build_summary_sheet(ws, detail_sheets):
    """Build summary sheet with cross-sheet links."""
    ws.cell(row=1, column=1, value='Summary')

    row = 3
    for sheet_name, totals_row, ending_row, gl_row in detail_sheets:
        ws.cell(row=row, column=1, value=sheet_name)
        # Use single quotes for sheet names with spaces
        ws.cell(row=row, column=2, value=f"='{sheet_name}'!O{totals_row}")
        ws.cell(row=row, column=3, value=f"='{sheet_name}'!O{ending_row}")
        ws.cell(row=row, column=4, value=f"='{sheet_name}'!N{gl_row}")
        row += 1


def main():
    parser = argparse.ArgumentParser(description='Build rollforward workbook')
    parser.add_argument('--base', required=True, help='Base CSV file')
    parser.add_argument('--gl', required=True, help='GL balances JSON')
    parser.add_argument('--output', required=True, help='Output XLSX file')
    parser.add_argument('--account-mapping', help='Account mapping JSON')
    args = parser.parse_args()

    # Load data
    base_data = load_csv(args.base)
    gl_data = load_json(args.gl)
    account_mapping = load_json(args.account_mapping) if args.account_mapping else {}

    # Build workbook
    wb = Workbook()
    wb.remove(wb.active)

    # Create Summary first
    summary_ws = wb.create_sheet('Summary', 0)

    # Build detail sheets
    detail_sheets = []
    for item in base_data:
        account = item.get('account_number', item.get('account', 'Unknown'))
        sheet_name = account_mapping.get(account, account)

        ws = wb.create_sheet(sheet_name)
        gl_balances = gl_data.get(account, {})
        totals_row, ending_row, gl_row = build_detail_sheet(ws, [item], gl_balances, sheet_name)
        detail_sheets.append((sheet_name, totals_row, ending_row, gl_row))

    # Build summary
    build_summary_sheet(summary_ws, detail_sheets)

    # Save
    wb.save(args.output)
    print(f"Saved: {args.output}")


if __name__ == '__main__':
    main()
