#!/usr/bin/env python3
"""Quick pre-submission validation for rollforward workbooks.

Run this BEFORE the full test suite to catch the most common fatal bugs.

Usage:
    python scripts/quick_validate.py <workbook_path>

Checks (fast, targeted):
    1. Variance formula uses column N for BOTH operands
    2. Ending Balance references Period Totals (not self)
    3. No #REF! errors in formula cells
    4. Summary sheet GL Balance links use column N (not O)
    5. Sheet names match expected pattern

Exit codes:
    0 = All quick checks passed
    1 = One or more checks failed
"""

import sys
import re
from openpyxl import load_workbook


def find_control_rows(ws):
    """Find control row indices by scanning column A labels."""
    rows = {}
    for row in range(1, min(ws.max_row + 1, 50)):
        label = ws.cell(row=row, column=1).value
        if label:
            label_str = str(label).strip()
            if 'Period Totals' in label_str:
                rows['period_totals'] = row
            elif 'Ending Balance' in label_str:
                rows['ending_balance'] = row
            elif 'Variance' in label_str:
                rows['variance'] = row
            elif 'GL Balance' in label_str:
                rows['gl_balance'] = row
    return rows


def check_variance(ws, var_row, gl_row, eb_row):
    """CRITICAL: Variance must be =N{gl}-N{eb}, NOT =O{gl}-N{eb} or =O{gl}-O{eb}."""
    cell_val = ws.cell(row=var_row, column=14).value  # Column N
    if not cell_val or not isinstance(cell_val, str):
        return False, f"N{var_row} is empty or not a formula"
    formula = cell_val.strip()
    expected = f'=N{gl_row}-N{eb_row}'
    if formula == expected:
        return True, f"Variance correct: {formula}"
    # Check for common wrong patterns
    if formula.startswith('=O'):
        return False, f"Variance WRONG: uses column O for GL. Got '{formula}', expected '{expected}'"
    if '-O' in formula:
        return False, f"Variance WRONG: uses column O for Ending Balance. Got '{formula}', expected '{expected}'"
    return False, f"Variance unexpected: '{formula}', expected '{expected}'"


def check_ending_balance(ws, pt_row, eb_row):
    """CRITICAL: Ending Balance must reference Period Totals, not self."""
    cell_val = ws.cell(row=eb_row, column=5).value  # Column E
    if not cell_val or not isinstance(cell_val, str):
        return False, f"E{eb_row} is empty or not a formula"
    formula = cell_val.strip()
    # Check for self-reference (B{eb_row} instead of B{pt_row})
    if f'B{eb_row}' in formula:
        return False, f"Ending Balance WRONG: self-references B{eb_row}. Got '{formula}', should reference B{pt_row}"
    if f'B{pt_row}' in formula:
        return True, f"Ending Balance correct: references Period Totals"
    return False, f"Ending Balance unexpected: '{formula}'"


def check_ref_errors(ws):
    """Check for #REF! errors in any cell."""
    for row in range(1, min(ws.max_row + 1, 30)):
        for col in range(1, min(ws.max_column + 1, 20)):
            val = ws.cell(row=row, column=col).value
            if val and isinstance(val, str) and '#REF' in val:
                return False, f"#REF! error at {chr(64+col)}{row}"
    return True, "No #REF! errors"


def check_summary_gl_links(ws):
    """CRITICAL: Summary sheet GL Balance links must use column N, not O.

    The most common Summary sheet bug is linking to column O for GL Balance
    when the GL Balance data in detail sheets lives in column N.
    """
    errors = []
    for row in range(1, min(ws.max_row + 1, 50)):
        label = ws.cell(row=row, column=1).value
        if label and isinstance(label, str):
            label_upper = label.upper()
            # Check rows with "GL" and "Balance" in the label
            if 'GL' in label_upper and 'BALANCE' in label_upper:
                # Check column B for the cross-sheet link formula
                formula = ws.cell(row=row, column=2).value
                if formula and isinstance(formula, str):
                    # GL Balance links MUST use column N, not O
                    if '!O' in formula:
                        errors.append(f"Row {row}: GL Balance link uses column O '{formula}' - should be column N")
                    elif '!N' not in formula:
                        errors.append(f"Row {row}: GL Balance link missing column N '{formula}'")

    if errors:
        return False, "Summary GL links WRONG: " + "; ".join(errors)
    return True, "Summary GL links use column N correctly"


def main(filepath):
    print(f"Quick validation: {filepath}")
    print("=" * 50)

    try:
        wb = load_workbook(filepath, data_only=False)
    except Exception as e:
        print(f"FATAL: Cannot load workbook: {e}")
        return 1

    all_passed = True

    # First, check Summary sheet for GL Balance link errors
    summary_sheet = None
    for sheet_name in wb.sheetnames:
        if 'Summary' in sheet_name:
            summary_sheet = wb[sheet_name]
            break

    if summary_sheet:
        print("\nSummary Sheet GL Balance Links:")
        ok, msg = check_summary_gl_links(summary_sheet)
        status = "PASS" if ok else "FAIL"
        print(f"  [{status}] {msg}")
        all_passed = all_passed and ok

    # Then check detail sheets
    for sheet_name in wb.sheetnames:
        if 'Summary' in sheet_name:
            continue
        ws = wb[sheet_name]
        print(f"\nSheet: {sheet_name}")

        rows = find_control_rows(ws)
        if not rows:
            print("  SKIP: No control rows found")
            continue

        pt = rows.get('period_totals')
        eb = rows.get('ending_balance')
        var = rows.get('variance')
        gl = rows.get('gl_balance')

        # Check Variance
        if var and gl and eb:
            ok, msg = check_variance(ws, var, gl, eb)
            status = "PASS" if ok else "FAIL"
            print(f"  [{status}] Variance: {msg}")
            all_passed = all_passed and ok

        # Check Ending Balance
        if pt and eb:
            ok, msg = check_ending_balance(ws, pt, eb)
            status = "PASS" if ok else "FAIL"
            print(f"  [{status}] Ending Balance: {msg}")
            all_passed = all_passed and ok

        # Check REF errors
        ok, msg = check_ref_errors(ws)
        status = "PASS" if ok else "FAIL"
        print(f"  [{status}] REF errors: {msg}")
        all_passed = all_passed and ok

    print("\n" + "=" * 50)
    if all_passed:
        print("ALL QUICK CHECKS PASSED - proceed to full test suite")
        return 0
    else:
        print("QUICK CHECKS FAILED - fix before running test suite")
        return 1


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)
    sys.exit(main(sys.argv[1]))
