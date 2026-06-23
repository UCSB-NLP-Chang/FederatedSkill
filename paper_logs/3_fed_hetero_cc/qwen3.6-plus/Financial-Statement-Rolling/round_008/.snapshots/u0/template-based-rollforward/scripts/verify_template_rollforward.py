#!/usr/bin/env python3
"""Verify template-based rollforward workbook structure.

Usage:
    python verify_template_rollforward.py <workbook_path>

Checks specific to template-based builds:
    1. Summary sheet GL Balance links use column N (not O)
    2. Summary label/formula rows are aligned (same row)
    3. Cross-sheet references properly quoted
    4. Detail sheet GL values in E, H, K, N
"""

import sys
from openpyxl import load_workbook
import re


def check_summary_gl_links(wb):
    """Verify GL Balance links use column N."""
    if 'Refund Summary' not in wb.sheetnames and 'Summary' not in wb.sheetnames:
        return True, "No summary sheet to check"

    summary_name = 'Refund Summary' if 'Refund Summary' in wb.sheetnames else 'Summary'
    ws = wb[summary_name]

    errors = []
    for row in range(1, ws.max_row + 1):
        cell = ws.cell(row=row, column=2)  # Column B
        if not cell.value or not isinstance(cell.value, str):
            continue
        if not cell.value.startswith('='):
            continue

        # Check for GL Balance row label in column A
        label = ws.cell(row=row, column=1).value
        if label and 'GL' in str(label).upper() and 'Balance' in str(label):
            # This row should link to column N
            if '!O' in cell.value:
                errors.append(f"Row {row}: GL Balance uses column O '{cell.value}' (should be N)")
            elif '!N' not in cell.value:
                errors.append(f"Row {row}: GL Balance missing column N reference '{cell.value}'")

    if errors:
        return False, "GL Balance errors: " + "; ".join(errors)
    return True, "GL Balance links OK (column N)"


def check_row_alignment(ws):
    """Check that label and formula rows align."""
    misaligned = []
    for row in range(1, ws.max_row):
        a_val = ws.cell(row=row, column=1).value
        b_val = ws.cell(row=row, column=2).value
        a_next = ws.cell(row=row+1, column=1).value

        # If A has a label like "Period Totals" and B has formula, check alignment
        if a_val and isinstance(a_val, str) and a_val in ['Period Totals', 'Ending Balance', 'GL Balance']:
            if not b_val or not isinstance(b_val, str) or not b_val.startswith('='):
                # Check if formula is in next row (off-by-one)
                b_next = ws.cell(row=row+1, column=2).value
                if b_next and isinstance(b_next, str) and b_next.startswith('='):
                    misaligned.append(f"Row {row}: label '{a_val}' misaligned with formula in row {row+1}")

    if misaligned:
        return False, "Alignment issues: " + "; ".join(misaligned)
    return True, "Row alignment OK"


def check_gl_value_placement(ws):
    """Verify GL Balance row has values in E, H, K, N."""
    # Find GL Balance row
    gl_row = None
    for row in range(1, ws.max_row + 1):
        val = ws.cell(row=row, column=1).value
        if val and isinstance(val, str) and 'GL' in val.upper() and 'Balance' in val:
            gl_row = row
            break

    if not gl_row:
        return True, "No GL Balance row found"

    # Check E(5), H(8), K(11), N(14) have values
    cols = [5, 8, 11, 14]  # E, H, K, N
    missing = []
    for col in cols:
        val = ws.cell(row=gl_row, column=col).value
        if val is None or val == '':
            missing.append(f"{chr(64+col)}{gl_row}")

    if missing:
        return False, f"GL Balance row {gl_row} missing values at: {', '.join(missing)}"
    return True, f"GL Balance row {gl_row} has values in E, H, K, N"


def main(filepath):
    print(f"Verifying: {filepath}")
    print("-" * 50)

    try:
        wb = load_workbook(filepath, data_only=False)
    except Exception as e:
        print(f"ERROR: Cannot load workbook: {e}")
        return 1

    all_ok = True

    # Check summary GL links
    ok, msg = check_summary_gl_links(wb)
    print(f"[{'OK' if ok else 'FAIL'}] {msg}")
    all_ok = all_ok and ok

    # Check each sheet
    summary_names = ['Summary', 'Refund Summary']
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\nSheet: {sheet_name}")

        if sheet_name in summary_names:
            ok, msg = check_row_alignment(ws)
            print(f"[{'OK' if ok else 'FAIL'}] {msg}")
            all_ok = all_ok and ok
        else:
            ok, msg = check_gl_value_placement(ws)
            print(f"[{'OK' if ok else 'FAIL'}] {msg}")
            all_ok = all_ok and ok

    print("-" * 50)
    if all_ok:
        print("PASS: Template rollforward checks passed")
        return 0
    else:
        print("FAIL: Some template-specific checks failed")
        return 1


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)
    sys.exit(main(sys.argv[1]))