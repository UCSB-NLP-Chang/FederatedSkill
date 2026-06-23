#!/usr/bin/env python3
"""Validate rollforward workbook formulas - MUST pass before saving.

CRITICAL CHECK: Variance formula must use column N for both operands.
Wrong formula (=O{gl_row}-N{ending_row}) causes verification failure.

Usage:
    python scripts/validate_formulas.py <workbook_path>
"""

import sys
import re
from openpyxl import load_workbook


def validate_variance_formula(ws, variance_row=11, gl_row=12, ending_row=10):
    """Assert Variance uses column N for BOTH operands (not O for GL).

    CORRECT:   =N{gl_row}-N{ending_row}
    WRONG:     =O{gl_row}-N{ending_row}
    """
    # Find variance row by label
    variance_cell = None
    gl_cell = None
    for row in range(6, 20):
        label = ws[f'A{row}'].value
        if label == 'Variance':
            variance_row = row
        elif label == 'GL Balance':
            gl_row = row
        elif label == 'Ending Balance':
            ending_row = row

    # Check column O (Reserve column) for variance formula
    formula = ws[f'O{variance_row}'].value
    if formula and isinstance(formula, str) and formula.startswith('='):
        # Check for wrong pattern: O{gl_row} appears
        if re.search(rf'O{gl_row}', formula):
            raise AssertionError(
                f"Variance formula uses column O (wrong): {formula}\n"
                f"GL Balance is in column N, not O.\n"
                f"CORRECT formula: =N{gl_row}-N{ending_row}"
            )
        # Check for correct pattern: N{gl_row}-N{ending_row}
        if not re.search(rf'=N{gl_row}-N{ending_row}', formula):
            print(f"WARNING: Variance formula pattern unexpected: {formula}")
            print(f"Expected: =N{gl_row}-N{ending_row}")
        else:
            print(f"OK: Variance formula correct: {formula}")
    return True


def validate_ending_balance_reference(ws, ending_row=10, totals_row=9):
    """Assert Ending Balance Beginning cell references Period Totals, not self."""
    # Find rows by label
    for row in range(6, 20):
        label = ws[f'A{row}'].value
        if label == 'Period Totals':
            totals_row = row
        elif label == 'Ending Balance':
            ending_row = row

    # Check E column (Jul Ending) - should reference B{totals_row}
    formula = ws[f'E{ending_row}'].value
    if formula and isinstance(formula, str) and formula.startswith('='):
        # Check for self-reference: B{ending_row} instead of B{totals_row}
        if re.search(rf'=B{ending_row}\+', formula):
            raise AssertionError(
                f"Ending Balance self-references: {formula}\n"
                f"Beginning Balance must reference Period Totals row B{totals_row}, "
                f"not own empty cell B{ending_row}"
            )
        print(f"OK: Ending Balance references correct: {formula}")
    return True


def validate_sheet_order(wb, expected_first='Summary'):
    """Summary sheet must be first."""
    sheets = wb.sheetnames
    if sheets[0] != expected_first:
        raise AssertionError(f"Sheet order wrong: {sheets[0]} first, expected {expected_first}")
    print(f"OK: Sheet order: {sheets}")
    return True


def validate_formulas_present(ws):
    """Check that control rows have formulas, not hardcoded values."""
    control_labels = ['Period Totals', 'Ending Balance', 'Variance']
    found = []
    for row in range(6, 20):
        label = ws[f'A{row}'].value
        if label in control_labels:
            # Check column E has formula
            val = ws[f'E{row}'].value
            if val and isinstance(val, str) and val.startswith('='):
                found.append(label)
    if len(found) < 2:
        raise AssertionError(f"Missing formula rows. Found: {found}")
    print(f"OK: Control rows with formulas: {found}")
    return True


def main(filepath):
    print(f"Validating: {filepath}")
    print("-" * 50)

    wb = load_workbook(filepath, data_only=False)

    # Validate sheet order
    validate_sheet_order(wb)

    # Validate each detail sheet
    for sheet_name in wb.sheetnames:
        if 'Summary' not in sheet_name:
            ws = wb[sheet_name]
            print(f"\nSheet: {sheet_name}")
            try:
                validate_formulas_present(ws)
                validate_ending_balance_reference(ws)
                validate_variance_formula(ws)
            except AssertionError as e:
                print(f"FAIL: {e}")
                return False

    print("\n" + "-" * 50)
    print("PASS: All validations passed")
    return True


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)
    success = main(sys.argv[1])
    sys.exit(0 if success else 1)
