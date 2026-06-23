#!/usr/bin/env python3
"""Verify openpyxl workbook structure against requirements."""
import sys
import openpyxl

def verify(path, expected_sheets=None, expected_named_ranges=None, check_labels=None):
    wb = openpyxl.load_workbook(path)
    errors = []
    
    if expected_sheets and wb.sheetnames != expected_sheets:
        errors.append(f"Sheet order mismatch: {wb.sheetnames}")
        
    if expected_named_ranges is not None:
        actual = len(wb.defined_names.definedName)
        if actual != expected_named_ranges:
            errors.append(f"Named ranges: {actual} != {expected_named_ranges}")
            
    if check_labels:
        for sheet_name, labels in check_labels.items():
            ws = wb.get(sheet_name)
            if not ws:
                errors.append(f"Missing sheet: {sheet_name}")
                continue
            found = {str(row[0]).strip() for row in ws.iter_rows(min_col=1, max_col=1, values_only=True) if row[0]}
            for lbl in labels:
                if lbl not in found:
                    errors.append(f"Missing label '{lbl}' in {sheet_name}")
                    
    if errors:
        print("VERIFICATION FAILED:")
        for e in errors: print(f"  - {e}")
        return False
    print("VERIFICATION PASSED")
    return True

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: verify_workbook.py <path_to_xlsx>")
        sys.exit(1)
    verify(sys.argv[1])
