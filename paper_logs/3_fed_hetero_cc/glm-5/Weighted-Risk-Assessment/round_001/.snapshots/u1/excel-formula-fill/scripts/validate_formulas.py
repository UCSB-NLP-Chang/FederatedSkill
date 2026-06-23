#!/usr/bin/env python3
"""Validate that formulas in an Excel workbook are correctly written.

Usage: python3 validate_formulas.py <input_workbook> <output_workbook>

Reads the Data sheet from the input workbook and checks that:
1. All highlighted cells in Task sheet contain formulas
2. Number formats are preserved
3. Formulas are present and categorized by type

Does NOT evaluate formulas. Use the manual verification step (step 5 in SKILL.md)
to check computed correctness.
"""
import sys
import openpyxl


def read_data_sheet(wb):
    """Extract all values from the Data sheet into a dict keyed by (row, col)."""
    if 'Data' not in wb.sheetnames:
        print("ERROR: No 'Data' sheet found")
        return None
    ws = wb['Data']
    data = {}
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            if cell.value is not None:
                data[(cell.row, cell.column)] = cell.value
    return data


def read_task_formulas(wb):
    """Extract all formulas from the Task sheet."""
    if 'Task' not in wb.sheetnames:
        print("ERROR: No 'Task' sheet found")
        return []
    ws = wb['Task']
    formulas = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith('='):
                formulas.append({
                    'cell': cell.coordinate,
                    'formula': cell.value,
                    'row': cell.row,
                    'col': cell.column,
                })
    return formulas


def main():
    if len(sys.argv) < 3:
        print(f"Usage: {sys.argv[0]} <input.xlsx> <output.xlsx>")
        sys.exit(1)

    input_path = sys.argv[1]
    output_path = sys.argv[2]

    wb_in = openpyxl.load_workbook(input_path, data_only=True)
    wb_out = openpyxl.load_workbook(output_path)

    data = read_data_sheet(wb_in)
    formulas = read_task_formulas(wb_out)

    print(f"Data sheet entries: {len(data)}")
    print(f"Formulas written: {len(formulas)}")

    # Print sample Data values for manual verification
    print("\n=== Sample Data sheet values (for manual verification) ===")
    ws_data = wb_in['Data']
    sample_count = 0
    for row in range(1, min(15, ws_data.max_row + 1)):
        for col in range(1, min(10, ws_data.max_column + 1)):
            val = ws_data.cell(row=row, column=col).value
            if val is not None:
                print(f"  [{row},{col}]: {val}")
                sample_count += 1
                if sample_count >= 10:
                    break
        if sample_count >= 10:
            break

    # Group formulas by type
    lookup_count = 0
    derived_count = 0
    stats_count = 0
    weighted_count = 0

    for f in formulas:
        formula = f['formula']
        if 'INDEX' in formula:
            lookup_count += 1
        elif 'SUMPRODUCT' in formula:
            weighted_count += 1
        elif any(fn in formula for fn in ['MIN', 'MAX', 'MEDIAN', 'AVERAGE', 'PERCENTILE']):
            stats_count += 1
        else:
            derived_count += 1

    print(f"\nFormula breakdown:")
    print(f"  Lookup (INDEX+MATCH): {lookup_count}")
    print(f"  Derived calculations: {derived_count}")
    print(f"  Statistics: {stats_count}")
    print(f"  Weighted mean: {weighted_count}")

    # Print formulas for inspection
    print("\n=== Formulas written (sample) ===")
    for f in formulas[:10]:
        print(f"  {f['cell']}: {f['formula'][:60]}{'...' if len(f['formula']) > 60 else ''}")
    if len(formulas) > 10:
        print(f"  ... and {len(formulas) - 10} more")

    # Check that all highlighted cells have formulas
    ws_task_in = wb_in['Task']
    ws_task_out = wb_out['Task']
    missing = 0
    for row in ws_task_in.iter_rows(min_row=1, max_row=ws_task_in.max_row, max_col=ws_task_in.max_column):
        for cell in row:
            if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb and 'FFF2CC' in str(cell.fill.fgColor.rgb):
                out_cell = ws_task_out.cell(row=cell.row, column=cell.column)
                if out_cell.value is None or (isinstance(out_cell.value, str) and not out_cell.value.startswith('=')):
                    missing += 1
                    print(f"  WARNING: Highlighted cell {cell.coordinate} has no formula (value={out_cell.value})")

    if missing == 0:
        print("\n[PASS] All highlighted cells contain formulas.")
    else:
        print(f"\n[FAIL] {missing} highlighted cells are missing formulas!")

    # Verify formatting preserved
    format_issues = 0
    for row in ws_task_in.iter_rows(min_row=1, max_row=ws_task_in.max_row, max_col=ws_task_in.max_column):
        for cell in row:
            if cell.fill and cell.fill.fgColor and 'FFF2CC' in str(cell.fill.fgColor.rgb):
                out_cell = ws_task_out.cell(row=cell.row, column=cell.column)
                if out_cell.number_format != cell.number_format:
                    format_issues += 1
                    print(f"  FORMAT MISMATCH {cell.coordinate}: expected '{cell.number_format}', got '{out_cell.number_format}'")

    if format_issues == 0:
        print("[PASS] All number formats preserved.")
    else:
        print(f"[FAIL] {format_issues} number format mismatches found!")

    # Reminder about manual verification
    print("\n=== NEXT STEP ===")
    print("Run manual verification (SKILL.md step 5):")
    print("  1. Use sample Data values above to compute 2-3 expected results")
    print("  2. Compare against what your INDEX/MATCH formulas should return")


if __name__ == '__main__':
    main()
