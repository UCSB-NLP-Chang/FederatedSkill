#!/usr/bin/env python3
"""Validate that formulas in an Excel workbook are present and formatting preserved.

Usage: python3 validate_formulas.py <input_workbook> <output_workbook>

Reads the Data sheet from the input workbook, then checks that formulas
were written in the output workbook and formatting preserved.
Also checks for common reference locking issues in SUMPRODUCT formulas.
Does NOT verify computed correctness — agent must spot-check manually.
"""
import sys
import re
import openpyxl


def read_data_sheet(wb):
    """Extract all values from the Data sheet into a dict keyed by (row, col)."""
    if 'Data' not in wb.sheetnames:
        print("ERROR: No 'Data' sheet found")
        return None
    ws = wb['Data']
    data = {}
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            if cell.value is not None:
                data[(cell.row, cell.column)] = cell.value
    return data


def read_task_formulas(wb):
    """Extract all formulas from the Task sheet."""
    if 'Task' not in wb.sheetnames:
        print("ERROR: No 'Task' sheet found")
        return []
    ws = wb['Task']
    formulas = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith('='):
                formulas.append({
                    'cell': cell.coordinate,
                    'formula': cell.value,
                    'row': cell.row,
                    'col': cell.column,
                })
    return formulas


def check_reference_locking(formulas):
    """Check for common reference locking issues in formulas."""
    issues = []
    for f in formulas:
        formula = f['formula']
        # Check SUMPRODUCT for row-absolute patterns (should have $ before row numbers)
        if 'SUMPRODUCT' in formula:
            # Find range references in SUMPRODUCT args
            ranges = re.findall(r'[A-Z]+\d+:[A-Z]+\d+', formula)
            for r in ranges:
                if '$' not in r:
                    issues.append(f"  WARNING: {f['cell']} SUMPRODUCT range '{r}' missing absolute reference ($)")
        # Check statistics formulas for row-absolute in multi-row ranges
        if any(fn in formula for fn in ['MIN', 'MAX', 'MEDIAN', 'AVERAGE', 'PERCENTILE']):
            ranges = re.findall(r'[A-Z]+\d+:[A-Z]+\d+', formula)
            for r in ranges:
                # Check if it's a multi-row range (more than 1 row)
                match = re.match(r'([A-Z]+)(\d+):([A-Z]+)(\d+)', r)
                if match:
                    start_row = int(match.group(2))
                    end_row = int(match.group(4))
                    if end_row - start_row > 0 and '$' not in r:
                        issues.append(f"  WARNING: {f['cell']} statistics range '{r}' may need row-absolute ($) for multi-row data")
    return issues


def main():
    if len(sys.argv) < 3:
        print(f"Usage: {sys.argv[0]} <input.xlsx> <output.xlsx>")
        sys.exit(1)

    input_path = sys.argv[1]
    output_path = sys.argv[2]

    wb_in = openpyxl.load_workbook(input_path, data_only=True)
    wb_out = openpyxl.load_workbook(output_path)

    data = read_data_sheet(wb_in)
    formulas = read_task_formulas(wb_out)

    print(f"Data sheet entries: {len(data)}")
    print(f"Formulas written: {len(formulas)}")

    # Group formulas by type based on pattern
    lookup_count = 0
    derived_count = 0
    stats_count = 0
    weighted_count = 0

    for f in formulas:
        formula = f['formula']
        if 'INDEX' in formula:
            lookup_count += 1
        elif 'SUMPRODUCT' in formula:
            weighted_count += 1
        elif any(fn in formula for fn in ['MIN', 'MAX', 'MEDIAN', 'AVERAGE', 'PERCENTILE']):
            stats_count += 1
        else:
            derived_count += 1

    print(f"\nFormula breakdown:")
    print(f"  Lookup (INDEX+MATCH): {lookup_count}")
    print(f"  Derived calculations: {derived_count}")
    print(f"  Statistics: {stats_count}")
    print(f"  Weighted mean: {weighted_count}")

    # Check that all highlighted cells have formulas
    ws_task_in = wb_in['Task']
    ws_task_out = wb_out['Task']
    missing = 0
    for row in ws_task_in.iter_rows(min_row=1, max_row=ws_task_in.max_row, max_col=ws_task_in.max_column):
        for cell in row:
            if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb and 'FFF2CC' in str(cell.fill.fgColor.rgb):
                out_cell = ws_task_out.cell(row=cell.row, column=cell.column)
                if out_cell.value is None or (isinstance(out_cell.value, str) and not out_cell.value.startswith('=')):
                    missing += 1
                    print(f"  WARNING: Highlighted cell {cell.coordinate} has no formula (value={out_cell.value})")

    if missing == 0:
        print("\n[PASS] All highlighted cells contain formulas.")
    else:
        print(f"\n[FAIL] {missing} highlighted cells are missing formulas!")

    # Verify formatting preserved
    format_issues = 0
    for row in ws_task_in.iter_rows(min_row=1, max_row=ws_task_in.max_row, max_col=ws_task_in.max_column):
        for cell in row:
            if cell.fill and cell.fill.fgColor and 'FFF2CC' in str(cell.fill.fgColor.rgb):
                out_cell = ws_task_out.cell(row=cell.row, column=cell.column)
                if out_cell.number_format != cell.number_format:
                    format_issues += 1
                    print(f"  FORMAT MISMATCH {cell.coordinate}: expected '{cell.number_format}', got '{out_cell.number_format}'")

    if format_issues == 0:
        print("[PASS] All number formats preserved.")
    else:
        print(f"[FAIL] {format_issues} number format mismatches found!")

    # Check reference locking
    locking_issues = check_reference_locking(formulas)
    if locking_issues:
        print("\nReference locking warnings:")
        for issue in locking_issues:
            print(issue)
    else:
        print("\n[PASS] No obvious reference locking issues detected.")

    print("\n=== NEXT STEP ===")
    print("Validation complete. Manual spot-check of calculated values is still required:")
    print("  1. Load Data sheet values manually")
    print("  2. Compute expected values for 2-3 cells per block")
    print("  3. Compare against what your formulas should return")


if __name__ == '__main__':
    main()
