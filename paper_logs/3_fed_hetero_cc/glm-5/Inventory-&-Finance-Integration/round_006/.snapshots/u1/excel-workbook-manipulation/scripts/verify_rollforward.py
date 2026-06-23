#!/usr/bin/env python3
"""Verify rollforward schedule control rows have complete formulas.

Catches the most common failure: control rows with formulas only in the
totals column (O), leaving month columns (B-N) empty.

Usage:
    python verify_rollforward.py <workbook_path> [sheet_name]

If sheet_name is omitted, checks all sheets.
"""
import sys
import openpyxl
from openpyxl.utils import get_column_letter

def verify_control_rows(ws, control_rows=None, month_start=2, month_end=14, totals_col=15):
    """Verify control rows have formulas across ALL month columns.

    Args:
        ws: worksheet to verify
        control_rows: list of row numbers to check (default [13, 14, 15])
        month_start: first month column (default 2 = B)
        month_end: last month column (default 14 = N)
        totals_col: totals column (default 15 = O)

    Returns:
        list of (row, col, issue) tuples for any problems found
    """
    if control_rows is None:
        control_rows = [13, 14, 15]  # Month Totals, Ending Balance, Variance

    issues = []

    for row in control_rows:
        # Check each month column has content
        for col in range(month_start, month_end + 1):
            cell = ws.cell(row=row, column=col)
            val = cell.value
            if val is None:
                issues.append((row, col, f"Missing formula/value in {get_column_letter(col)}{row}"))

        # Check totals column has content
        cell = ws.cell(row=row, column=totals_col)
        if cell.value is None:
            issues.append((row, totals_col, f"Missing formula/value in totals column O{row}"))

    return issues

def verify_gl_balance_row(ws, gl_row=16, totals_col=15):
    """Verify GL Balance row only has value in totals column.

    GL Balance should be a static value in column O only, not in month columns.
    """
    issues = []

    # Month columns should be empty
    for col in range(2, totals_col):
        cell = ws.cell(row=gl_row, column=col)
        if cell.value is not None:
            issues.append((gl_row, col, f"GL Balance should only be in column O, found value in {get_column_letter(col)}{gl_row}"))

    # Totals column should have value
    cell = ws.cell(row=gl_row, column=totals_col)
    if cell.value is None:
        issues.append((gl_row, totals_col, f"Missing GL Balance value in O{gl_row}"))

    return issues

def verify_workbook(path, sheet_name=None):
    """Verify rollforward structure in workbook."""
    wb = openpyxl.load_workbook(path, data_only=False)

    sheets_to_check = [sheet_name] if sheet_name else wb.sheetnames
    all_issues = []

    for name in sheets_to_check:
        if name not in wb.sheetnames:
            print(f"Sheet '{name}' not found")
            continue

        ws = wb[name]
        print(f"\n=== Checking {name} ===")

        # Check control rows
        issues = verify_control_rows(ws)
        for row, col, msg in issues:
            print(f"  ISSUE: {msg}")
        all_issues.extend(issues)

        # Check GL Balance row
        gl_issues = verify_gl_balance_row(ws)
        for row, col, msg in gl_issues:
            print(f"  ISSUE: {msg}")
        all_issues.extend(gl_issues)

        if not issues and not gl_issues:
            print(f"  OK: Control rows complete")

    return len(all_issues) == 0

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python verify_rollforward.py <workbook_path> [sheet_name]")
        sys.exit(1)

    path = sys.argv[1]
    sheet = sys.argv[2] if len(sys.argv) > 2 else None

    ok = verify_workbook(path, sheet)
    sys.exit(0 if ok else 1)