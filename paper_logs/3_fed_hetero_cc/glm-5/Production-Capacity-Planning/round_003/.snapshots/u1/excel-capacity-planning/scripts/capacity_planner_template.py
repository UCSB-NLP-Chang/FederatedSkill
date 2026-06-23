#!/usr/bin/env python3
"""
Template for iterative capacity/backlog planning from Excel data.

IMPORTANT: CAPACITY_RULES and THRESHOLDS below are EXAMPLES only.
You MUST adapt them to match your specific task requirements before running.
Check the task prompt for exact capacity values per days-worked.
"""
import openpyxl
import sys

def run_planning(input_path, output_plan_path, output_summary_path):
    wb_in = openpyxl.load_workbook(input_path, data_only=True)
    ws_in = wb_in.active
    
    # 1. Detect orientation & extract data
    # Scan first few rows for week numbers to find header row
    header_row_idx = None
    week_col_start = None
    for r_idx, row in enumerate(ws_in.iter_rows(min_row=1, max_row=10, values_only=False), 1):
        for c_idx, cell in enumerate(row, 1):
            if isinstance(cell.value, (int, float)) and 1 <= cell.value <= 53:
                header_row_idx = r_idx
                week_col_start = c_idx
                break
        if header_row_idx: break
            
    if not header_row_idx:
        raise ValueError("Could not detect week headers. Check Excel layout.")
        
    # Extract weeks
    weeks = []
    for cell in ws_in[header_row_idx][week_col_start-1:]:
        if cell.value is not None:
            weeks.append(cell.value)
            
    # Find demand row (look for "Demand" or "Total" in first column)
    demand_row_idx = None
    for r_idx, row in enumerate(ws_in.iter_rows(min_row=1, max_row=20, values_only=False), 1):
        first_cell = row[0]
        if first_cell.value and ("demand" in str(first_cell.value).lower() or "total" in str(first_cell.value).lower()):
            demand_row_idx = r_idx
            break
            
    if not demand_row_idx:
        raise ValueError("Could not detect demand row.")
        
    demands = []
    for cell in ws_in[demand_row_idx][week_col_start-1:]:
        demands.append(cell.value if cell.value is not None else 0.0)
        
    # Trim to match weeks length
    demands = demands[:len(weeks)]
    
    # ============================================================
    # CONFIGURATION - ADAPT THESE TO MATCH YOUR TASK PROMPT
    # The values below are EXAMPLES. Check your task for actual values.
    # ============================================================
    INITIAL_BACKLOG = 0.0  # Set from task prompt
    
    # EXAMPLE capacity rules - REPLACE with task-specific values
    # Common patterns: 6-day=180hrs(with OT), 5-day=150hrs, 4-day=120hrs
    # But VERIFY against your specific task requirements!
    CAPACITY_RULES = {
        6: {"std_hrs": 160, "ot_hrs": 20, "total": 180},
        5: {"std_hrs": 150, "ot_hrs": 0, "total": 150},
        4: {"std_hrs": 120, "ot_hrs": 0, "total": 120}
    }
    
    plan_rows = []
    calc_backlog = INITIAL_BACKLOG
    
    for i, week in enumerate(weeks):
        demand = demands[i]
        # Decision logic: determine days worked based on backlog & demand
        # ADAPT this threshold logic to match your task requirements
        days = 6 if (calc_backlog > 0 or demand > 150) else 4
        cap = CAPACITY_RULES[days]["total"]
        ot = CAPACITY_RULES[days]["ot_hrs"]
        
        start_pd = max(0, calc_backlog) # Reported past due
        calc_backlog = calc_backlog + demand - cap
        end_pd = max(0, calc_backlog)   # Reported end backlog (optional)
        
        plan_rows.append([week, days, demand, cap, start_pd, calc_backlog, ot])
        
    # Write Excel
    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = "Plan"
    # ADAPT headers to match task requirements exactly
    headers = ["Week", "Days Worked", "Scheduled Demand (Std Hrs)", "Weekly Capacity (Std Hrs)", 
               "Start of Week Past Due (Std Hrs)", "End of Week Backlog/Buffer (Std Hrs)", "Overtime Hours"]
    ws_out.append(headers)
    for row in plan_rows:
        ws_out.append(row)
    wb_out.save(output_plan_path)
    
    # Write Summary - ADAPT format to match task requirements
    first_5 = next((r[0] for r in plan_rows if r[1] == 5), None)
    first_4 = next((r[0] for r in plan_rows if r[1] == 4), None)
    summary_lines = [
        f"First_Week_5_Days: {first_5 if first_5 else 'N/A'}",
        f"First_Week_4_Days: {first_4 if first_4 else 'N/A'}",
        "Summary: Adapt this line to explain the catch-up trajectory and step-down logic."
    ]
    with open(output_summary_path, "w") as f:
        f.write("\n".join(summary_lines) + "\n")
        
    print("Planning complete.")

if __name__ == "__main__":
    if len(sys.argv) != 4:
        print("Usage: python capacity_planner_template.py <input.xlsx> <output_plan.xlsx> <output_summary.txt>")
        sys.exit(1)
    run_planning(sys.argv[1], sys.argv[2], sys.argv[3])
