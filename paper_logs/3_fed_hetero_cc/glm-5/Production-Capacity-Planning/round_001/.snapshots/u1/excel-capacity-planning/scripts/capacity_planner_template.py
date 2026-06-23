#!/usr/bin/env python3
"""
Template for iterative capacity/backlog planning from Excel data.
Adapt CAPACITY_RULES and THRESHOLDS to match task requirements.
"""
import openpyxl
import sys

def run_planning(input_path, output_plan_path, output_summary_path):
    wb_in = openpyxl.load_workbook(input_path, data_only=True)
    ws_in = wb_in.active

    # Extract data: assume col A = period, col B = demand
    rows = list(ws_in.iter_rows(min_row=2, values_only=True))
    periods = [r[0] for r in rows]
    demands = [r[1] for r in rows]

    # Configuration (ADAPT THESE)
    INITIAL_BACKLOG = 0.0
    CAPACITY_RULES = {
        6: {"std_hrs": 160, "ot_hrs": 20, "total": 180},
        5: {"std_hrs": 150, "ot_hrs": 0, "total": 150},
        4: {"std_hrs": 120, "ot_hrs": 0, "total": 120}
    }

    plan_rows = []
    current_backlog = INITIAL_BACKLOG

    for i, period in enumerate(periods):
        demand = demands[i]
        # Decision logic: determine days worked based on backlog & demand
        # Example: if backlog > 0 or demand > 150 -> 6 days
        days = 6 if (current_backlog > 0 or demand > 150) else 4
        cap = CAPACITY_RULES[days]["total"]
        ot = CAPACITY_RULES[days]["ot_hrs"]

        start_backlog = current_backlog
        end_backlog = start_backlog + demand - cap
        current_backlog = end_backlog # Keep negative for buffer tracking, or max(0, end_backlog) per rules

        plan_rows.append([period, days, demand, cap, start_backlog, end_backlog, ot])

    # Write Excel
    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = "Plan"
    headers = ["Week", "Days Worked", "Scheduled Demand (Std Hrs)", "Weekly Capacity (Std Hrs)",
               "Start of Week Past Due (Std Hrs)", "End of Week Backlog/Buffer (Std Hrs)", "Overtime Hours"]
    ws_out.append(headers)
    for row in plan_rows:
        ws_out.append(row)
    wb_out.save(output_plan_path)

    # Write Summary
    first_5 = next((r[0] for r in plan_rows if r[1] == 5), None)
    first_4 = next((r[0] for r in plan_rows if r[1] == 4), None)
    summary_lines = [
        f"First_Week_5_Days: {first_5 if first_5 else 'N/A'}",
        f"First_Week_4_Days: {first_4 if first_4 else 'N/A'}",
        "Summary: Adapt this line to explain the catch-up trajectory and step-down logic."
    ]
    with open(output_summary_path, "w") as f:
        f.write("\n".join(summary_lines) + "\n")

    print("Planning complete.")

if __name__ == "__main__":
    if len(sys.argv) != 4:
        print("Usage: python capacity_planner_template.py <input.xlsx> <output_plan.xlsx> <output_summary.txt>")
        sys.exit(1)
    run_planning(sys.argv[1], sys.argv[2], sys.argv[3])