#!/usr/bin/env python3
"""Helper for creating formatted Excel files from extracted image data."""

from openpyxl import Workbook
from typing import List, Dict, Any

def create_excel_from_records(
    filepath: str,
    records: List[Dict[str, Any]],
    sheet_name: str = "data",
    columns: List[str] = None
) -> None:
    """
    Create an Excel file from a list of dictionaries.

    Args:
        filepath: Output Excel file path
        records: List of dictionaries with extracted data
        sheet_name: Name for the Excel sheet
        columns: Column order (if None, uses keys from first record)
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    if not records:
        wb.save(filepath)
        return

    # Determine column order
    if columns is None:
        columns = list(records[0].keys())

    # Write header
    for col_idx, col_name in enumerate(columns, 1):
        ws.cell(row=1, column=col_idx, value=col_name)

    # Write data rows - pass raw values, do NOT round or format
    for row_idx, record in enumerate(records, 2):
        for col_idx, col_name in enumerate(columns, 1):
            value = record.get(col_name)
            ws.cell(row=row_idx, column=col_idx, value=value)

    wb.save(filepath)


def verify_excel_content(filepath: str, expected_rows: int) -> Dict[str, Any]:
    """
    Verify Excel file has expected structure.

    Returns dict with verification results.
    """
    from openpyxl import load_workbook

    wb = load_workbook(filepath)
    ws = wb.active

    # Count actual data rows (excluding header)
    data_rows = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(cell is not None for cell in row):
            data_rows += 1

    return {
        "sheet_name": ws.title,
        "total_rows": ws.max_row,
        "data_rows": data_rows,
        "columns": ws.max_column,
        "expected_rows": expected_rows,
        "match": data_rows == expected_rows
    }


if __name__ == "__main__":
    # Example usage
    import sys

    # Demo: verify an existing file
    if len(sys.argv) > 1:
        result = verify_excel_content(sys.argv[1], int(sys.argv[2]) if len(sys.argv) > 2 else 0)
        print(f"Verification: {result}")
    else:
        print("Usage: python create_excel.py <filepath> [expected_rows]")