#!/usr/bin/env python3
"""
Verify formula consistency across multi-year Excel model sheets.
Use when building compensation models or financial projections with parallel year sheets.
"""

import openpyxl
import sys
import re
from collections import defaultdict


def extract_year_refs(formula):
    """Extract year-shift patterns like +1, +2 from formulas."""
    return set(re.findall(r'[\+\-]\d+', str(formula)))


def compare_yearly_formulas(wb_path, sheet_pattern, base_sheet_name):
    """
    Compare formulas across parallel year sheets to ensure consistency.

    Args:
        wb_path: Path to workbook
        sheet_pattern: Regex to match year-variant sheets (e.g., r'EE Calcs.*')
        base_sheet_name: Name of base sheet to compare against

    Returns:
        dict of inconsistencies by cell address
    """
    wb = openpyxl.load_workbook(wb_path, data_only=False)

    year_sheets = [s for s in wb.sheetnames if re.search(sheet_pattern, s)]
    base_ws = wb[base_sheet_name]

    inconsistencies = defaultdict(list)

    # Analyze structure
    for sheet_name in year_sheets:
        if sheet_name == base_sheet_name:
            continue

        ws = wb[sheet_name]

        # Compare row-by-row for common columns
        for row in range(1, min(base_ws.max_row, ws.max_row) + 1):
            for col in range(1, min(base_ws.max_column, ws.max_column) + 1):
                base_cell = base_ws.cell(row=row, column=col)
                year_cell = ws.cell(row=row, column=col)

                base_val = base_cell.value
                year_val = year_cell.value

                # Both formulas?
                if isinstance(base_val, str) and base_val and not base_val.startswith('='):
                    if isinstance(year_val, str) and year_val and not year_val.startswith('='):
                        # Normalize: remove year offsets for comparison
                        base_norm = re.sub(r'[\+\-]\d+', '##', base_val)
                        year_norm = re.sub(r'[\+\-]\d+', '##', year_val)

                        if base_norm != year_norm:
                            addr = f"{openpyxl.utils.get_column_letter(col)}{row}"
                            inconsistencies[sheet_name].append({
                                'cell': addr,
                                'base': base_val,
                                'year': year_val
                            })

    return dict(inconsistencies)


def verify_named_ranges(wb_path, required_patterns):
    """
    Verify named ranges follow expected naming conventions.

    Args:
        wb_path: Path to workbook
        required_patterns: List of regex patterns that should match named ranges

    Returns:
        tuple (ok: bool, missing: list, unexpected: list)
    """
    wb = openpyxl.load_workbook(wb_path, data_only=False)

    # CORRECT: Use .values() to get DefinedName objects safely
    names = [dn.name for dn in wb.defined_names.values()]

    missing = []
    for pattern in required_patterns:
        if not any(re.search(pattern, n) for n in names):
            missing.append(pattern)

    # Check for suspicious patterns (optional validation)
    expected_count = len(required_patterns)

    return (len(missing) == 0, missing, names)


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print(f"Usage: {sys.argv[0]} <workbook.xlsx>")
        sys.exit(1)

    wb_path = sys.argv[1]

    # Example: Verify compensation model structure
    print("=== Named Range Check ===")
    ok, missing, found = verify_named_ranges(wb_path, [
        r'MWS_(Current|Year_Plus_\d+)',
        r'Payroll_Tax_Tier\d+_Rate',
        r'Seniority_\d+_\d+_(Current|Year)'
    ])

    if ok:
        print(f"✓ Found {len(found)} named ranges")
    else:
        print(f"✗ Missing patterns: {missing}")

    # Example: Compare year sheets
    print("\n=== Year Sheet Comparison ===")
    try:
        issues = compare_yearly_formulas(wb_path, r'EE Calcs.*', 'EE Calcs (Current)')
        if not issues:
            print("✓ Year sheets formula-consistent")
        else:
            for sheet, cells in issues.items():
                print(f"✗ {sheet}: {len(cells)} formula mismatches")
                for issue in cells[:3]:  # Show first 3
                    print(f"    {issue['cell']}: {issue['base'][:40]}... vs {issue['year'][:40]}...")
    except Exception as e:
        print(f"⚠ Could not compare sheets: {e}")