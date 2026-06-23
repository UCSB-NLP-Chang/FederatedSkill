#!/usr/bin/env python3
"""Verify rollforward workbook structure, formulas, and formatting.

Supports deferred revenue, accrual, and warranty reserve variants.
"""
import sys
import openpyxl


def verify(path):
    wb = openpyxl.load_workbook(path)
    errors = []

    # Detect summary sheet name (Deferred, Warranty, etc.)
    summary_name = None
    if wb.sheetnames[0] in ["Deferred Summary", "Warranty Summary", "Summary"]:
        summary_name = wb.sheetnames[0]
    else:
        errors.append(f"First sheet should be Summary, got: {wb.sheetnames[0]}")

    detail_sheets = [name for name in wb.sheetnames if name != summary_name]

    for name in detail_sheets:
        ws = wb[name]

        # Find header row (usually row 5 or 6)
        header_row = None
        for r in range(1, 10):
            if ws.cell(row=r, column=1).value in ["Customer", "Claim Group", "Accrual Bucket"]:
                header_row = r
                break

        if not header_row:
            errors.append(f"{name}: Could not find header row")
            continue

        # Find control rows by label
        totals_row = None
        ending_row = None
        variance_row = None
        gl_row = None

        for r in range(header_row + 1, header_row + 20):
            label = ws.cell(row=r, column=1).value
            if label == "Period Totals":
                totals_row = r
            elif label == "Ending Balance":
                ending_row = r
            elif label == "Variance":
                variance_row = r
            elif label == "GL Balance":
                gl_row = r

        if not all([totals_row, ending_row, variance_row, gl_row]):
            errors.append(f"{name}: Missing control rows (totals={totals_row}, ending={ending_row}, variance={variance_row}, gl={gl_row})")
            continue

        # Check for formulas in key columns (O for 4-month, L for 3-month)
        # Check column O (15) for total formulas
        o_totals = ws.cell(row=totals_row, column=15).value
        o_ending = ws.cell(row=ending_row, column=15).value

        if not isinstance(o_totals, str) or not o_totals.startswith("="):
            errors.append(f"{name}: Missing formula in O{totals_row} (Period Totals)")
        if not isinstance(o_ending, str) or not o_ending.startswith("="):
            errors.append(f"{name}: Missing formula in O{ending_row} (Ending Balance)")

        # Check number formatting on monetary cells
        for r in range(header_row + 1, gl_row + 1):
            for c in range(2, 16):  # B through O
                cell = ws.cell(row=r, column=c)
                if isinstance(cell.value, (int, float)) or (isinstance(cell.value, str) and cell.value.startswith("=")):
                    if cell.number_format != '#,##0.00':
                        errors.append(f"{name}: Cell {cell.coordinate} missing #,##0.00 format")

    if errors:
        print("VERIFICATION FAILED:")
        for e in errors:
            print(f"  - {e}")
        sys.exit(1)
    else:
        print(f"All checks passed! Summary: {summary_name}, Details: {detail_sheets}")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: verify_workbook.py <workbook_path>")
        sys.exit(1)
    verify(sys.argv[1])