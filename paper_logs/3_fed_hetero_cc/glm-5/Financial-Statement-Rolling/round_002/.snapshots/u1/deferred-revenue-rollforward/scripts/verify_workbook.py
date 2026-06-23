#!/usr/bin/env python3
"""Verify deferred revenue workbook structure, formulas, and formatting.

Usage:
    python3 verify_workbook.py <workbook_path>
"""
import sys
import openpyxl


def verify(path):
    wb = openpyxl.load_workbook(path)
    errors = []

    # Check sheet order
    if wb.sheetnames[0] != "Deferred Summary":
        errors.append(f"First sheet must be 'Deferred Summary', got '{wb.sheetnames[0]}'")

    for name in wb.sheetnames:
        ws = wb[name]
        if name == "Deferred Summary":
            continue

        # Check headers exist
        headers = [ws.cell(row=5, column=c).value for c in range(1, 18)]
        if not all(h for h in headers[:5]):
            errors.append(f"{name}: Missing headers at row 5")

        # Check control rows exist with formulas
        for r in range(10, 14):
            label = ws.cell(row=r, column=1).value
            if not label:
                errors.append(f"{name}: Missing control row label at row {r}")
            o_val = ws.cell(row=r, column=15).value
            if not isinstance(o_val, str) or not o_val.startswith("="):
                errors.append(f"{name}: Missing formula in O{r}")

        # Check number format on monetary cells
        for r in range(6, 10):
            for c in range(2, 15):
                cell = ws.cell(row=r, column=c)
                if isinstance(cell.value, (int, float)) and cell.number_format != '#,##0.00':
                    errors.append(f"{name}: Cell {cell.coordinate} missing #,##0.00 format")

    if errors:
        print("VERIFICATION FAILED:")
        for e in errors:
            print(f"  - {e}")
        sys.exit(1)
    else:
        print("All checks passed!")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: verify_workbook.py <workbook_path>")
        sys.exit(1)
    verify(sys.argv[1])
