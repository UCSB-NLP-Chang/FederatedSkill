#!/usr/bin/env python3
"""
Template for multi-sheet Excel pivot reports using pandas.
DO NOT use openpyxl pivot API - use pd.pivot_table() instead.
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font

def create_pivot_report(source_csv: str, output_path: str, lookup_df: pd.DataFrame = None,
                        join_key: str = None, derived_cols: list = None,
                        pivot_specs: list = None):
    """
    Create Excel report with pivot tables and enriched source data.

    Args:
        source_csv: Path to source CSV/Excel file
        output_path: Path for output Excel file
        lookup_df: Optional lookup DataFrame (from PDF extraction)
        join_key: Column name for merge
        derived_cols: List of dicts with 'name', 'formula' (lambda function)
        pivot_specs: List of dicts with 'sheet_name', 'index', 'values', 'aggfunc', 'columns'
    """
    df = pd.read_csv(source_csv) if source_csv.endswith('.csv') else pd.read_excel(source_csv)

    # Join with lookup if provided
    if lookup_df is not None and join_key:
        # Align join-key types
        df[join_key] = df[join_key].astype(str).str.strip()
        lookup_df[join_key] = lookup_df[join_key].astype(str).str.strip()
        df = df.merge(lookup_df, on=join_key, how='left', suffixes=('', '_lookup'))

        # Validate merge
        if len(df) == 0:
            raise ValueError(f"Merge produced 0 rows - join key '{join_key}' mismatch")

    # Add derived columns
    if derived_cols:
        for col in derived_cols:
            df[col['name']] = col['formula'](df)

    # Create workbook
    wb = Workbook()
    header_font = Font(bold=True)

    def write_df(ws, df, columns=None):
        data = df[columns] if columns else df
        for r_idx, row in enumerate(dataframe_to_rows(data, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
        for cell in ws[1]:
            cell.font = header_font

    # Source data sheet
    ws_source = wb.active
    ws_source.title = 'SourceData'
    write_df(ws_source, df)

    # Pivot sheets
    if pivot_specs:
        for spec in pivot_specs:
            ws_pivot = wb.create_sheet(spec['sheet_name'])
            pivot = pd.pivot_table(
                df,
                values=spec.get('values'),
                index=spec.get('index'),
                columns=spec.get('columns'),
                aggfunc=spec.get('aggfunc', 'mean')
            ).reset_index()

            # Flatten column names if multi-index from columns parameter
            if isinstance(pivot.columns, pd.MultiIndex):
                pivot.columns = [str(c) if c else '' for c in pivot.columns]

            write_df(ws_pivot, pivot)

    wb.save(output_path)
    print(f"Report saved to {output_path}")
    return output_path

def verify_report(xlsx_path: str, expected_sheets: list = None):
    """Verify the generated Excel file structure."""
    wb = pd.ExcelFile(xlsx_path)
    print(f"Sheets: {wb.sheet_names}")

    if expected_sheets:
        for sheet in expected_sheets:
            if sheet not in wb.sheet_names:
                raise ValueError(f"Missing sheet: {sheet}")

    for sheet in wb.sheet_names:
        df = wb.parse(sheet)
        print(f"  {sheet}: {df.shape}")

    return wb.sheet_names

if __name__ == '__main__':
    import sys
    if len(sys.argv) < 3:
        print("Usage: python create_pivot_report.py <source.csv> <output.xlsx>")
        sys.exit(1)
    create_pivot_report(sys.argv[1], sys.argv[2])