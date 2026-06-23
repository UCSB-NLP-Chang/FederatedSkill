#!/usr/bin/env python3
"""
Verify Excel formula integrity after generation.
Run this after creating a workbook to catch #REF! and #NAME? issues early.
"""
import openpyxl
import sys
from pathlib import Path

def verify_workbook(filepath):
    """Check for common formula errors in generated workbook."""
    wb = openpyxl.load_workbook(filepath, data_only=False)
    issues = []
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    formula = cell.value
                    
                    # Check for unquoted sheet names with spaces
                    if any(c in formula for c in [' ', '(', ')', '-']):
                        if '!' in formula and not ("'" in formula or '"' in formula):
                            # Sheet reference without quotes
                            if any(x in formula for x in ['Calcs', 'EE ']):
                                issues.append(f"{sheet_name}!{cell.coordinate}: Possible unquoted sheet ref in: {formula[:50]}...")
                    
                    # Check for Python escape artifacts
                    if '\\' in formula:
                        issues.append(f"{sheet_name}!{cell.coordinate}: Contains backslash (Python artifact): {formula}")
    
    # Check defined names
    print(f"Defined names: {len(wb.defined_names)}")
    for name in wb.defined_names:
        print(f"  {name}")
    
    if issues:
        print(f"\nFound {len(issues)} potential issues:")
        for issue in issues[:10]:  # Show first 10
            print(f"  - {issue}")
        if len(issues) > 10:
            print(f"  ... and {len(issues) - 10} more")
        return 1
    else:
        print("No obvious formula issues detected.")
        return 0

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(f"Usage: {sys.argv[0]} <workbook.xlsx>")
        sys.exit(1)
    
    sys.exit(verify_workbook(sys.argv[1]))