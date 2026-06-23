#!/usr/bin/env python3
"""Reusable openpyxl workbook structural verifier.
Adjust EXPECTED_* constants per task before running.
"""
import sys
import openpyxl

# --- TASK-SPECIFIC CONSTANTS (Override these) ---
EXPECTED_SHEETS = ["Summary", "Assumptions", "Roster", "Calculations", "EE Calcs (Current)", "EE Calcs (Yr+1)", "EE Calcs (Yr+2)"]
EXPECTED_ROWS = {"Roster": 103, "EE Calcs (Current)": 107}
EXPECTED_NAMED_RANGES = 39
FORMULA_CHECK_SHEET = "Summary"
FORMULA_CHECK_CELLS = ["C27", "C28", "C29"]  # Sample cells that must contain formulas
# ------------------------------------------------

def verify(path):
    wb = openpyxl.load_workbook(path)
    errors = []

    # 1. Sheet order
    if wb.sheetnames != EXPECTED_SHEETS:
        errors.append(f"Sheet mismatch. Expected: {EXPECTED_SHEETS}, Got: {wb.sheetnames}")
    else:
        print("✓ Sheet order matches")

    # 2. Row counts
    for sheet_name, count in EXPECTED_ROWS.items():
        if sheet_name in wb.sheetnames:
            actual = wb[sheet_name].max_row
            if actual != count:
                errors.append(f"{sheet_name} row count: {actual} != {count}")
    if not any("row count" in e for e in errors):
        print("✓ Row counts match")

    # 3. Named ranges
    dn_list = wb.defined_names
    dn_count = len(dn_list)
    if dn_count != EXPECTED_NAMED_RANGES:
        errors.append(f"Named ranges: {dn_count} != {EXPECTED_NAMED_RANGES}")
    else:
        print(f"✓ Named ranges: {dn_count}")

    # 4. Formula presence check
    if FORMULA_CHECK_SHEET in wb.sheetnames:
        ws = wb[FORMULA_CHECK_SHEET]
        missing_formulas = []
        for cell_ref in FORMULA_CHECK_CELLS:
            cell = ws[cell_ref]
            if not cell.value or not str(cell.value).startswith("="):
                missing_formulas.append(cell_ref)
        if missing_formulas:
            errors.append(f"Missing formulas in {FORMULA_CHECK_SHEET}: {missing_formulas}")
        else:
            print(f"✓ Formula checks passed in {FORMULA_CHECK_SHEET}")

    if errors:
        print("\n❌ Verification FAILED:")
        for e in errors:
            print(f"  - {e}")
        sys.exit(1)
    else:
        print("\n✅ Verification complete. Workbook structure is valid.")

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python verify_xlsx.py <path_to_workbook.xlsx>")
        sys.exit(1)
    verify(sys.argv[1])
