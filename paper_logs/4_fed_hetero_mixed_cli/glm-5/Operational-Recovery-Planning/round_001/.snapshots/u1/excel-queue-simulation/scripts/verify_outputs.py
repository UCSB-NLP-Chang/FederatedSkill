#!/usr/bin/env python3
"""Verify Excel queue simulation outputs against standard constraints."""
import sys
import openpyxl

def verify_summary(path):
    with open(path) as f:
        lines = f.readlines()
    if len(lines) != 3:
        print(f"FAIL: Expected 3 lines, got {len(lines)}")
        sys.exit(1)
    summary = lines[2].strip()
    words = summary.split()
    if len(words) > 60:
        print(f"FAIL: Summary too long ({len(words)} words, max 60)")
        sys.exit(1)
    sentences = summary.count('.')
    if sentences > 3:
        print(f"FAIL: Too many sentences ({sentences}, max 3)")
        sys.exit(1)
    print("Summary: OK")

def verify_workbook(path, sheet_name="Plan", expected_rows=40):
    wb = openpyxl.load_workbook(path)
    if sheet_name not in wb.sheetnames:
        print(f"FAIL: Sheet '{sheet_name}' not found")
        sys.exit(1)
    ws = wb[sheet_name]
    if ws.max_row != expected_rows + 1:
        print(f"FAIL: Expected {expected_rows+1} rows, got {ws.max_row}")
        sys.exit(1)
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    if len(headers) < 7:
        print(f"FAIL: Expected at least 7 headers, got {len(headers)}")
        sys.exit(1)
    weeks = [ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)]
    if weeks != list(range(1, expected_rows + 1)):
        print("FAIL: Weeks not ascending 1..N")
        sys.exit(1)
    print("Workbook: OK")

if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage: verify_outputs.py <summary.txt> <workbook.xlsx>")
        sys.exit(1)
    verify_summary(sys.argv[1])
    verify_workbook(sys.argv[2])
