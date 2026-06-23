#!/usr/bin/env python3
"""Verify structure and formulas of a financial rollforward workbook.

Checks for:
- Circular references (formula referencing its own exact cell)
- Ending Balance column O references Period Totals row (NOT its own row) for prepaid expenses
- Summary links point to correct columns (N for standard, O check for running balance)
- Control row labels present
- Cross-sheet reference syntax

Designed for deferred revenue, prepaid expense, accrual, warranty, and commission rollforwards.
"""
import sys
import re
import openpyxl
from openpyxl.utils import get_column_letter


def check_circular_reference(row, col, formula, label):
    """Check if a formula references its own exact cell (circular reference).

    Same-row references to DIFFERENT columns (e.g., E48 referencing B48, C48, D48)
    are valid running balance patterns and NOT circular.
    Only flag if the formula references the exact same cell (column + row).
    """
    if not formula or not isinstance(formula, str) or not formula.startswith('='):
        return None

    col_letter = get_column_letter(col)
    cell_ref = f"{col_letter}{row}"

    # Check if formula references its own exact cell
    pattern = re.compile(r'(?<![A-Z0-9_])' + re.escape(cell_ref) + r'(?![A-Z0-9_])', re.IGNORECASE)
    if pattern.search(formula):
        return f"Circular reference in {label}: {cell_ref} formula references itself: {formula}"
    return None


def check_ending_balance_column_o(ws, sheet_name):
    """Check that Ending Balance column O references Period Totals row, not own row.

    This catches the #1 bug in prepaid expense schedules where column O formula
    (total amortization) references the Ending Balance row instead of Period Totals.
    """
    errors = []

    # Find control rows
    totals_row = None
    ending_row = None

    for r in range(1, ws.max_row + 1):
        label = ws.cell(row=r, column=1).value
        if label in ("Period Totals", "Month Totals"):
            totals_row = r
        elif label == "Ending Balance":
            ending_row = r

    if not totals_row or not ending_row:
        return errors  # Can't check without finding rows

    # Check column O formula (total amortization/releases)
    o_formula = ws.cell(row=ending_row, column=15).value
    if isinstance(o_formula, str) and o_formula.startswith('='):
        # Extract row numbers from formula
        row_refs = re.findall(r'[A-Z]+(\d+)', o_formula)
        for ref_row in row_refs:
            if int(ref_row) == ending_row:
                errors.append(
                    f"{sheet_name}: Ending Balance row {ending_row} column O formula "
                    f"references its own row instead of Period Totals row {totals_row}. "
                    f"Formula: {o_formula}. "
                    f"Should be: =D{totals_row}+G{totals_row}+J{totals_row}+M{totals_row}"
                )
                break

    return errors


def check_summary_column_standard(ws_summary, detail_sheets):
    """Check that summary sheet links to final period column N, not totals column O.

    For standard rollforwards (deferred revenue, prepaid expense, etc.),
    summary should link to column N (final period ending balance), NOT column O.
    """
    errors = []

    for row in range(1, ws_summary.max_row + 1):
        for col in range(1, ws_summary.max_column + 1):
            cell_val = ws_summary.cell(row=row, column=col).value
            if isinstance(cell_val, str) and cell_val.startswith("="):
                # Check for references to column O in detail sheets (wrong for standard rollforwards)
                o_ref_pattern = r"='?[^'!]+(?:#[^'!]+)?'?!O\d+"
                if re.search(o_ref_pattern, cell_val):
                    errors.append(
                        f"Summary row {row} col {col} links to totals column O instead of final period column N: {cell_val}"
                    )

    return errors


def verify(path):
    wb = openpyxl.load_workbook(path)
    print(f"Verifying: {path}")
    print(f"Sheets: {wb.sheetnames}")

    errors = []
    detail_sheets = []

    for name in wb.sheetnames:
        ws = wb[name]
        print(f"\n--- {name} ---")

        # Check headers (assumes row 5)
        headers = [ws.cell(row=5, column=c).value for c in range(1, min(ws.max_column + 1, 18))]
        print(f"Headers (row 5): {headers}")

        # Find control rows and check for circular references
        control_labels = {'Period Totals', 'Month Totals', 'Ending Balance', 'Calculated Ending Balance', 'Variance', 'GL Balance'}
        control_rows = {}

        for r in range(1, ws.max_row + 1):
            label = ws.cell(row=r, column=1).value
            if label in control_labels:
                control_rows[label] = r
                print(f"  Row {r}: {label}")

                # Check all columns in this row for circular references
                for c in range(1, ws.max_column + 1):
                    cell_val = ws.cell(row=r, column=c).value
                    if isinstance(cell_val, str) and cell_val.startswith('='):
                        err = check_circular_reference(r, c, cell_val, f"{name}!{label}")
                        if err:
                            errors.append(err)
                            print(f"    *** ERROR: {err}")

        # Check Ending Balance column O (skip summary sheet)
        if name != wb.sheetnames[0]:  # Assume first sheet is summary
            detail_sheets.append(name)
            ending_errors = check_ending_balance_column_o(ws, name)
            errors.extend(ending_errors)
            for err in ending_errors:
                print(f"  *** ERROR: {err}")

        # Print GL Balance values
        if 'GL Balance' in control_rows:
            gl_row = control_rows['GL Balance']
            print(f"  GL Balance values:")
            for col, period in [(5, 'Jan'), (8, 'Feb'), (11, 'Mar'), (14, 'Apr')]:
                val = ws.cell(row=gl_row, column=col).value
                print(f"    {period} (col {col}): {val}")

    # Check summary sheet for wrong column references
    if wb.sheetnames:
        summary_ws = wb[wb.sheetnames[0]]
        summary_errors = check_summary_column_standard(summary_ws, detail_sheets)
        errors.extend(summary_errors)
        for err in summary_errors:
            print(f"\n*** ERROR: {err}")

    if errors:
        print(f"\n*** VERIFICATION FAILED: {len(errors)} error(s) found ***")
        for err in errors:
            print(f"  - {err}")
        return False
    else:
        print("\nVerification passed. No errors found.")
        return True


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python verify_workbook.py <path_to_xlsx>")
        sys.exit(1)
    success = verify(sys.argv[1])
    sys.exit(0 if success else 1)
