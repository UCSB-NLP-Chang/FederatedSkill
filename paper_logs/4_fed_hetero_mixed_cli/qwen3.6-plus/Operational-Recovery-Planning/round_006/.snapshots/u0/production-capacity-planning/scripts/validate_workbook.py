#!/usr/bin/env python3
"""Reusable validation helpers for Excel workbook verification."""

from openpyxl import load_workbook
from datetime import datetime, date
from typing import List, Dict, Any, Optional, Callable, Set


def to_date(val: Any) -> Optional[date]:
    """Canonical date conversion for openpyxl cell values.
    
    Handles datetime.datetime, datetime.date, and None.
    Use this BEFORE any date comparison to avoid TypeError.
    """
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    # Fallback: try parsing string
    try:
        return datetime.strptime(str(val), "%Y-%m-%d").date()
    except (ValueError, TypeError):
        return None


def validate_sheet_names(wb_path: str, expected_sheets: List[str]) -> Dict[str, bool]:
    """Verify all expected sheets exist with exact names."""
    wb = load_workbook(wb_path)
    results = {}
    for sheet in expected_sheets:
        results[sheet] = sheet in wb.sheetnames
    return results


def validate_date_range(
    wb_path: str,
    sheet_name: str,
    date_col: int,
    start_row: int,
    expected_start: date,
    expected_end: date
) -> Dict[str, Any]:
    """Verify date range in a sheet covers expected span."""
    wb = load_workbook(wb_path)
    ws = wb[sheet_name]

    first_date = to_date(ws.cell(row=start_row, column=date_col).value)
    last_row = start_row
    while ws.cell(row=last_row + 1, column=date_col).value:
        last_row += 1
    last_date = to_date(ws.cell(row=last_row, column=date_col).value)

    return {
        'first_date': first_date,
        'last_date': last_date,
        'expected_first': expected_start,
        'expected_last': expected_end,
        'first_match': first_date == expected_start if first_date else False,
        'last_match': last_date == expected_end if last_date else False,
        'row_count': last_row - start_row + 1
    }


def validate_weekend_zero_production(
    wb_path: str,
    sheet_name: str,
    date_col: int,
    production_cols: List[int],
    start_row: int,
    holidays: Optional[List[date]] = None
) -> Dict[str, Any]:
    """Verify production columns are zero on weekends and holidays."""
    wb = load_workbook(wb_path)
    ws = wb[sheet_name]
    holiday_set = set(holidays) if holidays else set()

    violations = []
    row = start_row

    while True:
        cell_val = ws.cell(row=row, column=date_col).value
        if not cell_val:
            break

        cell_date = to_date(cell_val)
        if cell_date is None:
            row += 1
            continue

        is_weekend = cell_date.weekday() >= 5
        is_holiday = cell_date in holiday_set

        if is_weekend or is_holiday:
            for col in production_cols:
                prod_val = ws.cell(row=row, column=col).value
                if prod_val and prod_val != 0:
                    violations.append({
                        'row': row,
                        'date': str(cell_date),
                        'column': col,
                        'value': prod_val
                    })
        row += 1

    return {
        'valid': len(violations) == 0,
        'violation_count': len(violations),
        'violations': violations[:10]
    }


def validate_cumulative_formulas(
    wb_path: str,
    sheet_name: str,
    cumul_cols: List[int],
    start_row: int,
    expected_first_formula: str = None,
    expected_subsequent_formula: str = None
) -> Dict[str, Any]:
    """Verify cumulative formulas are correctly structured."""
    wb = load_workbook(wb_path)
    ws = wb[sheet_name]

    results = {}

    for col in cumul_cols:
        first_formula = ws.cell(row=start_row, column=col).value
        second_formula = ws.cell(row=start_row + 1, column=col).value

        results[f'col_{col}'] = {
            'first_row_formula': str(first_formula) if first_formula else None,
            'second_row_formula': str(second_formula) if second_formula else None,
            'is_formula_first': str(first_formula).startswith('=') if first_formula else False,
            'is_formula_second': str(second_formula).startswith('=') if second_formula else False
        }

    return results


def validate_po_quantities(
    wb_path: str,
    sheet_name: str,
    date_col: int,
    po_cols: Dict[int, Dict[date, int]],
    start_row: int
) -> Dict[str, Any]:
    """Verify PO quantities on specific dates.

    po_cols format: {column_num: {date: expected_quantity}}
    """
    wb = load_workbook(wb_path)
    ws = wb[sheet_name]

    results = {'valid': True, 'mismatches': []}

    row = start_row
    date_to_row = {}

    while True:
        cell_val = ws.cell(row=row, column=date_col).value
        if not cell_val:
            break

        cell_date = to_date(cell_val)
        if cell_date:
            date_to_row[cell_date] = row
        row += 1

    for col, date_quantities in po_cols.items():
        for exp_date, exp_qty in date_quantities.items():
            if exp_date in date_to_row:
                actual = ws.cell(row=date_to_row[exp_date], column=col).value
                if actual != exp_qty:
                    results['valid'] = False
                    results['mismatches'].append({
                        'date': str(exp_date),
                        'column': col,
                        'expected': exp_qty,
                        'actual': actual
                    })
            else:
                results['valid'] = False
                results['mismatches'].append({
                    'date': str(exp_date),
                    'column': col,
                    'error': 'Date not found in sheet'
                })

    return results


def validate_formula_vs_constant(
    wb_path: str,
    sheet_name: str,
    formula_cols: List[int],
    constant_cols: List[int],
    start_row: int,
    end_row: int
) -> Dict[str, Any]:
    """Verify formula columns contain formulas and constant columns contain values."""
    wb = load_workbook(wb_path)
    ws = wb[sheet_name]

    results = {'formula_errors': [], 'constant_errors': []}

    for row in range(start_row, end_row + 1):
        for col in formula_cols:
            cell = ws.cell(row=row, column=col)
            if cell.data_type != 'f':  # 'f' = formula
                results['formula_errors'].append({
                    'row': row,
                    'col': col,
                    'value': str(cell.value)[:50]
                })

        for col in constant_cols:
            cell = ws.cell(row=row, column=col)
            if cell.data_type == 'f':
                results['constant_errors'].append({
                    'row': row,
                    'col': col,
                    'formula': str(cell.value)[:50]
                })

    results['valid'] = len(results['formula_errors']) == 0 and len(results['constant_errors']) == 0
    return results


def validate_shift_days(
    wb_path: str,
    sheet_name: str,
    date_col: int,
    production_cols: List[int],
    start_row: int,
    threshold_date: date,
    min_shift_days: int = 20,
    max_shift_days: int = 24,
    elevated_capacity: int = 170,
    standard_capacity_func: Optional[Callable[[date], int]] = None,
    holidays: Optional[List[date]] = None
) -> Dict[str, Any]:
    """Validate shift-day high-capacity window constraints.
    
    Checks:
    - Shift days are working days on/after threshold_date
    - Shift day count is within [min_shift_days, max_shift_days]
    - Production on shift days <= elevated_capacity
    - Production on non-shift days <= standard capacity
    """
    wb = load_workbook(wb_path)
    ws = wb[sheet_name]
    holiday_set = set(holidays) if holidays else set()

    results = {
        'valid': True,
        'shift_days': [],
        'shift_day_count': 0,
        'errors': []
    }

    # Collect all dates and production values
    row = start_row
    dates = []
    productions = {col: [] for col in production_cols}

    while True:
        cell_val = ws.cell(row=row, column=date_col).value
        if not cell_val:
            break
        
        cell_date = to_date(cell_val)
        if cell_date:
            dates.append(cell_date)
            for col in production_cols:
                productions[col].append(ws.cell(row=row, column=col).value)
        row += 1

    # Identify shift days (any day where production exceeds standard capacity)
    shift_day_indices = set()
    for idx, d in enumerate(dates):
        is_weekend = d.weekday() >= 5
        is_holiday = d in holiday_set
        if is_weekend or is_holiday:
            continue
        
        std_cap = standard_capacity_func(d) if standard_capacity_func else 135
        for col in production_cols:
            prod = productions[col][idx]
            if prod and prod > std_cap:
                shift_day_indices.add(idx)
                break

    # Validate shift days
    shift_dates = [dates[i] for i in shift_day_indices]
    results['shift_days'] = [str(d) for d in shift_dates]
    results['shift_day_count'] = len(shift_dates)

    if len(shift_dates) < min_shift_days:
        results['valid'] = False
        results['errors'].append(f"Shift day count {len(shift_dates)} < {min_shift_days}")
    if len(shift_dates) > max_shift_days:
        results['valid'] = False
        results['errors'].append(f"Shift day count {len(shift_dates)} > {max_shift_days}")

    for idx in shift_day_indices:
        d = dates[idx]
        if d < threshold_date:
            results['valid'] = False
            results['errors'].append(f"Shift day {d} is before threshold {threshold_date}")
        for col in production_cols:
            prod = productions[col][idx]
            if prod and prod > elevated_capacity:
                results['valid'] = False
                results['errors'].append(f"Shift day {d}: production {prod} > {elevated_capacity}")

    # Validate non-shift days respect standard caps
    for idx, d in enumerate(dates):
        if idx in shift_day_indices:
            continue
        is_weekend = d.weekday() >= 5
        is_holiday = d in holiday_set
        if is_weekend or is_holiday:
            continue
        
        std_cap = standard_capacity_func(d) if standard_capacity_func else 135
        for col in production_cols:
            prod = productions[col][idx]
            if prod and prod > std_cap:
                results['valid'] = False
                results['errors'].append(f"Non-shift day {d}: production {prod} > standard cap {std_cap}")

    return results