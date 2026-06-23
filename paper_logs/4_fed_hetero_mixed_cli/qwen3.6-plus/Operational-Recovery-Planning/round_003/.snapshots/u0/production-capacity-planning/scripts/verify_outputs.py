#!/usr/bin/env python3
"""Verify Excel time-series outputs and summary text files.

Usage: python3 verify_outputs.py <plan.xlsx> <summary.txt> [expected_weeks_start] [expected_weeks_end]

Checks:
- Excel sheet name
- Header row exact match
- Data row count
- Week sequence continuity
- Summary file line count and format
"""
import sys
import openpyxl

def verify_excel(path, expected_sheet, expected_headers, start_week, end_week):
    wb = openpyxl.load_workbook(path)
    errors = []

    if expected_sheet not in wb.sheetnames:
        errors.append(f"Sheet '{expected_sheet}' not found. Found: {wb.sheetnames}")
        return errors

    ws = wb[expected_sheet]

    # Check headers
    actual_headers = [ws.cell(row=1, column=c).value for c in range(1, len(expected_headers) + 1)]
    if actual_headers != expected_headers:
        errors.append(f"Header mismatch.\nExpected: {expected_headers}\nActual:   {actual_headers}")

    # Check data rows
    expected_count = end_week - start_week + 1
    actual_count = ws.max_row - 1
    if actual_count != expected_count:
        errors.append(f"Row count mismatch. Expected {expected_count}, got {actual_count}")

    # Check week sequence
    weeks = [ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)]
    expected_weeks = list(range(start_week, end_week + 1))
    if weeks != expected_weeks:
        errors.append(f"Week sequence mismatch.\nExpected: {expected_weeks[:5]}...{expected_weeks[-5:]}\nActual:   {weeks[:5]}...{weeks[-5:]}")

    # Check for None rows
    for r in range(2, ws.max_row + 1):
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, len(expected_headers) + 1)]
        if all(v is None for v in row_vals):
            errors.append(f"Row {r} is all-None (should be removed)")

    return errors

def verify_summary(path, expected_lines=3):
    errors = []
    with open(path, 'r') as f:
        content = f.read()

    lines = content.strip().split('\n')
    if len(lines) != expected_lines:
        errors.append(f"Summary line count: expected {expected_lines}, got {len(lines)}")

    # Check key format lines
    for line in lines:
        if line.startswith('First_Week_'):
            if ': ' not in line:
                errors.append(f"Key-value format error: {line}")

    return errors

if __name__ == '__main__':
    if len(sys.argv) < 3:
        print("Usage: verify_outputs.py <plan.xlsx> <summary.txt> [start_week] [end_week]")
        sys.exit(1)

    plan_path = sys.argv[1]
    summary_path = sys.argv[2]
    start_week = int(sys.argv[3]) if len(sys.argv) > 3 else 4
    end_week = int(sys.argv[4]) if len(sys.argv) > 4 else 52

    expected_headers = [
        'Week',
        'Days Worked',
        'Scheduled Demand (Std Hrs)',
        'Weekly Capacity (Std Hrs)',
        'Start of Week Past Due (Std Hrs)',
        'End of Week Backlog/Buffer (Std Hrs)',
        'Overtime Hours'
    ]

    all_errors = []
    all_errors.extend(verify_excel(plan_path, 'Plan', expected_headers, start_week, end_week))
    all_errors.extend(verify_summary(summary_path))

    if all_errors:
        print("VERIFICATION FAILED:")
        for e in all_errors:
            print(f"  - {e}")
        sys.exit(1)
    else:
        print("VERIFICATION PASSED: All checks OK")
        sys.exit(0)