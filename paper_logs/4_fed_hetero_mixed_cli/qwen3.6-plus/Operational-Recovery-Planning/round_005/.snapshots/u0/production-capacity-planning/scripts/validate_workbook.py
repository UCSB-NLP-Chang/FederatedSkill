#!/usr/bin/env python3
"""Reusable validation helpers for Excel workbook verification."""

from openpyxl import load_workbook
from datetime import datetime, date
from typing import List, Dict, Any, Optional, Callable


def validate_sheet_names(wb_path: str, expected_sheets: List[str]) -> Dict[str, bool]:
    """Verify all expected sheets exist with exact names."""
    wb = load_workbook(wb_path)
    results = {}
    for sheet in expected_sheets:
        results[sheet] = sheet in wb.sheetnames
    return results


def validate_date_range(
    wb_path: str,
    sheet_name: str,
    date_col: int,
    start_row: int,
    expected_start: date,
    expected_end: date
) -> Dict[str, Any]:
    """Verify date range in a sheet covers expected span."""
    wb = load_workbook(wb_path)
    ws = wb[sheet_name]

    first_date = ws.cell(row=start_row, column=date_col).value
    last_row = start_row
    while ws.cell(row=last_row + 1, column=date_col).value:
        last_row += 1
    last_date = ws.cell(row=last_row, column=date_col).value

    # Convert datetime.datetime to date if needed
    if hasattr(first_date, 'date') and callable(first_date.date):
        first_date = first_date.date()
    if hasattr(last_date, 'date') and callable(last_date.date):
        last_date = last_date.date()

    return {
        'first_date': first_date,
        'last_date': last_date,
        'expected_first': expected_start,
        'expected_last': expected_end,
        'first_match': first_date == expected_start if first_date else False,
        'last_match': last_date == expected_end if last_date else False,
        'row_count': last_row - start_row + 1
    }


def validate_weekend_zero_production(
    wb_path: str,
    sheet_name: str,
    date_col: int,
    production_cols: List[int],
    start_row: int
) -> Dict[str, Any]:
    """Verify production columns are zero on weekends."""
    wb = load_workbook(wb_path)
    ws = wb[sheet_name]

    violations = []
    row = start_row

    while True:
        cell_val = ws.cell(row=row, column=date_col).value
        if not cell_val:
            break

        cell_date = cell_val if isinstance(cell_val, date) else cell_val
        if hasattr(cell_date, 'date') and callable(cell_date.date):
            cell_date = cell_date.date()

        if hasattr(cell_date, 'weekday') and cell_date.weekday() >= 5:
            for col in production_cols:
                prod_val = ws.cell(row=row, column=col).value
                if prod_val and prod_val != 0:
                    violations.append({
                        'row': row,
                        'date': str(cell_date),
                        'column': col,
                        'value': prod_val
                    })
        row += 1

    return {
        'valid': len(violations) == 0,
        'violation_count': len(violations),
        'violations': violations[:10]
    }


def validate_cumulative_formulas(
    wb_path: str,
    sheet_name: str,
    cumul_cols: List[int],
    start_row: int,
    expected_first_formula: str = None,
    expected_subsequent_formula: str = None
) -> Dict[str, Any]:
    """Verify cumulative formulas are correctly structured."""
    wb = load_workbook(wb_path)
    ws = wb[sheet_name]

    results = {}

    for col in cumul_cols:
        first_formula = ws.cell(row=start_row, column=col).value
        second_formula = ws.cell(row=start_row + 1, column=col).value

        results[f'col_{col}'] = {
            'first_row_formula': str(first_formula) if first_formula else None,
            'second_row_formula': str(second_formula) if second_formula else None,
            'is_formula_first': str(first_formula).startswith('=') if first_formula else False,
            'is_formula_second': str(second_formula).startswith('=') if second_formula else False
        }

    return results


def validate_po_quantities(
    wb_path: str,
    sheet_name: str,
    date_col: int,
    po_cols: Dict[int, Dict[date, int]],
    start_row: int
) -> Dict[str, Any]:
    """Verify PO quantities on specific dates.

    po_cols format: {column_num: {date: expected_quantity}}
    """
    wb = load_workbook(wb_path)
    ws = wb[sheet_name]

    results = {'valid': True, 'mismatches': []}

    row = start_row
    date_to_row = {}

    while True:
        cell_val = ws.cell(row=row, column=date_col).value
        if not cell_val:
            break

        cell_date = cell_val if isinstance(cell_val, date) else cell_val
        if hasattr(cell_date, 'date') and callable(cell_date.date):
            cell_date = cell_date.date()

        date_to_row[cell_date] = row
        row += 1

    for col, date_quantities in po_cols.items():
        for exp_date, exp_qty in date_quantities.items():
            if exp_date in date_to_row:
                actual = ws.cell(row=date_to_row[exp_date], column=col).value
                if actual != exp_qty:
                    results['valid'] = False
                    results['mismatches'].append({
                        'date': str(exp_date),
                        'column': col,
                        'expected': exp_qty,
                        'actual': actual
                    })
            else:
                results['valid'] = False
                results['mismatches'].append({
                    'date': str(exp_date),
                    'column': col,
                    'error': 'Date not found in sheet'
                })

    return results


def validate_formula_vs_constant(
    wb_path: str,
    sheet_name: str,
    formula_cols: List[int],
    constant_cols: List[int],
    start_row: int,
    end_row: int
) -> Dict[str, Any]:
    """Verify formula columns contain formulas and constant columns contain values."""
    wb = load_workbook(wb_path)
    ws = wb[sheet_name]

    results = {'formula_errors': [], 'constant_errors': []}

    for row in range(start_row, end_row + 1):
        for col in formula_cols:
            cell = ws.cell(row=row, column=col)
            if cell.data_type != 'f':  # 'f' = formula
                results['formula_errors'].append({
                    'row': row,
                    'col': col,
                    'value': str(cell.value)[:50]
                })

        for col in constant_cols:
            cell = ws.cell(row=row, column=col)
            if cell.data_type == 'f':
                results['constant_errors'].append({
                    'row': row,
                    'col': col,
                    'formula': str(cell.value)[:50]
                })

    results['valid'] = len(results['formula_errors']) == 0 and len(results['constant_errors']) == 0
    return results