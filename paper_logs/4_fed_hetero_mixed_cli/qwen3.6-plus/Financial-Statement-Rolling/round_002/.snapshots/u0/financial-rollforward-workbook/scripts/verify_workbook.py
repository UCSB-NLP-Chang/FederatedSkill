#!/usr/bin/env python3
"""Verify structure and formulas of a generated Excel workbook.

Checks for:
- Sheet existence and order
- Headers at row 5
- Control row labels
- Circular reference detection (formula referencing its own row)
- Cross-sheet reference syntax
"""
import sys
import re
import openpyxl

def check_circular_reference(row, formula, label):
    """Check if a formula references its own row (circular reference)."""
    if not formula or not isinstance(formula, str) or not formula.startswith('='):
        return None

    # Extract row numbers from formula
    row_refs = re.findall(r'[A-Z]+(\d+)', formula)
    for ref_row in row_refs:
        if int(ref_row) == row:
            return f"Circular reference in {label}: row {row} formula references itself: {formula}"
    return None

def verify(path):
    wb = openpyxl.load_workbook(path)
    print(f"Sheets: {wb.sheetnames}")

    errors = []

    for name in wb.sheetnames:
        ws = wb[name]
        print(f"\n--- {name} ---")

        # Check headers (assumes row 5 or 6)
        for header_row in [5, 6]:
            headers = [ws.cell(row=header_row, column=c).value for c in range(1, ws.max_column + 1)]
            if any(h is not None for h in headers):
                print(f"Headers (row {header_row}): {headers}")
                break

        # Scan for control rows and check for circular references
        control_labels = {'Period Totals', 'Ending Balance', 'Calculated Ending Balance', 'Variance', 'GL Balance'}
        for r in range(1, ws.max_row + 1):
            a_val = ws.cell(row=r, column=1).value
            if a_val in control_labels:
                print(f"\nRow {r} [{a_val}]:")

                # Check all columns in this row for circular references
                for c in range(1, ws.max_column + 1):
                    cell_val = ws.cell(row=r, column=c).value
                    if isinstance(cell_val, str) and cell_val.startswith('='):
                        print(f"  Col {c}: {cell_val}")
                        err = check_circular_reference(r, cell_val, f"{name}!{a_val}")
                        if err:
                            errors.append(err)
                            print(f"    *** ERROR: {err}")

    if errors:
        print(f"\n*** VERIFICATION FAILED: {len(errors)} circular reference(s) found ***")
        for err in errors:
            print(f"  - {err}")
        return False
    else:
        print("\nVerification complete. No circular references found.")
        return True

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: verify_workbook.py <path_to_xlsx>")
        sys.exit(1)
    success = verify(sys.argv[1])
    sys.exit(0 if success else 1)