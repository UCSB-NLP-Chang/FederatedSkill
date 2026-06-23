#!/usr/bin/env python3
"""
Reusable openpyxl workbook structural and formula verifier.
Adjust EXPECTED_* constants per task before running.
"""
import sys
import openpyxl

# --- TASK-SPECIFIC CONSTANTS (Override these) ---
EXPECTED_SHEETS = ["Summary", "Assumptions", "Roster", "Calculations"]
EXPECTED_ROWS = {}  # e.g., {"Roster": 103, "Calculations": 107}
EXPECTED_NAMED_RANGES = 0  # Set to expected count
FORMULA_CHECK_SHEET = "Summary"
FORMULA_CHECK_CELLS = []  # e.g., ["C27", "C28", "C29"]
# ------------------------------------------------


def verify(path):
    wb = openpyxl.load_workbook(path, data_only=False)
    errors = []

    # 1. Sheet order
    if wb.sheetnames != EXPECTED_SHEETS:
        errors.append(f"Sheet mismatch. Expected: {EXPECTED_SHEETS}, Got: {wb.sheetnames}")
    else:
        print("OK: Sheet order matches")

    # 2. Row counts
    for sheet_name, count in EXPECTED_ROWS.items():
        if sheet_name in wb.sheetnames:
            actual = wb[sheet_name].max_row
            if actual != count:
                errors.append(f"{sheet_name} row count: {actual} != {count}")
    if EXPECTED_ROWS and not any("row count" in e for e in errors):
        print("OK: Row counts match")

    # 3. Named ranges
    if EXPECTED_NAMED_RANGES > 0:
        dn_count = len(list(wb.defined_names))
        if dn_count != EXPECTED_NAMED_RANGES:
            errors.append(f"Named ranges: {dn_count} != {EXPECTED_NAMED_RANGES}")
        else:
            print(f"OK: Named ranges: {dn_count}")

    # 4. Formula presence check
    if FORMULA_CHECK_SHEET in wb.sheetnames and FORMULA_CHECK_CELLS:
        ws = wb[FORMULA_CHECK_SHEET]
        missing_formulas = []
        for cell_ref in FORMULA_CHECK_CELLS:
            cell = ws[cell_ref]
            if not cell.value or not str(cell.value).startswith("="):
                missing_formulas.append(cell_ref)
        if missing_formulas:
            errors.append(f"Missing formulas in {FORMULA_CHECK_SHEET}: {missing_formulas}")
        else:
            print(f"OK: Formula checks passed in {FORMULA_CHECK_SHEET}")

    # 5. Formula integrity check (backslash artifacts, unquoted sheet refs)
    formula_issues = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    formula = cell.value
                    # Check for Python escape artifacts
                    if '\\' in formula:
                        formula_issues.append(f"{sheet_name}!{cell.coordinate}: Contains backslash: {formula[:60]}...")
                    # Check for unquoted sheet names with spaces/special chars
                    if '!' in formula and any(c in formula for c in [' ', '(', ')']):
                        if "'" not in formula:
                            formula_issues.append(f"{sheet_name}!{cell.coordinate}: Possible unquoted sheet ref: {formula[:60]}...")

    if formula_issues:
        errors.append(f"Formula integrity issues: {len(formula_issues)} found")
        for issue in formula_issues[:5]:
            print(f"  WARN: {issue}")
        if len(formula_issues) > 5:
            print(f"  ... and {len(formula_issues) - 5} more")
    else:
        print("OK: No formula integrity issues")

    if errors:
        print("\nFAIL: Verification errors:")
        for e in errors:
            print(f"  - {e}")
        sys.exit(1)
    else:
        print("\nPASS: Workbook structure valid.")
        sys.exit(0)


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(f"Usage: python3 {sys.argv[0]} <path_to_workbook.xlsx>")
        sys.exit(1)
    verify(sys.argv[1])
