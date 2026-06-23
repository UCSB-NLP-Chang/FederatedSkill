#!/usr/bin/env python3
"""
Template for extracting structured data from images using OCR.
Adapt the parse_date(), parse_price(), and extract_order_id() functions for your specific use case.
"""
import os
import re
from datetime import datetime
from pathlib import Path
from PIL import Image, ImageEnhance, ImageFilter
import pytesseract
from openpyxl import Workbook

def preprocess_image(img):
    """Apply preprocessing to improve OCR accuracy."""
    img = img.convert('L')  # Grayscale
    enhancer = ImageEnhance.Contrast(img)
    img = enhancer.enhance(2.0)  # Increase contrast
    img = img.filter(ImageFilter.SHARPEN)
    return img

def extract_text_multi_strategy(img_path):
    """Extract text using multiple OCR strategies for robustness."""
    img = Image.open(img_path)
    texts = []

    # Strategy 1: Default
    try:
        texts.append(pytesseract.image_to_string(img))
    except Exception:
        pass

    # Strategy 2: Preprocessed
    try:
        processed = preprocess_image(img)
        texts.append(pytesseract.image_to_string(processed))
    except Exception:
        pass

    # Strategy 3: Different PSM modes
    for psm in [6, 4, 11, 12]:
        try:
            texts.append(pytesseract.image_to_string(img, config=f'--psm {psm}'))
        except Exception:
            pass

    return '\n'.join(texts)

def parse_date(text, priority_labels=None):
    """
    Extract date from text with priority labels.

    Args:
        text: OCR text to parse
        priority_labels: List of date labels in priority order (e.g., ['EXP', 'EXPIRY', 'MFG'])

    Returns:
        Date string in ISO format (YYYY-MM-DD) or None
    """
    if priority_labels is None:
        priority_labels = ['EXP', 'EXPIRY']  # Default: prioritize expiry dates

    text_upper = text.upper()
    lines = text_upper.split('\n')

    # Try to find date by priority label
    for label in priority_labels:
        for line in lines:
            if label in line:
                date = extract_date_from_line(line)
                if date:
                    return date

    # Fallback: find any date in text
    return extract_date_from_line(text_upper)

def extract_date_from_line(line):
    """Extract and parse date from a line of text."""
    # DD/MM/YYYY or DD-MM-YYYY
    match = re.search(r'(\d{1,2})[/-](\d{1,2})[/-](\d{4})', line)
    if match:
        d, m, y = int(match.group(1)), int(match.group(2)), int(match.group(3))
        try:
            return datetime(y, m, d).strftime('%Y-%m-%d')
        except ValueError:
            pass

    # MM/YYYY or MM-YYYY (convert to first of month)
    match = re.search(r'(\d{1,2})[/-](\d{4})', line)
    if match:
        m, y = int(match.group(1)), int(match.group(2))
        try:
            return datetime(y, m, 1).strftime('%Y-%m-%d')
        except ValueError:
            pass

    # YYYY-MM-DD
    match = re.search(r'(\d{4})[/-](\d{1,2})[/-](\d{1,2})', line)
    if match:
        y, m, d = int(match.group(1)), int(match.group(2)), int(match.group(3))
        try:
            return datetime(y, m, d).strftime('%Y-%m-%d')
        except ValueError:
            pass

    return None

def parse_price(text):
    """Extract price from text, stripping currency symbols."""
    # Match price patterns with optional currency symbols
    patterns = [
        r'(?:RM|MYR|\$)\s*(\d+\.\d{2})',
        r'(?:RM|MYR|\$)\s*(\d+)',
        r'(\d+\.\d{2})\s*(?:EACH|each)?',
    ]

    for pattern in patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            return float(match.group(1))

    return None

def extract_order_id(text):
    """Extract order ID from OCR text."""
    patterns = [
        r'(ORD[-_]?(?:\d{4}[-_]?\d{5}))',
        r'(SO[-_]?(?:\d{4}[-_]?\d{3}))',
        r'(INV[-_]?\d{8})',
        r'(?:ORDER|INVOICE)[:\s#]*(\d{6,12})',
    ]
    for pattern in patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            return match.group(1)
    return None

def load_reference_set(filepath):
    """Load known valid values from a reference file (one per line)."""
    valid_values = set()
    with open(filepath, 'r') as f:
        for line in f:
            line = line.strip()
            if line and line.lower() != 'order_id':  # Skip header
                valid_values.add(line)
    return valid_values

def process_images_to_excel(img_dir, output_path, sheet_name='data',
                            reference_file=None, track_duplicates=True):
    """
    Process all images in directory and output to Excel.

    Args:
        img_dir: Directory containing images
        output_path: Path for output Excel file
        sheet_name: Name for the Excel sheet
        reference_file: Optional path to reference file for ID validation
        track_duplicates: If True, null fields for duplicate IDs
    """
    # Load reference set if provided
    valid_ids = None
    if reference_file and os.path.exists(reference_file):
        valid_ids = load_reference_set(reference_file)
        print(f"Loaded {len(valid_ids)} reference IDs")

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(['filename', 'order_id', 'date', 'total_amount'])  # Header row

    seen_ids = set()  # Track duplicates
    img_files = sorted(Path(img_dir).glob('*.jpg'))

    for img_path in img_files:
        text = extract_text_multi_strategy(img_path)
        order_id = extract_order_id(text)
        date = parse_date(text)
        price = parse_price(text)

        # Handle duplicates
        if track_duplicates and order_id and order_id in seen_ids:
            print(f"  Duplicate order ID {order_id}, setting all fields to null")
            order_id = None
            date = None
            price = None
        elif order_id:
            seen_ids.add(order_id)

        # Validate against reference if provided
        if valid_ids is not None and order_id and order_id not in valid_ids:
            print(f"  Warning: {order_id} not in reference set")

        ws.append([img_path.name, order_id or '', date or '', price or ''])
        print(f"Processed {img_path.name}: order_id={order_id}, date={date}, price={price}")

    # Remove any trailing empty rows
    while ws.max_row > 1:
        last_row = ws[ws.max_row]
        if all(cell.value is None or cell.value == '' for cell in last_row):
            ws.delete_rows(ws.max_row)
        else:
            break

    wb.save(output_path)
    print(f"Saved {len(img_files)} records to {output_path}")

if __name__ == '__main__':
    # Example usage - adapt paths as needed
    process_images_to_excel(
        '/app/workspace/dataset/img',
        '/app/workspace/output.xlsx',
        reference_file='/app/workspace/dataset/known_orders.csv'
    )