#!/usr/bin/env python3
"""Verify rollforward workbook structure and formulas against reconciliation constraints."""
import sys
import openpyxl

def verify(path):
    try:
        wb = openpyxl.load_workbook(path)
    except Exception as e:
        print(f"FAIL: Cannot load workbook: {e}")
        return False

    errors = []
    if len(wb.sheetnames) < 2:
        errors.append("Expected at least 2 sheets (Summary + 1 Detail)")
        print("\n".join(errors))
        return False

    summary_name = wb.sheetnames[0]
    detail_names = wb.sheetnames[1:]

    for d_name in detail_names:
        ws = wb[d_name]
        # Check control row labels in Col A
        for r, label in [(12, "Month Totals"), (13, "Ending Balance"), (14, "Variance"), (15, "GL Balance")]:
            val = ws.cell(row=r, column=1).value
            if val is None or str(val).strip() == "":
                errors.append(f"[{d_name}] Row {r} Col A is empty (expected {label} label)")
        
        # Check numeric types in data rows (6-11, Cols B-N)
        for r in range(6, 12):
            for c in range(2, 15):
                v = ws.cell(row=r, column=c).value
                if v is not None and not isinstance(v, (int, float)):
                    errors.append(f"[{d_name}] R{r}C{c} is {type(v).__name__}, expected numeric")
        
        # Check Month Totals formulas (Row 12)
        for c in range(2, 15):
            f = ws.cell(row=12, column=c).value
            if not isinstance(f, str) or not f.startswith("=SUM("):
                errors.append(f"[{d_name}] R12C{c} missing SUM formula")
                
        # Check Ending Balance formula (Row 13, Col N)
        f = ws.cell(row=13, column=14).value
        if not isinstance(f, str) or "N12" not in f:
            errors.append(f"[{d_name}] R13C14 (Ending Balance) should reference N12")

    # Check Summary cross-sheet links
    ws_sum = wb[summary_name]
    found_links = False
    for row in ws_sum.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("='"):
                found_links = True
                break
    if not found_links:
        errors.append(f"[{summary_name}] No cross-sheet formulas found")

    if errors:
        print("VERIFICATION FAILED:")
        for e in errors:
            print(f"  - {e}")
        return False
    else:
        print("VERIFICATION PASSED: Structure and formulas look correct.")
        return True

if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("Usage: verify_rollforward.py <workbook.xlsx>")
        sys.exit(1)
    sys.exit(0 if verify(sys.argv[1]) else 1)
