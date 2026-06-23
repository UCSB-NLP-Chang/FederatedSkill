#!/usr/bin/env python3
"""Safe formula injector for openpyxl workbooks.
Populates a rectangular range with a formula template, preserving existing formatting.
"""
import openpyxl
from openpyxl.utils import get_column_letter

def inject_formulas(workbook_path, output_path, sheet_name, start_row, end_row, start_col, end_col, formula_template):
    wb = openpyxl.load_workbook(workbook_path)
    ws = wb[sheet_name]

    for r in range(start_row, end_row + 1):
        for c in range(start_col, end_col + 1):
            col_letter = get_column_letter(c)
            cell = ws[f"{col_letter}{r}"]
            cell.value = formula_template.format(row=r, col=col_letter)

    wb.save(output_path)
    print(f"Saved to {output_path}")

if __name__ == "__main__":
    # Example usage:
    # inject_formulas("template.xlsx", "output.xlsx", "Task", 12, 17, 8, 12,
    #                 "=INDEX(Data!$H$21:$L$38, MATCH($D{row}, Data!$D$21:$D$38, 0), MATCH({col}$10, Data!$H$4:$L$4, 0))")
    pass
