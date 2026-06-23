#!/usr/bin/env python3
"""Reusable openpyxl workbook structural verifier.
Adjust EXPECTED_* constants per task before running.
"""
import sys
import openpyxl

# --- TASK-SPECIFIC CONSTANTS (Override these) ---
EXPECTED_SHEETS = ["Summary", "Assumptions", "Roster", "Calculations"]
EXPECTED_ROWS = {}
EXPECTED_NAMED_RANGES = 0
FORMULA_CHECK_SHEET = "Summary"
FORMULA_CHECK_CELLS = []
# ------------------------------------------------

def verify(path):
    wb = openpyxl.load_workbook(path, data_only=False)
    errors = []

    # 1. Sheet order
    if EXPECTED_SHEETS and wb.sheetnames[:len(EXPECTED_SHEETS)] != EXPECTED_SHEETS:
        errors.append(f"Sheet mismatch. Expected: {EXPECTED_SHEETS}, Got: {wb.sheetnames}")
    else:
        print(f"OK Sheet order: {wb.sheetnames}")

    # 2. Row counts
    for sheet_name, count in EXPECTED_ROWS.items():
        if sheet_name in wb.sheetnames:
            actual = wb[sheet_name].max_row
            if actual != count:
                errors.append(f"{sheet_name} row count: {actual} != {count}")
    if EXPECTED_ROWS and not any("row count" in e for e in errors):
        print("OK Row counts match")

    # 3. Named ranges
    dn_list = wb.defined_names
    dn_count = len(dn_list)
    if EXPECTED_NAMED_RANGES > 0 and dn_count != EXPECTED_NAMED_RANGES:
        errors.append(f"Named ranges: {dn_count} != {EXPECTED_NAMED_RANGES}")
    else:
        print(f"OK Named ranges: {dn_count}")

    # 4. Formula presence check
    if FORMULA_CHECK_SHEET in wb.sheetnames:
        ws = wb[FORMULA_CHECK_SHEET]
        missing_formulas = []
        for cell_ref in FORMULA_CHECK_CELLS:
            cell = ws[cell_ref]
            if not cell.value or not str(cell.value).startswith("="):
                missing_formulas.append(cell_ref)
        if missing_formulas:
            errors.append(f"Missing formulas in {FORMULA_CHECK_SHEET}: {missing_formulas}")
        elif FORMULA_CHECK_CELLS:
            print(f"OK Formula checks passed in {FORMULA_CHECK_SHEET}")

    # 5. Scan for common formula issues
    formula_issues = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    formula = cell.value
                    # Check for backslash artifacts
                    if '\\' in formula:
                        formula_issues.append(f"{sheet_name}!{cell.coordinate}: backslash in formula")
                    # Check for unquoted sheet refs with spaces
                    if '!' in formula and any(c in formula for c in [' ', '(', ')']):
                        if "'" not in formula:
                            formula_issues.append(f"{sheet_name}!{cell.coordinate}: unquoted sheet ref")

    if formula_issues:
        errors.extend(formula_issues[:10])
    else:
        print("OK No formula issues detected")

    if errors:
        print("\nFAILED Verification:")
        for e in errors:
            print(f"  - {e}")
        sys.exit(1)
    else:
        print("\nPASS Verification complete.")

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python verify_xlsx.py <path_to_workbook.xlsx>")
        sys.exit(1)
    verify(sys.argv[1])
