#!/usr/bin/env python3
"""Verify structure and formulas of a generated Excel workbook."""
import sys
import openpyxl

def verify(path):
    wb = openpyxl.load_workbook(path)
    print(f"Sheets: {wb.sheetnames}")
    
    for name in wb.sheetnames:
        ws = wb[name]
        print(f"\n--- {name} ---")
        # Check headers (assumes row 5)
        headers = [ws.cell(row=5, column=c).value for c in range(1, ws.max_column + 1)]
        print(f"Headers: {headers}")
        
        # Scan for control rows and print formulas in column O (15)
        control_labels = {'Period Totals', 'Ending Balance', 'Variance', 'GL Balance'}
        for r in range(1, ws.max_row + 1):
            a_val = ws.cell(row=r, column=1).value
            if a_val in control_labels:
                o_val = ws.cell(row=r, column=15).value
                print(f"Row {r} [{a_val}]: O={o_val}")
                
    print("\nVerification complete.")

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: verify_workbook.py <path_to_xlsx>")
        sys.exit(1)
    verify(sys.argv[1])
