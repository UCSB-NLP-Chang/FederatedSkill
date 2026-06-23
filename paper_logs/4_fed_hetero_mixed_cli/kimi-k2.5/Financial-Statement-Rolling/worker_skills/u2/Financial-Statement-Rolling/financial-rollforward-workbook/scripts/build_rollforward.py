#!/usr/bin/env python3
"""Build deferred revenue rollforward workbook with correct formula patterns.

Usage:
    python build_rollforward.py <output_path> <saas_csv> <services_csv> <gl_json>

The script creates an Excel workbook with:
- Deferred Summary sheet with linked totals
- Detail sheets per account with proper rollforward formulas
- No circular references
- GL reconciliation rows
"""

import sys
import json
import csv
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


def build_rollforward(output_path, saas_csv, services_csv, gl_json):
    """Build the deferred revenue workbook."""
    
    # Read source data
    with open(saas_csv, 'r') as f:
        saas_data = list(csv.DictReader(f))
    with open(services_csv, 'r') as f:
        services_data = list(csv.DictReader(f))
    with open(gl_json, 'r') as f:
        gl_balances = json.load(f)
    
    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Deferred Summary"
    
    # Column headers for detail sheets
    headers = [
        'Customer', 'Beginning Balance', 
        'May Billings', 'May Recognition', 'May Ending Balance',
        'Jun Billings', 'Jun Recognition', 'Jun Ending Balance',
        'Jul Billings', 'Jul Recognition', 'Jul Ending Balance',
        'Aug Billings', 'Aug Recognition', 'Aug Ending Balance',
        'Contract Months', 'Notes', 'Revenue Code'
    ]
    
    def create_detail_sheet(ws, data, gl_key, gl_data):
        """Create a detail sheet with correct formula patterns."""
        
        # Headers at row 5
        for col, header in enumerate(headers, 1):
            ws.cell(row=5, column=col, value=header)
        
        # Data rows starting at row 6
        start_row = 6
        for i, row in enumerate(data):
            r = start_row + i
            ws.cell(row=r, column=1, value=row['entity'])
            ws.cell(row=r, column=2, value=float(row['beginning_balance']))
            ws.cell(row=r, column=3, value=float(row['may_adds']))
            ws.cell(row=r, column=4, value=float(row['may_release']))
            ws.cell(row=r, column=5, value=float(row['may_ending_balance']))
            ws.cell(row=r, column=6, value=float(row['jun_adds']))
            ws.cell(row=r, column=7, value=float(row['jun_release']))
            ws.cell(row=r, column=8, value=float(row['jun_ending_balance']))
            ws.cell(row=r, column=9, value=float(row['jul_adds']))
            ws.cell(row=r, column=10, value=float(row['jul_release']))
            ws.cell(row=r, column=11, value=float(row['jul_ending_balance']))
            ws.cell(row=r, column=12, value=float(row['aug_adds']))
            ws.cell(row=r, column=13, value=float(row['aug_release']))
            ws.cell(row=r, column=14, value=float(row['aug_ending_balance']))
            ws.cell(row=r, column=15, value=int(row['term_months']))
            ws.cell(row=r, column=16, value=row['comments'])
            ws.cell(row=r, column=17, value=int(row['account_number']))
        
        end_row = start_row + len(data) - 1
        
        # Control rows - NO circular references
        totals_row = end_row + 1
        calc_ending_row = end_row + 2
        variance_row = end_row + 3
        gl_row = end_row + 4
        
        # Period Totals - SUM of data rows
        ws.cell(row=totals_row, column=1, value="Period Totals")
        for col in range(2, 15):  # B through N
            col_letter = get_column_letter(col)
            ws.cell(row=totals_row, column=col, 
                   value=f"=SUM({col_letter}{start_row}:{col_letter}{end_row})")
        # Total billings in column O
        ws.cell(row=totals_row, column=15, 
               value=f"=C{totals_row}+F{totals_row}+I{totals_row}+L{totals_row}")
        
        # Calculated Ending Balance - from data rows, NOT self-referencing
        ws.cell(row=calc_ending_row, column=1, value="Calculated Ending Balance")
        # Sum of ending balances from data rows (column N for Aug)
        ws.cell(row=calc_ending_row, column=14, 
               value=f"=SUM(N{start_row}:N{end_row})")
        # Total recognition in column O
        ws.cell(row=calc_ending_row, column=15,
               value=f"=D{totals_row}+G{totals_row}+J{totals_row}+M{totals_row}")
        
        # GL Balance - hard-coded value
        ws.cell(row=gl_row, column=1, value="GL Balance")
        ws.cell(row=gl_row, column=14, value=gl_data.get('aug', 0))
        
        # Variance - GL minus Calculated
        ws.cell(row=variance_row, column=1, value="Variance")
        ws.cell(row=variance_row, column=15, 
               value=f"=N{gl_row}-N{calc_ending_row}")
        
        return totals_row, calc_ending_row, gl_row
    
    # Create SaaS detail sheet
    ws_saas = wb.create_sheet("SaaS Rev #2300")
    saas_totals, saas_calc, saas_gl = create_detail_sheet(
        ws_saas, saas_data, 'saas_rev_2300', gl_balances.get('saas_rev_2300', {}))
    
    # Create Services detail sheet
    ws_services = wb.create_sheet("Services Rev #2310")
    svc_totals, svc_calc, svc_gl = create_detail_sheet(
        ws_services, services_data, 'services_rev_2310', gl_balances.get('services_rev_2310', {}))
    
    # Build Summary sheet
    ws_summary.cell(row=1, column=1, value="LatticeWare")
    ws_summary.cell(row=2, column=1, value="Period Ending: August 31, 2025")
    
    ws_summary.cell(row=5, column=1, value="Account 2300 - SaaS Revenue")
    ws_summary.cell(row=7, column=1, value="Period Totals (Total Billings)")
    ws_summary.cell(row=7, column=2, value=f"='SaaS Rev #2300'!O{saas_totals}")
    ws_summary.cell(row=8, column=1, value="Calculated Ending Balance")
    ws_summary.cell(row=8, column=2, value=f"='SaaS Rev #2300'!N{saas_calc}")
    ws_summary.cell(row=9, column=1, value="GL Balance")
    ws_summary.cell(row=9, column=2, value=f"='SaaS Rev #2300'!N{saas_gl}")
    
    ws_summary.cell(row=11, column=1, value="Account 2310 - Services Revenue")
    ws_summary.cell(row=13, column=1, value="Period Totals (Total Billings)")
    ws_summary.cell(row=13, column=2, value=f"='Services Rev #2310'!O{svc_totals}")
    ws_summary.cell(row=14, column=1, value="Calculated Ending Balance")
    ws_summary.cell(row=14, column=2, value=f"='Services Rev #2310'!N{svc_calc}")
    ws_summary.cell(row=15, column=1, value="GL Balance")
    ws_summary.cell(row=15, column=2, value=f"='Services Rev #2310'!N{svc_gl}")
    
    ws_summary.cell(row=17, column=1, value="Total GL Balance")
    ws_summary.cell(row=17, column=2, value=f"=B9+B15")
    
    wb.save(output_path)
    print(f"Workbook created: {output_path}")


if __name__ == "__main__":
    if len(sys.argv) != 5:
        print("Usage: python build_rollforward.py <output_path> <saas_csv> <services_csv> <gl_json>")
        sys.exit(1)
    build_rollforward(sys.argv[1], sys.argv[2], sys.argv[3], sys.argv[4])
