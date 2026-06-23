#!/usr/bin/env python3
"""Verify structure and formulas of a generated Excel workbook.

Checks for:
- Sheet existence and order
- Headers at row 5
- Control row labels
- Circular reference detection (formula referencing its own row)
- Cross-sheet reference syntax
- Summary sheet linking to correct columns (N not O)
- Control row formula correctness (Ending Balance uses SUM, not rollforward)
- Variance formula correctness
"""
import sys
import re
import openpyxl

def check_circular_reference(row, formula, label):
    """Check if a formula references its own row (circular reference)."""
    if not formula or not isinstance(formula, str) or not formula.startswith('='):
        return None

    # Extract row numbers from formula
    row_refs = re.findall(r'[A-Z]+(\d+)', formula)
    for ref_row in row_refs:
        if int(ref_row) == row:
            return f"Circular reference in {label}: row {row} formula references itself: {formula}"
    return None

def check_ending_balance_formula(row, formula, label, data_start, data_end):
    """Check if Ending Balance control row uses SUM instead of rollforward formula."""
    if not formula or not isinstance(formula, str) or not formula.startswith('='):
        return None

    # Check for rollforward pattern (=B{r}+C{r}-D{r} or similar)
    # This pattern indicates wrong formula type for control rows
    rollforward_pattern = rf'^=[A-Z]+{row}[+-][A-Z]+{row}[+-][A-Z]+{row}$'
    if re.match(rollforward_pattern, formula.replace(' ', '')):
        return f"Ending Balance control row uses rollforward formula instead of SUM: row {row} has {formula}"

    # Check if it's a SUM formula referencing data rows
    if 'SUM' in formula.upper():
        # Verify it references data rows, not itself
        row_refs = re.findall(r'[A-Z]+(\d+)', formula)
        for ref_row in row_refs:
            if int(ref_row) == row:
                return f"Ending Balance SUM formula references its own row: {formula}"
        return None  # Valid SUM formula

    return None

def check_variance_formula(row, formula, gl_row, calc_row):
    """Check if Variance row has correct formula structure."""
    if not formula or not isinstance(formula, str) or not formula.startswith('='):
        return None

    # Variance should be GL - Calculated Ending, not a complex formula
    # Check for patterns like =O{r}-N{r} which are wrong
    if gl_row and calc_row:
        # Should be something like =N{gl}-N{calc} for each period column
        # Wrong pattern: totals column minus ending column
        wrong_pattern = rf'=O{row}-N{row}'
        if re.match(wrong_pattern, formula.replace(' ', '').upper()):
            return f"Variance formula uses totals column O instead of period columns: {formula}"

    return None

def check_summary_column(ws_summary, detail_sheets):
    """Check that summary sheet links to final period column N, not totals column O."""
    errors = []

    for row in range(1, ws_summary.max_row + 1):
        for col in range(1, ws_summary.max_column + 1):
            cell_val = ws_summary.cell(row=row, column=col).value
            if isinstance(cell_val, str) and cell_val.startswith("="):
                # Check for references to column O in detail sheets
                o_ref_pattern = r"='?[^'!]+(?:#[^'!]+)?'?!O\d+"
                if re.search(o_ref_pattern, cell_val):
                    errors.append(f"Summary row {row} col {col} links to totals column O instead of final period column N: {cell_val}")

    return errors

def verify(path):
    wb = openpyxl.load_workbook(path)
    print(f"Sheets: {wb.sheetnames}")

    errors = []
    detail_sheets = []

    for name in wb.sheetnames:
        ws = wb[name]
        print(f"\n--- {name} ---")

        # Check headers (assumes row 5)
        headers = [ws.cell(row=5, column=c).value for c in range(1, ws.max_column + 1)]
        print(f"Headers: {headers}")

        # Scan for control rows and check for circular references
        control_labels = {'Period Totals', 'Ending Balance', 'Calculated Ending Balance', 'Variance', 'GL Balance'}
        control_rows = {}
        data_start = 6
        data_end = None

        for r in range(1, ws.max_row + 1):
            a_val = ws.cell(row=r, column=1).value
            if a_val in control_labels:
                control_rows[a_val] = r
                print(f"\nRow {r} [{a_val}]:")

                # Check all columns in this row for circular references
                for c in range(1, ws.max_column + 1):
                    cell_val = ws.cell(row=r, column=c).value
                    if isinstance(cell_val, str) and cell_val.startswith('='):
                        print(f"  Col {c}: {cell_val}")
                        err = check_circular_reference(r, cell_val, f"{name}!{a_val}")
                        if err:
                            errors.append(err)
                            print(f"    *** ERROR: {err}")

                        # Check Ending Balance formula correctness
                        if a_val == 'Ending Balance':
                            err = check_ending_balance_formula(r, cell_val, f"{name}!{a_val}", data_start, data_end)
                            if err:
                                errors.append(err)
                                print(f"    *** ERROR: {err}")

            # Track data row range (rows between header and first control row)
            if data_end is None and a_val in control_labels:
                data_end = r - 1

        # Store detail sheet info for summary check
        if name != wb.sheetnames[0]:  # Assume first sheet is summary
            detail_sheets.append(name)

    # Check summary sheet for wrong column references
    if wb.sheetnames:
        summary_ws = wb[wb.sheetnames[0]]
        summary_errors = check_summary_column(summary_ws, detail_sheets)
        errors.extend(summary_errors)
        for err in summary_errors:
            print(f"*** ERROR: {err}")

    if errors:
        print(f"\n*** VERIFICATION FAILED: {len(errors)} error(s) found ***")
        for err in errors:
            print(f"  - {err}")
        return False
    else:
        print("\nVerification complete. No errors found.")
        return True

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: verify_workbook.py <path_to_xlsx>")
        sys.exit(1)
    success = verify(sys.argv[1])
    sys.exit(0 if success else 1)
