#!/usr/bin/env python3
"""
Template for MetroLink-style transit subsidy reconciliation workbook.
Assumes 4 months (Jan-Apr) in columns B-E, Total Amortization in F.
Control rows at 12-15. Summary cross-sheet formulas reference correct columns.
"""

import openpyxl
from openpyxl.utils import get_column_letter
import json
import csv

# --- Configuration ---
INPUT_JSON = 'gl_balances.json'
INPUT_CSV_BUS = 'bus_vendors.csv'
INPUT_CSV_RAIL = 'rail_vendors.csv'
OUTPUT_FILE = 'reconciliation.xlsx'

# --- Setup ---
wb = openpyxl.Workbook()

# Remove default sheet, will create named ones
wb.remove(wb.active)

# --- Helper Functions ---
def col_letter(idx):
    """Get column letter from 1-based index (1=A, 2=B)."""
    return get_column_letter(idx)

def create_detail_sheet(wb, sheet_name, vendors_csv, gl_balance):
    """Create a detail sheet with vendors and control rows."""
    ws = wb.create_sheet(title=sheet_name)

    # Headers (Row 1)
    # Columns: A=Vendor, B=Jan, C=Feb, D=Mar, E=Apr, F=Total Amortization
    headers = ['Partner / Vendor', 'Jan', 'Feb', 'Mar', 'Apr', 'Total Amortization']
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=header)

    # Data rows (6-11)
    vendors = []
    with open(vendors_csv, 'r') as f:
        reader = csv.DictReader(f)
        for row in reader:
            vendors.append(row)

    for idx, vendor in enumerate(vendors, start=6):
        # Write vendor name
        ws.cell(row=idx, column=1, value=vendor['vendor'])
        # Write monthly amortization as floats (raw, no rounding)
        ws.cell(row=idx, column=2, value=float(vendor['jan_amortization']))
        ws.cell(row=idx, column=3, value=float(vendor['feb_amortization']))
        ws.cell(row=idx, column=4, value=float(vendor['mar_amortization']))
        ws.cell(row=idx, column=5, value=float(vendor['apr_amortization']))
        # Total amortization (sum of months) - raw float, no rounding
        total = float(vendor['jan_amortization']) + float(vendor['feb_amortization']) + \
                float(vendor['mar_amortization']) + float(vendor['apr_amortization'])
        ws.cell(row=idx, column=6, value=total)

    # Control Rows
    # Row 12: Month Totals (formulas summing rows 6-11 for each month column)
    for col_idx in range(2, 7):  # B through F
        col = col_letter(col_idx)
        ws.cell(row=12, column=col_idx, value=f"=SUM({col}6:{col}11)")

    # Row 13: Ending Balance (hardcoded April GL balance in Apr column and Total column)
    ws.cell(row=13, column=5, value=float(gl_balance))  # Apr column (E)
    ws.cell(row=13, column=6, value=float(gl_balance))  # Total column (F)

    # Row 14: Variance (0 or formula = E13 - E15)
    ws.cell(row=14, column=5, value=0)
    ws.cell(row=14, column=6, value=0)

    # Row 15: GL Balance (hardcoded)
    ws.cell(row=15, column=5, value=float(gl_balance))
    ws.cell(row=15, column=6, value=float(gl_balance))

    return ws

def create_summary_sheet(wb, sheet_names, gl_balances):
    """Create summary sheet with cross-sheet formulas."""
    ws = wb.create_sheet(title='Transit Summary', index=0)

    # Title rows
    ws.cell(row=1, column=1, value='MetroLink Transit Authority')
    ws.cell(row=2, column=1, value='Pass Liability Reconciliation')
    ws.cell(row=3, column=1, value='All amounts in USD')

    # Headers (Row 5)
    ws.cell(row=5, column=1, value='Pool / Program')
    ws.cell(row=5, column=2, value='Amount')

    # Data rows - cross-sheet references
    # April is column E (5th column), Month Totals is row 12
    # Use dynamic column calculation
    apr_col = col_letter(5)  # E
    total_col = col_letter(6)  # F

    for idx, (sheet_name, gl_bal) in enumerate(zip(sheet_names, gl_balances), start=6):
        # Reference the April column (E) in the Month Totals row (12)
        formula = f"='{sheet_name}'!{apr_col}12"
        ws.cell(row=idx, column=1, value=sheet_name)
        ws.cell(row=idx, column=2, value=formula)

    # Row 8: Month Totals
    ws.cell(row=8, column=1, value='Month Totals')
    ws.cell(row=8, column=2, value='=SUM(B6:B7)')

    # Row 9: Ending Balance (references the Amount column B, row 8)
    ws.cell(row=9, column=1, value='Ending Balance')
    ws.cell(row=9, column=2, value='=B8')

    # Row 12-13: Total Amortization links
    ws.cell(row=12, column=1, value='Bus Total Amortization')
    ws.cell(row=12, column=2, value=f"='{sheet_names[0]}'!{total_col}12")

    ws.cell(row=13, column=1, value='Rail Total Amortization')
    ws.cell(row=13, column=2, value=f"='{sheet_names[1]}'!{total_col}12")

    # Row 14: Total Amortization
    ws.cell(row=14, column=1, value='Total Amortization')
    ws.cell(row=14, column=2, value='=SUM(B12:B13)')

    # Row 16: GL Balance
    ws.cell(row=16, column=1, value='GL Balance')
    ws.cell(row=16, column=2, value='=B9+B14')

    return ws

# --- Main Execution ---
if __name__ == '__main__':
    # Load GL balances
    with open(INPUT_JSON, 'r') as f:
        gl_data = json.load(f)

    bus_gl = gl_data['bus_program_4310']['apr']
    rail_gl = gl_data['rail_program_4320']['apr']

    # Create detail sheets
    create_detail_sheet(wb, 'Bus Program #4310', INPUT_CSV_BUS, bus_gl)
    create_detail_sheet(wb, 'Rail Program #4320', INPUT_CSV_RAIL, rail_gl)

    # Create summary (moves to index 0)
    create_summary_sheet(wb, ['Bus Program #4310', 'Rail Program #4320'], [bus_gl, rail_gl])

    # Verify before save
    print("Sheet names:", wb.sheetnames)
    print("\nVerifying formulas...")

    # Check Bus sheet row 12 formulas
    bus_ws = wb['Bus Program #4310']
    print(f"Bus B12: {bus_ws['B12'].value}")  # Should be =SUM(B6:B11)
    print(f"Bus E13 (Ending): {bus_ws['E13'].value}")  # Should be float

    # Check Summary formulas
    sum_ws = wb['Transit Summary']
    print(f"Summary B6: {sum_ws['B6'].value}")  # Should be ='Bus Program #4310'!E12
    print(f"Summary B9: {sum_ws['B9'].value}")  # Should be =B8

    # Save
    wb.save(OUTPUT_FILE)
    print(f"\nSaved to {OUTPUT_FILE}")
