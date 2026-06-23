#!/usr/bin/env python3
"""Daily multi-scenario production recovery simulation (B2)."""

import argparse
import json
import openpyxl
from datetime import date, timedelta, datetime
from typing import Dict, List, Any, Tuple


def to_date(val):
    """Convert openpyxl date value to date object."""
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    return None


def build_calendar(start: date, end: date, holidays: List[date]) -> List[date]:
    """Generate working days excluding weekends and holidays."""
    working_days = []
    current = start
    while current <= end:
        if current.weekday() < 5 and current not in holidays:
            working_days.append(current)
        current += timedelta(days=1)
    return working_days


def distribute_exact(total: int, valid_days: List[date], capacity_func: callable) -> Dict[date, int]:
    """Distribute total across valid days with remainder front-loaded."""
    if not valid_days:
        return {}

    units_per_day, remainder = divmod(total, len(valid_days))

    distribution = {}
    for i, day in enumerate(valid_days):
        value = units_per_day + (1 if i < remainder else 0)
        cap = capacity_func(day)
        distribution[day] = min(value, cap)

    return distribution


def capacity_standard(day: date, transition_date: date, before: int, after: int) -> int:
    """Standard tiered capacity based on date."""
    if day >= transition_date:
        return after
    return before


def capacity_with_shift(day: date, shift_days: List[date], shift_cap: int,
                        transition_date: date, before: int, after: int) -> int:
    """Capacity with shift window override."""
    if day in shift_days:
        return shift_cap
    return capacity_standard(day, transition_date, before, after)


def create_scenario_sheet(wb: openpyxl.Workbook, scenario_name: str,
                          constraints: Dict, distribution: Dict[str, Dict[date, int]]) -> None:
    """Create one scenario sheet with data and formulas."""
    ws = wb.create_sheet(title=scenario_name)

    working_days = constraints['working_days']
    date_to_row = {d: i + 4 for i, d in enumerate(working_days)}

    # Headers
    ws['B3'] = 'Date'
    ws['C3'] = 'Web Planned Production'
    ws['D3'] = 'Web PO Due'
    ws['E3'] = 'Web Cumulative Open'
    ws['F3'] = 'DB Planned Production'
    ws['G3'] = 'DB PO Due'
    ws['H3'] = 'DB Cumulative Open'
    ws['I3'] = 'Network Production'
    ws['J3'] = 'Total Production'

    # Dates: first row literal, subsequent formula
    first_day = working_days[0]
    ws['B4'] = first_day  # datetime.date literal

    for row in range(5, 4 + len(working_days)):
        ws[f'B{row}'] = f'=B{row-1}+1'

    # Production values (constants)
    for day, row in date_to_row.items():
        web_prod = distribution['Web'].get(day, 0)
        db_prod = distribution['DB'].get(day, 0)
        net_prod = distribution['Network'].get(day, 0)

        ws.cell(row=row, column=3, value=web_prod)
        ws.cell(row=row, column=6, value=db_prod)
        ws.cell(row=row, column=9, value=net_prod)

        # PO due on specific dates (constants)
        if day in constraints.get('web_po_dates', {}):
            ws.cell(row=row, column=4, value=constraints['web_po_dates'][day])
        if day in constraints.get('db_po_dates', {}):
            ws.cell(row=row, column=7, value=constraints['db_po_dates'][day])

    # Cumulative formulas
    ws['E4'] = '=D4-C4'
    ws['H4'] = '=G4-F4'
    ws['J4'] = '=C4+F4+I4'

    for row in range(5, 4 + len(working_days)):
        ws[f'E{row}'] = f'=E{row-1}+D{row}-C{row}'
        ws[f'H{row}'] = f'=H{row-1}+G{row}-F{row}'
        ws[f'J{row}'] = f'=C{row}+F{row}+I{row}'

    # Weekend/holiday = 0
    for day, row in date_to_row.items():
        if day.weekday() >= 5 or day in constraints.get('holidays', []):
            ws.cell(row=row, column=3, value=0)
            ws.cell(row=row, column=6, value=0)
            ws.cell(row=row, column=9, value=0)


def generate_summary_md(output_path: str, scenarios: Dict[str, Dict], constraints: Dict) -> None:
    """Generate summary.md with bold labels."""
    lines = [
        "# Production Recovery Summary",
        "",
        "## Scenario Analysis",
        ""
    ]

    for scenario_name, results in scenarios.items():
        final_cum_web = results.get('final_cum_web', 0)
        final_cum_db = results.get('final_cum_db', 0)

        on_time_web = final_cum_web <= 0
        on_time_db = final_cum_db <= 0

        lines.append(f"### {scenario_name}")
        lines.append(f"- **Web Final Cumulative**: {final_cum_web}")
        lines.append(f"- **DB Final Cumulative**: {final_cum_db}")
        lines.append(f"- **Web On-Time**: {'Yes' if on_time_web else 'No'}")
        lines.append(f"- **DB On-Time**: {'Yes' if on_time_db else 'No'}")
        lines.append("")

    with open(output_path, 'w') as f:
        f.write('\n'.join(lines))


def main():
    parser = argparse.ArgumentParser(description='Daily production recovery simulation')
    parser.add_argument('--constraints', required=True, help='Constraints JSON file')
    parser.add_argument('--output-xlsx', required=True, help='Output Excel file')
    parser.add_argument('--output-summary', required=True, help='Output summary.md file')

    args = parser.parse_args()

    with open(args.constraints) as f:
        constraints = json.load(f)

    # Parse dates from constraints
    start_date = date.fromisoformat(constraints['start_date'])
    end_date = date.fromisoformat(constraints['end_date'])
    holidays = [date.fromisoformat(h) for h in constraints.get('holidays', [])]
    transition_date = date.fromisoformat(constraints.get('transition_date', start_date))

    # Build calendar
    working_days = build_calendar(start_date, end_date, holidays)
    constraints['working_days'] = working_days

    # Parse PO dates
    constraints['web_po_dates'] = {date.fromisoformat(d): q for d, q in constraints.get('web_po', {}).items()}
    constraints['db_po_dates'] = {date.fromisoformat(d): q for d, q in constraints.get('db_po', {}).items()}

    # Validate totals
    total_web_po = sum(constraints['web_po_dates'].values())
    total_db_po = sum(constraints['db_po_dates'].values())
    expected_web = constraints.get('web_total', total_web_po)
    expected_db = constraints.get('db_total', total_db_po)

    if total_web_po != expected_web:
        print(f"Warning: Web PO total {total_web_po} != expected {expected_web}")
    if total_db_po != expected_db:
        print(f"Warning: DB PO total {total_db_po} != expected {expected_db}")

    # Create workbook
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Generate scenarios
    scenario_names = constraints.get('scenarios', ['Scenario 1'])

    for scenario_name in scenario_names:
        # Calculate distributions based on scenario parameters
        web_start = date.fromisoformat(constraints.get('web_start', constraints['start_date']))
        db_start = date.fromisoformat(constraints.get('db_start', start_date))

        web_days = [d for d in working_days if d >= web_start]
        db_days = [d for d in working_days if d >= db_start]

        cap_before = constraints.get('capacity_before', 120)
        cap_after = constraints.get('capacity_after', 135)

        cap_func = lambda d: capacity_standard(d, transition_date, cap_before, cap_after)

        distribution = {
            'Web': distribute_exact(expected_web, web_days, cap_func),
            'DB': distribute_exact(expected_db, db_days, cap_func),
            'Network': {}  # Scenario-dependent
        }

        create_scenario_sheet(wb, scenario_name, constraints, distribution)

    wb.save(args.output_xlsx)

    # Generate summary
    generate_summary_md(args.output_summary, {}, constraints)

    print(f"Workbook created: {args.output_xlsx}")
    print(f"Working days: {len(working_days)}")


if __name__ == '__main__':
    main()