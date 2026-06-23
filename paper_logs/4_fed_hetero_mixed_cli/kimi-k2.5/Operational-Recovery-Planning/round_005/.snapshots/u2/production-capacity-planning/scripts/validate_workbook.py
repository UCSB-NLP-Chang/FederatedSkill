#!/usr/bin/env python3
"""Reusable validation helpers for Excel workbook verification."""

from openpyxl import load_workbook
from datetime import datetime, date
from typing import List, Dict, Any, Optional, Callable


def to_date(val):
    """Convert openpyxl date value to date object."""
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    return None


def validate_sheet_names(wb_path: str, expected_sheets: List[str]) -> Dict[str, bool]:
    """Verify all expected sheets exist with exact names."""
    wb = load_workbook(wb_path)
    results = {}
    for sheet in expected_sheets:
        results[sheet] = sheet in wb.sheetnames
    return results


def validate_date_range(
    wb_path: str,
    sheet_name: str,
    date_col: int,
    start_row: int,
    expected_start: date,
    expected_end: date
) -> Dict[str, Any]:
    """Verify date range in a sheet covers expected span."""
    wb = load_workbook(wb_path)
    ws = wb[sheet_name]

    first_date = to_date(ws.cell(row=start_row, column=date_col).value)
    last_row = start_row
    while ws.cell(row=last_row + 1, column=date_col).value:
        last_row += 1
    last_date = to_date(ws.cell(row=last_row, column=date_col).value)

    return {
        'first_date': first_date,
        'last_date': last_date,
        'expected_first': expected_start,
        'expected_last': expected_end,
        'first_match': first_date == expected_start if first_date else False,
        'last_match': last_date == expected_end if last_date else False,
        'row_count': last_row - start_row + 1
    }


def validate_weekend_zero_production(
    wb_path: str,
    sheet_name: str,
    date_col: int,
    production_cols: List[int],
    start_row: int,
    holidays: List[date] = None
) -> Dict[str, Any]:
    """Verify production columns are zero on weekends and holidays."""
    wb = load_workbook(wb_path)
    ws = wb[sheet_name]

    if holidays is None:
        holidays = []

    violations = []
    row = start_row

    while True:
        cell_val = ws.cell(row=row, column=date_col).value
        if not cell_val:
            break

        cell_date = to_date(cell_val)
        if cell_date and (cell_date.weekday() >= 5 or cell_date in holidays):
            for col in production_cols:
                prod_val = ws.cell(row=row, column=col).value
                if prod_val and prod_val != 0:
                    violations.append({
                        'row': row,
                        'date': str(cell_date),
                        'column': col,
                        'value': prod_val
                    })
        row += 1

    return {
        'valid': len(violations) == 0,
        'violation_count': len(violations),
        'violations': violations[:10]
    }


def validate_cumulative_formulas(
    wb_path: str,
    sheet_name: str,
    cumul_cols: List[int],
    start_row: int
) -> Dict[str, Any]:
    """Verify cumulative formulas are correctly structured."""
    wb = load_workbook(wb_path)
    ws = wb[sheet_name]

    results = {}

    for col in cumul_cols:
        first_formula = ws.cell(row=start_row, column=col).value
        second_formula = ws.cell(row=start_row + 1, column=col).value

        results[f'col_{col}'] = {
            'first_row_formula': str(first_formula) if first_formula else None,
            'second_row_formula': str(second_formula) if second_formula else None,
            'is_formula_first': str(first_formula).startswith('=') if first_formula else False,
            'is_formula_second': str(second_formula).startswith('=') if second_formula else False
        }

    return results


def validate_po_quantities(
    wb_path: str,
    sheet_name: str,
    date_col: int,
    po_cols: Dict[int, Dict[date, int]],
    start_row: int
) -> Dict[str, Any]:
    """Verify PO quantities on specific dates.

    po_cols format: {column_num: {date: expected_quantity}}
    """
    wb = load_workbook(wb_path)
    ws = wb[sheet_name]

    results = {'valid': True, 'mismatches': []}

    row = start_row
    date_to_row = {}

    # Build date to row mapping
    while True:
        cell_val = ws.cell(row=row, column=date_col).value
        if not cell_val:
            break
        cell_date = to_date(cell_val)
        if cell_date:
            date_to_row[cell_date] = row
        row += 1

    # Check each expected PO
    for col, date_quantities in po_cols.items():
        for exp_date, exp_qty in date_quantities.items():
            if exp_date in date_to_row:
                actual = ws.cell(row=date_to_row[exp_date], column=col).value
                if actual != exp_qty:
                    results['valid'] = False
                    results['mismatches'].append({
                        'date': str(exp_date),
                        'column': col,
                        'expected': exp_qty,
                        'actual': actual
                    })
            else:
                results['valid'] = False
                results['mismatches'].append({
                    'date': str(exp_date),
                    'column': col,
                    'error': 'Date not found in sheet'
                })

    return results


def validate_exact_totals(
    wb_path: str,
    sheet_name: str,
    production_cols: List[int],
    start_row: int,
    expected_totals: Dict[int, int]
) -> Dict[str, Any]:
    """Verify exact totals for production columns."""
    wb = load_workbook(wb_path)
    ws = wb[sheet_name]

    results = {'valid': True, 'mismatches': []}

    # Find last row
    last_row = start_row
    while ws.cell(row=last_row + 1, column=2).value:
        last_row += 1

    for col, expected in expected_totals.items():
        actual = sum(ws.cell(row=r, column=col).value or 0 for r in range(start_row, last_row + 1))
        if actual != expected:
            results['valid'] = False
            results['mismatches'].append({
                'column': col,
                'expected': expected,
                'actual': actual,
                'difference': actual - expected
            })

    return results


def run_full_validation(wb_path: str, constraints: Dict[str, Any]) -> Dict[str, Any]:
    """Run comprehensive validation against all constraints."""
    results = {}

    # Sheet names
    results['sheets'] = validate_sheet_names(wb_path, constraints.get('sheets', []))

    # Date range (per sheet)
    for sheet in constraints.get('sheets', []):
        if sheet in results['sheets'] and results['sheets'][sheet]:
            results[f'{sheet}_date_range'] = validate_date_range(
                wb_path, sheet,
                constraints['date_col'],
                constraints['start_row'],
                constraints['date_range'][0],
                constraints['date_range'][1]
            )

    # Weekend/holiday zero production
    for sheet in constraints.get('sheets', []):
        if sheet in results['sheets'] and results['sheets'][sheet]:
            results[f'{sheet}_weekend'] = validate_weekend_zero_production(
                wb_path, sheet,
                constraints['date_col'],
                constraints['production_cols'],
                constraints['start_row'],
                constraints.get('holidays', [])
            )

    # Exact totals
    for sheet in constraints.get('sheets', []):
        if sheet in results['sheets'] and results['sheets'][sheet]:
            results[f'{sheet}_totals'] = validate_exact_totals(
                wb_path, sheet,
                constraints['production_cols'],
                constraints['start_row'],
                constraints.get('expected_totals', {})
            )

    return results


if __name__ == '__main__':
    import argparse
    import json

    parser = argparse.ArgumentParser(description='Validate Excel workbook')
    parser.add_argument('--wb', required=True, help='Workbook path')
    parser.add_argument('--constraints', required=True, help='Constraints JSON file')
    parser.add_argument('--output', default=None, help='Output results JSON')

    args = parser.parse_args()

    with open(args.constraints) as f:
        constraints = json.load(f)

    results = run_full_validation(args.wb, constraints)

    if args.output:
        with open(args.output, 'w') as f:
            json.dump(results, f, indent=2)
    else:
        print(json.dumps(results, indent=2))