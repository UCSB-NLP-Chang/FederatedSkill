#!/usr/bin/env python3
"""Weekly catch-up simulation with step-down policy for queue recovery (B1)."""

import argparse
import json
import openpyxl
from datetime import date, datetime
from typing import Dict, List, Any, Tuple


def to_date(val):
    """Convert openpyxl date value to date object."""
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    return None


def find_demand_row(ws, labels=['demand', 'forecast', 'weekly']):
    """Find demand row by label, NOT by position."""
    for row in range(1, ws.max_row + 1):
        cell_val = ws.cell(row=row, column=1).value
        if cell_val:
            cell_str = str(cell_val).lower()
            if any(label in cell_str for label in labels):
                return row
    return None


def find_week_columns(ws, header_row, demand_row):
    """Find column-to-week mapping from header row."""
    week_cols = {}
    for col in range(2, ws.max_column + 1):
        header_val = ws.cell(row=header_row, column=col).value
        # Skip "Total" columns
        if header_val and 'total' in str(header_val).lower():
            continue
        if isinstance(header_val, (int, float)):
            week_num = int(header_val)
            week_cols[week_num] = col
    return week_cols


def extract_demand_values(ws, demand_row, week_cols):
    """Extract demand values for valid weeks."""
    demand_values = []
    for week_num in sorted(week_cols.keys()):
        col = week_cols[week_num]
        val = ws.cell(row=demand_row, column=col).value
        if isinstance(val, (int, float)) and val > 0:
            demand_values.append((week_num, col, float(val)))
    return demand_values


def step_down_simulation(initial_queue: float, demand_values: List[Tuple], params: Dict) -> Dict:
    """
    Run step-down simulation: 6 → 5 → 4 days per week.

    Returns: queue trajectory, milestones, final state.
    """
    step_days = [6, 5, 4]
    threshold = params.get('threshold', 4)
    daily_capacity = params['daily_capacity']

    current_step = 0
    queue = initial_queue
    trajectory = []
    milestones = {}

    for week_num, col, demand in demand_values:
        days_this_week = step_days[current_step]
        production = days_this_week * daily_capacity
        queue = queue + demand - production

        trajectory.append({
            'week': week_num,
            'demand': demand,
            'production': production,
            'days': days_this_week,
            'queue_end': queue
        })

        # Check step-down trigger
        if queue < threshold and current_step < len(step_days) - 1:
            current_step += 1

        # Track milestones
        if queue <= 0 and 'cleared' not in milestones:
            milestones['cleared'] = week_num

    return {
        'trajectory': trajectory,
        'milestones': milestones,
        'final_queue': queue,
        'final_step': current_step
    }


def generate_output_excel(output_path: str, simulation: Dict, params: Dict, demand_values: List) -> None:
    """Generate output Excel with simulation results."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = params.get('domain', 'Simulation')

    # Headers
    ws['A1'] = 'Week'
    ws['B1'] = 'Demand'
    ws['C1'] = 'Production'
    ws['D1'] = 'Days'
    ws['E1'] = 'Queue End'

    # Data rows
    for i, entry in enumerate(simulation['trajectory'], start=2):
        ws.cell(row=i, column=1, value=entry['week'])
        ws.cell(row=i, column=2, value=entry['demand'])
        ws.cell(row=i, column=3, value=entry['production'])
        ws.cell(row=i, column=4, value=entry['days'])
        ws.cell(row=i, column=5, value=entry['queue_end'])

    wb.save(output_path)


def generate_summary(output_path: str, simulation: Dict, params: Dict) -> None:
    """Generate summary.txt with milestone and narrative."""
    domain = params.get('domain', 'Queue Recovery')

    lines = [
        f"Domain: {domain}",
        f"Initial Queue: {params['initial_queue']}",
        f"Daily Capacity: {params['daily_capacity']} hours/day",
        f"Step-Down Threshold: {params.get('threshold', 4)}",
        ""
    ]

    if 'cleared' in simulation['milestones']:
        lines.append(f"Milestone: Queue cleared at week {simulation['milestones']['cleared']}")
    else:
        lines.append(f"Final Queue: {simulation['final_queue']} (not cleared)")

    lines.append("")
    lines.append(f"Final Step-Down Level: {['6 days', '5 days', '4 days'][simulation['final_step']]}")

    with open(output_path, 'w') as f:
        f.write('\n'.join(lines))


def main():
    parser = argparse.ArgumentParser(description='Weekly catch-up simulation')
    parser.add_argument('--input', required=True, help='Input Excel file')
    parser.add_argument('--output-xlsx', required=True, help='Output Excel file')
    parser.add_argument('--output-summary', required=True, help='Output summary.txt file')
    parser.add_argument('--params', required=True, help='Parameters JSON file')

    args = parser.parse_args()

    with open(args.params) as f:
        params = json.load(f)

    wb = openpyxl.load_workbook(args.input)
    ws = wb.active

    # Find demand row by label
    demand_row = find_demand_row(ws)
    if demand_row is None:
        raise ValueError("Could not find demand row in input file")

    # Find header row (typically above demand row)
    header_row = demand_row - 1

    # Find week columns from header
    week_cols = find_week_columns(ws, header_row, demand_row)

    # Extract demand values
    demand_values = extract_demand_values(ws, demand_row, week_cols)

    # Run simulation
    initial_queue = params['initial_queue']
    simulation = step_down_simulation(initial_queue, demand_values, params)

    # Generate outputs
    generate_output_excel(args.output_xlsx, simulation, params, demand_values)
    generate_summary(args.output_summary, simulation, params)

    print(f"Simulation complete. Final queue: {simulation['final_queue']}")
    if 'cleared' in simulation['milestones']:
        print(f"Queue cleared at week {simulation['milestones']['cleared']}")


if __name__ == '__main__':
    main()