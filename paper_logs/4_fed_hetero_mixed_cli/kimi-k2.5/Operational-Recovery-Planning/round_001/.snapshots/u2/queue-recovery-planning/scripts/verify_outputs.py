#!/usr/bin/env python3
"""Verify queue recovery planning outputs against standard constraints."""
import sys
import re
import openpyxl

def verify_summary(path):
    """Check summary.txt format and constraints."""
    with open(path) as f:
        lines = f.readlines()

    errors = []

    # Must have exactly 3 lines
    if len(lines) != 3:
        errors.append(f"Expected 3 lines, got {len(lines)}")
    else:
        # Line 1: First_Week_5_Days: <int>
        m1 = re.match(r'^First_Week_5_Days:\s*(\d+)\s*$', lines[0])
        if not m1:
            errors.append(f"Line 1 format error: expected 'First_Week_5_Days: <int>'")

        # Line 2: First_Week_4_Days: <int>
        m2 = re.match(r'^First_Week_4_Days:\s*(\d+)\s*$', lines[1])
        if not m2:
            errors.append(f"Line 2 format error: expected 'First_Week_4_Days: <int>'")

        # Line 3: Summary text
        summary = lines[2].strip()
        if not summary.startswith('Summary:'):
            errors.append(f"Line 3 must start with 'Summary:'")
        else:
            # Extract summary text after "Summary: "
            summary_text = summary[8:].strip()

            # Word count: ≤31 words
            words = summary_text.split()
            if len(words) > 31:
                errors.append(f"Summary too long: {len(words)} words (max 31)")

            # Sentence count: exactly 2
            # Count sentences by '. ' pattern (period followed by space or end)
            sentences = [s.strip() for s in re.split(r'\.\s*', summary_text) if s.strip()]
            if len(sentences) != 2:
                errors.append(f"Expected 2 sentences, got {len(sentences)}")

    if errors:
        print("Summary validation FAILED:")
        for e in errors:
            print(f"  - {e}")
        return False
    else:
        print("Summary: OK")
        return True

def verify_workbook(path, sheet_name="Plan", expected_rows=40):
    """Check plan.xlsx format and constraints."""
    errors = []

    try:
        wb = openpyxl.load_workbook(path)
    except Exception as e:
        print(f"Workbook validation FAILED: Cannot load file: {e}")
        return False

    if sheet_name not in wb.sheetnames:
        errors.append(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
        print("Workbook validation FAILED:")
        for e in errors:
            print(f"  - {e}")
        return False

    ws = wb[sheet_name]

    # Row count: header + 40 data rows
    if ws.max_row != expected_rows + 1:
        errors.append(f"Expected {expected_rows + 1} rows, got {ws.max_row}")

    # Column count: at least 7
    if ws.max_column < 7:
        errors.append(f"Expected at least 7 columns, got {ws.max_column}")
    else:
        # Verify header names
        expected_headers = [
            "Week",
            "On-Call Days",
            "Forecast Alert Load (Analyst Hrs)",
            "Weekly Triage Capacity (Analyst Hrs)",
            "Start-of-Week Alert Queue (Analyst Hrs)",
            "End-of-Week Alert Queue/Buffer (Analyst Hrs)",
            "Burnout Overtime Hours"
        ]
        actual_headers = [ws.cell(row=1, column=c).value for c in range(1, 8)]
        for i, (exp, act) in enumerate(zip(expected_headers, actual_headers), 1):
            if exp != act:
                errors.append(f"Column {i} header mismatch: expected '{exp}', got '{act}'")

    # Check weeks are 1..N ascending
    weeks = [ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)]
    if weeks != list(range(1, expected_rows + 1)):
        errors.append(f"Weeks not ascending 1..{expected_rows}")

    # Check for floating-point artifacts in numeric cells (columns 3-6)
    artifact_pattern = re.compile(r'\.\d{3,}$')
    for r in range(2, ws.max_row + 1):
        for c in range(3, 7):  # Columns 3-6 are numeric
            val = ws.cell(row=r, column=c).value
            if isinstance(val, float):
                # Check if more than 2 decimal places
                val_str = f"{val:.10f}".rstrip('0')
                if '.' in val_str:
                    decimals = len(val_str.split('.')[1])
                    if decimals > 2:
                        errors.append(f"Row {r}, Col {c}: floating-point artifact ({val})")

    if errors:
        print("Workbook validation FAILED:")
        for e in errors:
            print(f"  - {e}")
        return False
    else:
        print("Workbook: OK")
        return True

def main():
    if len(sys.argv) != 3:
        print("Usage: verify_outputs.py <summary.txt> <workbook.xlsx>")
        sys.exit(1)

    summary_ok = verify_summary(sys.argv[1])
    workbook_ok = verify_workbook(sys.argv[2])

    if summary_ok and workbook_ok:
        print("\nAll validations passed.")
        sys.exit(0)
    else:
        print("\nValidation FAILED. Fix the issues above.")
        sys.exit(1)

if __name__ == "__main__":
    main()
